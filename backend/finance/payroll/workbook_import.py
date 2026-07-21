"""Import daily payroll points from the legacy schedule workbook."""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel

from finance.enums import ApprovalStatus, PayrollPeriodStatus, PayrollRunStatus
from finance.models import (
    PayrollAuditEvent,
    PayrollDailyWorkEntry,
    PayrollPeriod,
    PayrollWorkSettings,
    PayrollWorkRecord,
)

from .access import has_payroll_permission
from .exceptions import PayrollOperationError, PayrollPermissionDenied
from .services import sync_daily_work_record

Employee = get_user_model()

MAX_WORKBOOK_SIZE = 10 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 80 * 1024 * 1024
IMPORT_MODES = {"skip_existing", "replace_existing"}
QUANTUM = Decimal("0.0001")
MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def _error(code: str, message: str, **details):
    raise PayrollOperationError(code, message, details=details)


def _normalize_name(value: str) -> str:
    value = value.casefold().replace("ё", "е")
    value = re.sub(r"[^a-zа-я0-9]+", " ", value)
    return " ".join(value.split())


def _employee_display_name(employee) -> str:
    parts = [employee.last_name, employee.first_name, employee.patronymic]
    return (
        " ".join(part.strip() for part in parts if part and part.strip())
        or f"Сотрудник #{employee.pk}"
    )


def _employee_variants(employee) -> set[str]:
    first = employee.first_name or ""
    last = employee.last_name or ""
    patronymic = employee.patronymic or ""
    candidates = {
        _employee_display_name(employee),
        employee.get_full_name(),
        f"{first} {last}",
        f"{last} {first}",
        first,
        last,
    }
    if patronymic:
        candidates.update(
            {
                f"{last} {first} {patronymic}",
                f"{first} {patronymic} {last}",
            }
        )
    aliases = (
        employee.attendance_aliases
        if isinstance(employee.attendance_aliases, list)
        else []
    )
    candidates.update(str(item) for item in aliases)
    return {
        normalized for item in candidates if (normalized := _normalize_name(str(item)))
    }


def _employee_payload(employee) -> dict:
    position = getattr(employee, "position", None)
    return {
        "id": employee.pk,
        "display_name": _employee_display_name(employee),
        "position": position.name if position is not None else None,
        "is_active": employee.is_active,
    }


def _sanitize_styles(payload: bytes) -> bytes:
    """Repair invalid third-party colour values without touching the source file."""

    source = io.BytesIO(payload)
    target = io.BytesIO()
    try:
        with zipfile.ZipFile(source, "r") as archive:
            total = sum(item.file_size for item in archive.infolist())
            if total > MAX_UNCOMPRESSED_SIZE:
                _error(
                    "WORKBOOK_TOO_LARGE", "Распакованный Excel-файл слишком большой."
                )
            with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as repaired:
                for item in archive.infolist():
                    content = archive.read(item.filename)
                    if item.filename == "xl/styles.xml":
                        content = re.sub(
                            rb'rgb="(?![0-9A-Fa-f]{8}")[^"]*"',
                            b'rgb="00000000"',
                            content,
                        )
                        # Several generators write the unsupported border value
                        # ``style="none"`` or CSS-like ``solid``.
                        content = content.replace(b' style="none"', b"")
                        content = content.replace(b'style="solid"', b'style="thin"')
                    repaired.writestr(item, content)
    except zipfile.BadZipFile:
        _error("INVALID_WORKBOOK", "Файл не является корректной книгой Excel.")
    return target.getvalue()


def _load_workbook(payload: bytes):
    try:
        return load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    except (TypeError, ValueError):
        try:
            repaired = _sanitize_styles(payload)
            return load_workbook(io.BytesIO(repaired), data_only=False, read_only=False)
        except PayrollOperationError:
            raise
        except Exception as exc:  # openpyxl exposes several format exceptions
            _error(
                "INVALID_WORKBOOK",
                "Не удалось прочитать структуру Excel-файла.",
                error=str(exc),
            )
    except (OSError, zipfile.BadZipFile) as exc:
        _error("INVALID_WORKBOOK", "Не удалось открыть Excel-файл.", error=str(exc))


def _cell_day(value) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int) and 1 <= value <= 31:
        return value
    if isinstance(value, float) and value.is_integer() and 1 <= value <= 31:
        return int(value)
    return None


def _month_number(value) -> int | None:
    if not isinstance(value, str):
        return None
    return MONTHS.get(_normalize_name(value))


def _block_year(
    sheet_title: str, worksheet, day_row: int, month_row: int
) -> tuple[int | None, int]:
    for row in range(day_row, month_row + 1):
        value = worksheet.cell(row, 1).value
        if (
            isinstance(value, (int, float))
            and int(value) == value
            and 2000 <= int(value) <= 2200
        ):
            return int(value), 2
    match = re.search(r"(20\d{2})", sheet_title)
    return (int(match.group(1)), 1) if match else (None, 0)


def _point_value(value) -> tuple[Decimal | None, str | None]:
    if value is None or isinstance(value, bool):
        return None, None
    if isinstance(value, str):
        stripped = value.replace("\xa0", " ").strip()
        if not stripped or stripped.startswith("="):
            return None, None
        try:
            number = Decimal(stripped.replace(",", "."))
        except InvalidOperation:
            return None, None
    elif isinstance(value, (datetime, date, time)):
        # Some cells containing 5.5 points are incorrectly formatted as dates.
        number = Decimal(str(to_excel(value)))
    else:
        try:
            number = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return None, None
    if not number.is_finite():
        return None, "Значение не является конечным числом"
    if number < 0:
        return None, "Отрицательные баллы не поддерживаются"
    return number.quantize(QUANTUM), None


def _find_period_blocks(workbook, period: PayrollPeriod) -> list[dict]:
    wanted_months: set[tuple[int, int]] = set()
    cursor = date(period.date_from.year, period.date_from.month, 1)
    end = date(period.date_to.year, period.date_to.month, 1)
    while cursor <= end:
        wanted_months.add((cursor.year, cursor.month))
        cursor = date(cursor.year + (cursor.month == 12), cursor.month % 12 + 1, 1)

    candidates: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for worksheet in workbook.worksheets:
        for month_row in range(1, worksheet.max_row + 1):
            month = _month_number(worksheet.cell(month_row, 1).value)
            if month is None:
                continue
            day_row = None
            day_columns = None
            for possible in range(month_row - 1, max(0, month_row - 4), -1):
                columns = {
                    column: day
                    for column in range(2, worksheet.max_column + 1)
                    if (day := _cell_day(worksheet.cell(possible, column).value))
                    is not None
                }
                if len(columns) >= 7:
                    day_row, day_columns = possible, columns
                    break
            if day_row is None or day_columns is None:
                continue
            year, confidence = _block_year(
                worksheet.title, worksheet, day_row, month_row
            )
            if year is None or (year, month) not in wanted_months:
                continue
            candidates[(year, month)].append(
                {
                    "worksheet": worksheet,
                    "sheet_name": worksheet.title,
                    "month_row": month_row,
                    "day_row": day_row,
                    "day_columns": day_columns,
                    "year": year,
                    "month": month,
                    "confidence": confidence,
                }
            )
    blocks = []
    missing = []
    for year_month in sorted(wanted_months):
        options = candidates.get(year_month, [])
        if not options:
            missing.append(f"{year_month[1]:02d}.{year_month[0]}")
            continue
        blocks.append(
            max(
                options, key=lambda item: (item["confidence"], len(item["day_columns"]))
            )
        )
    if missing:
        _error(
            "PERIOD_NOT_FOUND_IN_WORKBOOK",
            "В файле не найден график для выбранного периода.",
            missing_months=missing,
        )
    return blocks


def parse_workbook(payload: bytes, period: PayrollPeriod) -> dict:
    if not payload:
        _error("EMPTY_WORKBOOK", "Выберите непустой Excel-файл.")
    if len(payload) > MAX_WORKBOOK_SIZE:
        _error("WORKBOOK_TOO_LARGE", "Excel-файл должен быть не больше 10 МБ.")
    workbook = _load_workbook(payload)
    blocks = _find_period_blocks(workbook, period)
    rows: list[dict] = []
    for block in blocks:
        worksheet = block["worksheet"]
        later_month_rows = [
            row
            for row in range(block["month_row"] + 1, worksheet.max_row + 1)
            if _month_number(worksheet.cell(row, 1).value) is not None
        ]
        end_row = (min(later_month_rows) - 1) if later_month_rows else worksheet.max_row
        for row_number in range(block["month_row"] + 1, end_row + 1):
            raw_name = worksheet.cell(row_number, 1).value
            if (
                not isinstance(raw_name, str)
                or not raw_name.replace("\xa0", " ").strip()
            ):
                continue
            source_name = " ".join(raw_name.replace("\xa0", " ").split())
            entries = []
            invalid = []
            for column, day in block["day_columns"].items():
                try:
                    work_date = date(block["year"], block["month"], day)
                except ValueError:
                    continue
                if not period.date_from <= work_date <= period.date_to:
                    continue
                points, issue = _point_value(worksheet.cell(row_number, column).value)
                if issue:
                    invalid.append({"date": work_date.isoformat(), "message": issue})
                elif points is not None:
                    entries.append({"date": work_date, "points": points})
            if not entries and not invalid:
                continue
            rows.append(
                {
                    "row_key": f"{block['sheet_name']}:{row_number}",
                    "sheet_name": block["sheet_name"],
                    "row_number": row_number,
                    "source_name": source_name,
                    "entries": entries,
                    "invalid_cells": invalid,
                }
            )
    if not rows:
        _error(
            "NO_WORK_POINTS_FOUND",
            "В графике за выбранный период не найдено числовых значений выработки.",
        )
    return {
        "blocks": [
            {
                "sheet_name": item["sheet_name"],
                "year": item["year"],
                "month": item["month"],
            }
            for item in blocks
        ],
        "rows": rows,
    }


def _employees_with_variants():
    employees = list(
        Employee.objects.select_related("position").order_by(
            "last_name", "first_name", "id"
        )
    )
    return employees, {
        employee.pk: _employee_variants(employee) for employee in employees
    }


def _resolve_name(
    source_name: str, employees, variants_by_id
) -> tuple[int | None, list[int], str]:
    normalized = _normalize_name(source_name)
    exact = [
        employee.pk
        for employee in employees
        if normalized in variants_by_id[employee.pk]
    ]
    if len(exact) == 1:
        return exact[0], exact, "matched"
    if len(exact) > 1:
        return None, exact, "ambiguous"
    scored = []
    source_tokens = set(normalized.split())
    for employee in employees:
        variants = variants_by_id[employee.pk]
        ratio = max(
            (SequenceMatcher(None, normalized, item).ratio() for item in variants),
            default=0,
        )
        if source_tokens and any(
            source_tokens & set(item.split()) for item in variants
        ):
            ratio = max(ratio, 0.61)
        if ratio >= 0.5:
            scored.append((ratio, employee.pk))
    candidates = [employee_id for _, employee_id in sorted(scored, reverse=True)[:7]]
    return None, candidates, "unmatched"


def build_workbook_preview(
    period: PayrollPeriod, *, payload: bytes, filename: str
) -> dict:
    parsed = parse_workbook(payload, period)
    employees, variants_by_id = _employees_with_variants()
    existing = set(
        PayrollDailyWorkEntry.objects.filter(period=period).values_list(
            "employee_id", "work_date"
        )
    )
    employees_with_daily_entries = {employee_id for employee_id, _ in existing}
    employees_with_monthly_records = set(
        PayrollWorkRecord.objects.filter(period=period)
        .exclude(status=ApprovalStatus.VOIDED)
        .values_list("employee_id", flat=True)
    )
    public_rows = []
    matched_count = 0
    conflict_count = 0
    for row in parsed["rows"]:
        employee_id, candidates, status = _resolve_name(
            row["source_name"], employees, variants_by_id
        )
        if employee_id is not None:
            matched_count += 1
        date_conflicts = sum(
            (employee_id, item["date"]) in existing
            for item in row["entries"]
            if employee_id
        )
        monthly_only_conflict = bool(
            employee_id
            and employee_id in employees_with_monthly_records
            and employee_id not in employees_with_daily_entries
        )
        row_conflicts = len(row["entries"]) if monthly_only_conflict else date_conflicts
        conflict_count += row_conflicts
        public_rows.append(
            {
                "row_key": row["row_key"],
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "source_name": row["source_name"],
                "match_status": status,
                "matched_employee_id": employee_id,
                "candidate_employee_ids": candidates,
                "entry_count": len(row["entries"]),
                "points_total": str(
                    sum(
                        (item["points"] for item in row["entries"]), Decimal("0")
                    ).quantize(QUANTUM)
                ),
                "existing_count": row_conflicts,
                "existing_period_record": monthly_only_conflict,
                "invalid_cells": row["invalid_cells"],
                "entries": [
                    {"date": item["date"].isoformat(), "points": str(item["points"])}
                    for item in row["entries"]
                ],
            }
        )
    return {
        "file_name": Path(filename).name,
        "file_hash": hashlib.sha256(payload).hexdigest(),
        "period_id": period.pk,
        "period_lock_version": period.lock_version,
        "blocks": parsed["blocks"],
        "employees": [_employee_payload(employee) for employee in employees],
        "rows": public_rows,
        "summary": {
            "rows": len(public_rows),
            "matched": matched_count,
            "needs_mapping": len(public_rows) - matched_count,
            "entries": sum(item["entry_count"] for item in public_rows),
            "existing": conflict_count,
            "invalid": sum(len(item["invalid_cells"]) for item in public_rows),
        },
    }


def _ensure_import_allowed(period: PayrollPeriod):
    if period.status == PayrollPeriodStatus.CLOSED:
        _error("PERIOD_INPUTS_LOCKED", "Данные закрытого периода заблокированы.")
    if period.current_run is not None and period.current_run.status in {
        PayrollRunStatus.CALCULATED,
        PayrollRunStatus.REVIEW,
        PayrollRunStatus.APPROVED,
    }:
        _error("PERIOD_INPUTS_LOCKED", "Сначала верните текущий расчёт на исправление.")


@transaction.atomic
def apply_workbook_import(
    period_id: int,
    *,
    actor,
    payload: bytes,
    filename: str,
    mappings: dict[str, int | None],
    mode: str,
    expected_file_hash: str,
    expected_period_lock_version: int,
) -> dict:
    if (
        not actor
        or not actor.is_authenticated
        or not has_payroll_permission(actor, "finance.manage_payroll_inputs")
    ):
        raise PayrollPermissionDenied("finance.manage_payroll_inputs")
    if mode not in IMPORT_MODES:
        _error(
            "INVALID_WORKBOOK_IMPORT_MODE",
            "Неизвестный режим обработки существующих записей.",
        )
    actual_hash = hashlib.sha256(payload).hexdigest()
    if actual_hash != expected_file_hash:
        _error(
            "WORKBOOK_CHANGED",
            "После предпросмотра выбран другой файл. Повторите проверку.",
        )
    period = (
        PayrollPeriod.objects.select_for_update()
        .select_related("current_run")
        .get(pk=period_id)
    )
    _ensure_import_allowed(period)
    if period.lock_version != expected_period_lock_version:
        _error("STALE_PERIOD", "Период уже изменён; повторите предпросмотр.")

    parsed = parse_workbook(payload, period)
    employees, variants_by_id = _employees_with_variants()
    employees_by_id = {employee.pk: employee for employee in employees}
    resolved_rows = []
    unmapped_entry_count = 0
    for row in parsed["rows"]:
        automatic_id, _, _ = _resolve_name(
            row["source_name"], employees, variants_by_id
        )
        employee_id = mappings.get(row["row_key"], automatic_id)
        if employee_id is None:
            unmapped_entry_count += len(row["entries"])
            continue
        employee = employees_by_id.get(int(employee_id))
        if employee is None:
            _error(
                "EMPLOYEE_NOT_FOUND",
                "Выбранный сотрудник не найден.",
                employee_id=employee_id,
            )
        resolved_rows.append((row, employee))

    imported_keys: set[tuple[int, date]] = set()
    for row, employee in resolved_rows:
        for item in row["entries"]:
            key = (employee.pk, item["date"])
            if key in imported_keys:
                _error(
                    "DUPLICATE_WORKBOOK_MAPPING",
                    "Две строки файла сопоставлены одному сотруднику и одной дате.",
                    employee_id=employee.pk,
                    work_date=item["date"].isoformat(),
                )
            imported_keys.add(key)

    existing_by_key = {
        (entry.employee_id, entry.work_date): entry
        for entry in PayrollDailyWorkEntry.objects.select_for_update().filter(
            period=period,
            employee_id__in={employee.pk for _, employee in resolved_rows},
        )
    }
    employees_with_daily_entries = {employee_id for employee_id, _ in existing_by_key}
    employees_with_monthly_records = set(
        PayrollWorkRecord.objects.select_for_update()
        .filter(
            period=period,
            employee_id__in={employee.pk for _, employee in resolved_rows},
        )
        .exclude(status=ApprovalStatus.VOIDED)
        .values_list("employee_id", flat=True)
    )
    monthly_only_conflicts = (
        employees_with_monthly_records - employees_with_daily_entries
    )
    daily_target = PayrollWorkSettings.get_daily_target_points()
    summary = {
        "created": 0,
        "replaced": 0,
        "unchanged": 0,
        "skipped": unmapped_entry_count,
    }
    affected_employee_ids = set()
    safe_filename = Path(filename).name
    for row, employee in resolved_rows:
        for item in row["entries"]:
            if mode == "skip_existing" and employee.pk in monthly_only_conflicts:
                summary["skipped"] += 1
                continue
            key = (employee.pk, item["date"])
            entry = existing_by_key.get(key)
            if entry is not None and mode == "skip_existing":
                summary["skipped"] += 1
                continue
            note = f"Импорт из {safe_filename}; строка «{row['source_name']}»"
            if entry is not None:
                if entry.actual_points == item["points"]:
                    summary["unchanged"] += 1
                    continue
                previous = str(entry.actual_points)
                entry.actual_points = item["points"]
                entry.note = note
                entry.lock_version += 1
                action = "payroll.daily_work_entry_import_replaced"
                metadata = {"previous_actual_points": previous}
                summary["replaced"] += 1
            else:
                entry = PayrollDailyWorkEntry(
                    period=period,
                    employee=employee,
                    work_date=item["date"],
                    target_points=daily_target,
                    actual_points=item["points"],
                    note=note,
                )
                action = "payroll.daily_work_entry_import_created"
                metadata = {}
                summary["created"] += 1
            try:
                entry.full_clean()
                entry.save()
            except ValidationError as exc:
                _error(
                    "INPUT_VALIDATION_FAILED",
                    "Импортируемая выработка не прошла проверку.",
                    errors=exc.message_dict,
                )
            affected_employee_ids.add(employee.pk)
            PayrollAuditEvent.objects.create(
                actor=actor,
                action=action,
                object_type=entry._meta.label_lower,
                object_id=str(entry.pk),
                period=period,
                metadata={
                    "channel": "workbook_import",
                    "file_hash": actual_hash,
                    "file_name": safe_filename,
                    "source_name": row["source_name"],
                    "work_date": item["date"].isoformat(),
                    "actual_points": str(item["points"]),
                    **metadata,
                },
            )

    records = []
    for employee_id in sorted(affected_employee_ids):
        records.append(
            sync_daily_work_record(
                period=period,
                employee=employees_by_id[employee_id],
                actor=actor,
                channel="workbook_import",
            )
        )
    return {
        "mode": mode,
        "summary": summary,
        "record_ids": [item.pk for item in records],
    }


def parse_mappings(value) -> dict[str, int | None]:
    if isinstance(value, dict):
        raw = value
    else:
        try:
            raw = json.loads(value or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            _error(
                "INVALID_EMPLOYEE_MAPPINGS",
                "Не удалось прочитать сопоставление сотрудников.",
            )
    if not isinstance(raw, dict):
        _error(
            "INVALID_EMPLOYEE_MAPPINGS",
            "Сопоставление сотрудников должно быть объектом.",
        )
    parsed = {}
    for key, employee_id in raw.items():
        if not isinstance(key, str) or employee_id is None:
            parsed[str(key)] = None
            continue
        try:
            parsed[key] = int(employee_id)
        except (TypeError, ValueError):
            _error(
                "INVALID_EMPLOYEE_MAPPINGS",
                "Некорректный идентификатор сотрудника.",
                row_key=key,
            )
    return parsed
