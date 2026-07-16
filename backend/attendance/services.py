from __future__ import annotations

import logging
from io import BytesIO
from datetime import date, timedelta
from calendar import monthrange
from typing import Any

from django.conf import settings as django_settings
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from attendance.models import (
    AttendanceAutoSyncSettings,
    AttendanceAnalysisRun,
    AttendanceRecord,
    EmployeeWorkSchedule,
    StandardWorkSchedule,
)
from employees.constants import (
    ACTION_DISMISSED,
    ACTION_ON_DAY_OFF,
    ACTION_ON_LEAVE,
    ACTION_ON_MATERNITY,
    ACTION_ON_SICK_LEAVE,
    ACTION_REMOTE,
)
from employees.models import Employee
from employees.services.personnel_state import (
    EmployeePersonnelState,
    PERSONNEL_STATUS_NORMAL,
    resolve_employee_personnel_state,
)

MONTHS_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

DAYS_RU = {
    0: "Пн",
    1: "Вт",
    2: "Ср",
    3: "Чт",
    4: "Пт",
    5: "Сб",
    6: "Вс",
}

MATRIX_STATUS_LABELS = {
    "empty": "Нет записи",
    "technical": "Техсбой",
    "personnel_issue": "Кадровое событие",
    "underwork": "Недоработка",
    "late": "Опоздание",
    "overtime": "Переработка",
    "absent": "Отсутствие",
    "non_working": "Нерабочий",
    "normal": "Норма",
}

MATRIX_STATUS_SHORT_LABELS = {
    "empty": "",
    "technical": "Тех",
    "personnel_issue": "Кадр",
    "underwork": "НД",
    "late": "ОП",
    "overtime": "ПР",
    "absent": "Отс.",
    "non_working": "Нер",
    "normal": "OK",
}

logger = logging.getLogger(__name__)

PERSONNEL_STATUS_SHORT_LABELS = {
    ACTION_ON_LEAVE: "ОТП",
    ACTION_ON_SICK_LEAVE: "БЛ",
    ACTION_ON_DAY_OFF: "ОТГ",
    ACTION_ON_MATERNITY: "ДЕКР",
    ACTION_DISMISSED: "Вне штата",
    ACTION_REMOTE: "УД",
}

PERSONNEL_STATUS_LEGEND = {
    ACTION_ON_LEAVE: "Отпуск",
    ACTION_ON_SICK_LEAVE: "Больничный",
    ACTION_ON_DAY_OFF: "Отгул",
    ACTION_ON_MATERNITY: "Декрет",
    ACTION_DISMISSED: "Уволен / вне штата",
    ACTION_REMOTE: "Удаленка",
}

MATRIX_STATUS_FILLS = {
    "empty": "F1F5F9",
    "technical": "FEE2E2",
    "personnel_issue": "FEF3C7",
    "underwork": "FEF3C7",
    "late": "FEF3C7",
    "absent": "FEF3C7",
    "overtime": "DCFCE7",
    "non_working": "E2E8F0",
    "normal": "E0F2FE",
}

SUPPRESSED_EMPLOYEE_ISSUE_MARKERS = (
    "absence",
    "absent",
    "late",
    "early",
    "underwork",
    "отсутств",
    "опозд",
    "ранн",
    "недоработ",
)

ABSENCE_ISSUE_MARKERS = ("absence", "absent", "отсутств")

DEFAULT_WORK_SCHEDULE_PAYLOAD = {
    "start_time": "08:00",
    "end_time": "17:00",
    "expected_hours": 9,
    "workdays": list(EmployeeWorkSchedule.DEFAULT_WORKDAYS),
    "date_overrides": [],
}


def save_logstorm_attendance_result(
    *,
    employee: Employee,
    period_start: date,
    period_end: date,
    request_payload: dict[str, Any],
    response_payload: dict[str, Any],
    schedule_payload: dict[str, Any] | None = None,
    triggered_by: Employee | None = None,
) -> AttendanceAnalysisRun:
    records = response_payload.get("records")
    if not isinstance(records, list):
        records = []

    personnel_actions = list(
        employee.actions.filter(date__date__lte=period_end).order_by("date", "id")
    )

    with transaction.atomic():
        run = AttendanceAnalysisRun.objects.create(
            employee=employee,
            period_start=period_start,
            period_end=period_end,
            status=AttendanceAnalysisRun.STATUS_SUCCESS,
            schedule_payload=schedule_payload,
            request_payload=request_payload,
            response_payload=response_payload,
            triggered_by=triggered_by
            if getattr(triggered_by, "is_authenticated", False)
            else None,
        )

        for raw_record in records:
            if not isinstance(raw_record, dict):
                continue

            record_date = _parse_record_date(raw_record.get("date"))
            if record_date is None:
                continue
            personnel_state = resolve_employee_day_state(
                personnel_actions,
                record_date,
            )
            record_defaults = build_attendance_record_defaults(
                raw_record,
                personnel_state,
            )
            existing_record = AttendanceRecord.objects.filter(
                employee=employee,
                date=record_date,
            ).first()
            if (
                existing_record
                and existing_record.is_manually_edited
                and not raw_record.get("manual_edited")
            ):
                record_defaults = apply_manual_edit_payload_to_defaults(
                    record_defaults,
                    existing_record.manual_edit_payload,
                )
                record_defaults.update(
                    {
                        "is_manually_edited": True,
                        "manual_edit_payload": existing_record.manual_edit_payload,
                        "manual_edited_by": existing_record.manual_edited_by,
                        "manual_edited_at": existing_record.manual_edited_at,
                    }
                )

            AttendanceRecord.objects.update_or_create(
                employee=employee,
                date=record_date,
                defaults={"analysis_run": run, **record_defaults},
            )

    return run


def get_employee_work_schedule_payload(
    employee: Employee,
) -> dict[str, Any] | None:
    try:
        schedule = employee.work_schedule
    except EmployeeWorkSchedule.DoesNotExist:
        return None
    if schedule is None or not schedule.is_active:
        return None
    return schedule.to_logstorm_payload()


def get_default_work_schedule_payload() -> dict[str, Any]:
    return {
        **DEFAULT_WORK_SCHEDULE_PAYLOAD,
        "workdays": list(DEFAULT_WORK_SCHEDULE_PAYLOAD["workdays"]),
        "date_overrides": list(DEFAULT_WORK_SCHEDULE_PAYLOAD["date_overrides"]),
    }


def get_standard_work_schedule() -> StandardWorkSchedule | None:
    return StandardWorkSchedule.objects.order_by("id").first()


def get_standard_work_schedule_payload() -> dict[str, Any]:
    schedule = get_standard_work_schedule()
    if schedule is None:
        return get_default_work_schedule_payload()
    return schedule.to_logstorm_payload()


def get_attendance_auto_sync_settings() -> AttendanceAutoSyncSettings:
    settings, _ = AttendanceAutoSyncSettings.objects.get_or_create(
        singleton=True,
    )
    return settings


def get_attendance_auto_sync_employees():
    return Employee.objects.prefetch_related("actions").order_by("-id")


def get_attendance_auto_sync_stale_after() -> timedelta:
    configured_seconds = getattr(
        django_settings,
        "ATTENDANCE_AUTO_SYNC_STALE_AFTER_SECONDS",
        None,
    )
    if configured_seconds is None:
        configured_seconds = (
            getattr(django_settings, "CELERY_TASK_TIME_LIMIT", 30 * 60) + 300
        )
    try:
        seconds = int(configured_seconds)
    except (TypeError, ValueError):
        seconds = 30 * 60 + 300
    return timedelta(seconds=max(seconds, 60))


def is_attendance_auto_sync_stale(
    settings_obj: AttendanceAutoSyncSettings,
    now,
) -> bool:
    if settings_obj.last_status != AttendanceAutoSyncSettings.STATUS_RUNNING:
        return False
    if not settings_obj.last_started_at:
        return True
    return now - settings_obj.last_started_at > get_attendance_auto_sync_stale_after()


def run_attendance_auto_sync(*, force: bool = False) -> AttendanceAutoSyncSettings:
    from common.logstorm_attendance import (
        analyze_employee_attendance,
        build_logstorm_attendance_payload,
    )
    now = timezone.now()
    with transaction.atomic():
        recovered_stale_running = False
        settings = (
            AttendanceAutoSyncSettings.objects.select_for_update()
            .filter(singleton=True)
            .first()
        )
        if settings is None:
            settings = AttendanceAutoSyncSettings.objects.create(singleton=True)

        if settings.last_status == AttendanceAutoSyncSettings.STATUS_RUNNING:
            if not is_attendance_auto_sync_stale(settings, now):
                return settings

            recovered_stale_running = True
            stale_started_at = settings.last_started_at
            stale_message = (
                "Recovered stale running attendance auto-sync"
                f" started at {stale_started_at.isoformat()}"
                if stale_started_at
                else "Recovered stale running attendance auto-sync without start time"
            )
            logger.warning(stale_message)
            settings.last_status = AttendanceAutoSyncSettings.STATUS_FAILED
            settings.last_finished_at = now
            settings.last_error = stale_message
            settings.last_error_count = 1
            settings.next_run_at = now if settings.enabled else None
            settings.save(
                update_fields=[
                    "last_status",
                    "last_finished_at",
                    "last_error",
                    "last_error_count",
                    "next_run_at",
                    "updated_at",
                ]
            )

        if not force:
            if not settings.enabled:
                return settings
            if (
                not recovered_stale_running
                and settings.next_run_at
                and settings.next_run_at > now
            ):
                return settings

        settings.last_status = AttendanceAutoSyncSettings.STATUS_RUNNING
        settings.last_started_at = now
        settings.last_finished_at = None
        settings.last_error = ""
        settings.last_success_count = 0
        settings.last_error_count = 0
        if settings.enabled:
            settings.next_run_at = now + timedelta(
                minutes=settings.frequency_minutes
            )
        settings.save(
            update_fields=[
                "last_status",
                "last_started_at",
                "last_finished_at",
                "last_error",
                "last_success_count",
                "last_error_count",
                "next_run_at",
                "updated_at",
            ]
        )
        settings_id = settings.id
        lookback_days = settings.lookback_days

    period_end = timezone.localdate()
    period_start = period_end - timedelta(days=lookback_days - 1)
    success_count = 0
    error_count = 0
    errors: list[str] = []

    for employee in get_attendance_auto_sync_employees():
        schedule_payload = (
            get_employee_work_schedule_payload(employee)
            or get_standard_work_schedule_payload()
        )
        try:
            aliases = list(getattr(employee, "attendance_aliases", None) or [])
            result = analyze_employee_attendance(
                employee=employee,
                period_start=period_start,
                period_end=period_end,
                schedule=schedule_payload,
                aliases=aliases,
                client=None,
            )
            request_payload = build_logstorm_attendance_payload(
                employee=employee,
                period_start=period_start,
                period_end=period_end,
                schedule=schedule_payload,
                aliases=aliases,
            )
            save_logstorm_attendance_result(
                employee=employee,
                period_start=period_start,
                period_end=period_end,
                schedule_payload=request_payload.get("schedule"),
                request_payload=request_payload,
                response_payload=result,
                triggered_by=None,
            )
            success_count += 1
        except Exception as exc:
            error_count += 1
            errors.append(f"{_employee_display_name(employee)}: {exc}")

    finished_at = timezone.now()
    if error_count and success_count:
        status = AttendanceAutoSyncSettings.STATUS_PARTIAL
    elif error_count:
        status = AttendanceAutoSyncSettings.STATUS_FAILED
    else:
        status = AttendanceAutoSyncSettings.STATUS_SUCCESS

    with transaction.atomic():
        settings = AttendanceAutoSyncSettings.objects.select_for_update().get(
            pk=settings_id,
        )
        settings.last_status = status
        settings.last_finished_at = finished_at
        settings.last_success_count = success_count
        settings.last_error_count = error_count
        settings.last_error = "\n".join(errors)[:4000]
        settings.next_run_at = (
            finished_at + timedelta(minutes=settings.frequency_minutes)
            if settings.enabled
            else None
        )
        settings.save(
            update_fields=[
                "last_status",
                "last_finished_at",
                "last_success_count",
                "last_error_count",
                "last_error",
                "next_run_at",
                "updated_at",
            ]
        )
        return settings


def build_monthly_attendance_matrix(
    *,
    employees,
    records,
    year: int,
    month: int,
) -> dict[str, Any]:
    """Build an Excel-like monthly attendance matrix for API/UI usage."""
    employees = list(employees)
    records_by_key = {
        (record.employee_id, record.date): record
        for record in records
    }
    days_in_month = monthrange(year, month)[1]
    dates = [date(year, month, day) for day in range(1, days_in_month + 1)]

    rows = []
    for current_date in dates:
        row = {
            "date": current_date.isoformat(),
            "label": f"{current_date.day} ({DAYS_RU[current_date.weekday()]})",
            "weekday": DAYS_RU[current_date.weekday()],
            "is_weekend": current_date.weekday() >= 5,
            "cells": {},
        }
        for employee in employees:
            record = records_by_key.get((employee.id, current_date))
            row["cells"][str(employee.id)] = _monthly_matrix_cell(record)
        rows.append(row)

    summary_rows = [
        ("accounted_days", "Учтено(дней)", _summary_accounted_days),
        ("unaccounted_days", "Неучтено(дней)", _summary_unaccounted_days),
        ("worked_hours", "Отработано(часов)", _summary_worked_hours),
        ("late_days", "Опозданий", _summary_late_days),
        ("early_leave_days", "Ранних уходов", _summary_early_leave_days),
        ("overtime_days", "Переработок", _summary_overtime_days),
        ("absent_days", "Отсутствий", _summary_absent_days),
    ]
    summary = []
    for key, label, counter in summary_rows:
        summary.append(
            {
                "key": key,
                "label": label,
                "values": {
                    str(employee.id): counter(
                        [
                            records_by_key[(employee.id, current_date)]
                            for current_date in dates
                            if (employee.id, current_date) in records_by_key
                        ]
                    )
                    for employee in employees
                },
            }
        )

    return {
        "month": f"{year:04d}-{month:02d}",
        "month_label": f"{MONTHS_RU[month]} {year}",
        "employees": [
            {
                "id": employee.id,
                "name": _employee_display_name(employee),
                "email": getattr(employee, "email", "") or "",
            }
            for employee in employees
        ],
        "rows": rows,
        "summary": summary,
    }


def build_attendance_matrix_export_workbook(
    *,
    employees,
    records,
    period_start: date,
    period_end: date,
    record_comments: dict[int, list[dict[str, str]]] | None = None,
) -> bytes:
    employees = list(employees)
    records_by_month: dict[tuple[int, int], list[AttendanceRecord]] = {}
    for record in records:
        records_by_month.setdefault((record.date.year, record.date.month), []).append(
            record
        )

    workbook = Workbook()
    workbook.remove(workbook.active)

    for year, month in _month_pairs_between(period_start, period_end):
        matrix = build_monthly_attendance_matrix(
            employees=employees,
            records=records_by_month.get((year, month), []),
            year=year,
            month=month,
        )
        _append_attendance_matrix_sheet(
            workbook,
            matrix,
            record_comments=record_comments or {},
        )

    _append_attendance_matrix_legend_sheet(workbook)

    if not workbook.worksheets:
        workbook.create_sheet("Посещаемость")

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _month_pairs_between(
    period_start: date,
    period_end: date,
) -> list[tuple[int, int]]:
    result = []
    year = period_start.year
    month = period_start.month
    while year < period_end.year or (year == period_end.year and month <= period_end.month):
        result.append((year, month))
        month += 1
        if month > 12:
            month = 1
            year += 1
    return result


def _append_attendance_matrix_sheet(
    workbook: Workbook,
    matrix: dict[str, Any],
    *,
    record_comments: dict[int, list[dict[str, str]]],
) -> None:
    sheet = workbook.create_sheet(_safe_sheet_title(matrix["month_label"]))
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    summary_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.append(["Дата", "День", *[employee["name"] for employee in matrix["employees"]]])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in matrix["rows"]:
        sheet.append(
            [
                row["date"],
                row["weekday"],
                *[
                    _format_matrix_export_cell(row["cells"][str(employee["id"])])
                    for employee in matrix["employees"]
                ],
            ]
        )
        excel_row = sheet.max_row
        for index, excel_cell in enumerate(sheet[excel_row], start=1):
            excel_cell.border = border
            excel_cell.alignment = Alignment(vertical="top", wrap_text=True)
            if index > 2:
                matrix_cell = row["cells"][str(matrix["employees"][index - 3]["id"])]
                status = matrix_cell["status"]
                excel_cell.fill = PatternFill(
                    "solid",
                    fgColor=MATRIX_STATUS_FILLS.get(status, MATRIX_STATUS_FILLS["empty"]),
                )
                note = _format_matrix_export_note(
                    row=row,
                    cell=matrix_cell,
                    comments=record_comments.get(matrix_cell.get("record_id") or 0, []),
                )
                if note:
                    excel_cell.comment = Comment(note, "EUSRR")
                record_id = matrix_cell.get("record_id")
                if record_id:
                    excel_cell.hyperlink = _attendance_record_day_events_url(record_id)
                    excel_cell.font = Font(color="0563C1", underline="single")

    if matrix["summary"]:
        sheet.append([])
    for summary_row in matrix["summary"]:
        sheet.append(
            [
                summary_row["label"],
                "",
                *[
                    summary_row["values"].get(str(employee["id"]), 0)
                    for employee in matrix["employees"]
                ],
            ]
        )
        excel_row = sheet.max_row
        for cell in sheet[excel_row]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = summary_fill
            cell.border = border

    sheet.freeze_panes = "C2"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 10
    for column_index in range(3, len(matrix["employees"]) + 3):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18


def _append_attendance_matrix_legend_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Легенда")
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rows = [
        ("Код", "Расшифровка"),
        ("OK", "Норма"),
        ("ОП", "Опоздание"),
        ("НД", "Недоработка"),
        ("ПР", "Переработка"),
        ("Отс.", "Отсутствие"),
        ("Вых", "Выходной по графику/календарю"),
        ("Тех", "Технический сбой"),
        ("ОТП", "Отпуск"),
        ("БЛ", "Больничный"),
        ("ОТГ", "Отгул"),
        ("ДЕКР", "Декрет"),
        ("Вне штата", "Уволен или еще не принят в выбранный день"),
        ("Кадровое событие", "Проход при кадровом статусе вне штата"),
        ("+ Nч", "Зафиксирована работа в нерабочий день"),
        ("К: N", "Количество комментариев EUSRR к записи"),
        ("Ссылка", "Клик по ячейке записи открывает модал событий LogStorm в EUSRR"),
    ]
    for row in rows:
        sheet.append(row)
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == 1:
                cell.font = Font(bold=True, color="0F172A")
                cell.fill = header_fill
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 56


def _attendance_record_day_events_url(record_id: int) -> str:
    site_url = getattr(
        django_settings,
        "SITE_URL",
        "https://corp.robotail.pro",
    ).rstrip("/")
    return f"{site_url}/attendance?record={record_id}&events=1"


def _safe_sheet_title(value: str) -> str:
    forbidden = str.maketrans({char: " " for char in "[]:*?/\\\""})
    return value.translate(forbidden)[:31] or "Посещаемость"


def _format_matrix_export_cell(cell: dict[str, Any]) -> str:
    if cell["status"] == "empty":
        return ""
    return str(cell.get("display_text") or cell.get("short_label") or "")


def _format_matrix_export_note(
    *,
    row: dict[str, Any],
    cell: dict[str, Any],
    comments: list[dict[str, str]],
) -> str:
    if cell["status"] == "empty" and not comments:
        return ""

    lines = [
        f"Дата: {row['date']} ({row['weekday']})",
        f"Статус: {cell.get('primary_label') or MATRIX_STATUS_LABELS.get(cell['status'], cell['status'])}",
    ]
    details = [str(item) for item in cell.get("detail_lines", []) if item]
    if details:
        lines.extend(details)

    if comments:
        lines.append("")
        lines.append("Комментарии EUSRR:")
        for comment in comments:
            author = comment.get("author") or "Сотрудник"
            created_at = comment.get("created_at") or ""
            text = comment.get("text") or ""
            prefix = f"- {author}"
            if created_at:
                prefix = f"{prefix}, {created_at}"
            lines.append(f"{prefix}: {text}")

    return "\n".join(lines)


def _format_export_number(value: Any) -> str:
    if value is None or value == "":
        return "0"
    numeric = float(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.2f}".rstrip("0").rstrip(".")


def resolve_employee_day_state(
    personnel_actions,
    record_date: date,
) -> EmployeePersonnelState:
    employee = personnel_actions[0].employee if personnel_actions else None
    if employee is None:
        return EmployeePersonnelState()
    return resolve_employee_personnel_state(
        employee,
        record_date,
        actions=personnel_actions,
    )


def build_attendance_record_defaults(
    raw_record: dict[str, Any],
    personnel_state: EmployeePersonnelState,
) -> dict[str, Any]:
    is_remote_personnel_state = personnel_state.status == ACTION_REMOTE
    work_hours = _nullable_float(raw_record.get("work_hours"))
    has_work = bool(
        _nullable_string(raw_record.get("arrival_time"))
        or _nullable_string(raw_record.get("departure_time"))
        or (work_hours is not None and work_hours > 0)
    )
    is_workday = bool(raw_record.get("is_workday", True))
    if "effective_is_workday" in raw_record:
        effective_is_workday = bool(raw_record.get("effective_is_workday"))
    else:
        effective_is_workday = is_workday and not personnel_state.is_non_working
    is_overtime = bool(raw_record.get("is_overtime", False))
    overtime_hours = _nullable_float(raw_record.get("overtime_hours"))

    is_late = bool(raw_record.get("is_late", False))
    late_minutes = _nullable_int(raw_record.get("late_minutes"))
    is_early_leave = bool(raw_record.get("is_early_leave", False))
    early_leave_minutes = _nullable_int(raw_record.get("early_leave_minutes"))
    is_underwork = bool(raw_record.get("is_underwork", False))
    underwork_hours = _nullable_float(raw_record.get("underwork_hours"))
    is_absent = bool(raw_record.get("is_absent", False))

    statuses = _list_value(raw_record.get("statuses"))
    employee_issues = _list_value(raw_record.get("employee_issues"))
    technical_issues = _list_value(raw_record.get("technical_issues"))
    non_working_reason = _non_working_reason(
        is_workday=is_workday,
        effective_is_workday=effective_is_workday,
        personnel_state=personnel_state,
    )

    if (
        effective_is_workday
        and not is_remote_personnel_state
        and not has_work
        and not is_absent
        and _has_absence_issue([*statuses, *employee_issues])
    ):
        is_absent = True

    if is_absent and not has_work:
        is_underwork = False
        underwork_hours = None

    if is_remote_personnel_state:
        is_late = False
        late_minutes = None
        is_early_leave = False
        early_leave_minutes = None
        is_underwork = False
        underwork_hours = None
        is_absent = False
        statuses = _suppress_personnel_non_working_issues(statuses)
        employee_issues = _suppress_personnel_non_working_issues(employee_issues)
        technical_issues = _suppress_personnel_non_working_issues(technical_issues)

    if not effective_is_workday:
        is_late = False
        late_minutes = None
        is_early_leave = False
        early_leave_minutes = None
        is_underwork = False
        underwork_hours = None
        is_absent = False
        statuses = _suppress_personnel_non_working_issues(statuses)
        employee_issues = _suppress_personnel_non_working_issues(employee_issues)
        technical_issues = _suppress_personnel_non_working_issues(technical_issues)

        if has_work:
            is_overtime = True
            if overtime_hours is None:
                overtime_hours = work_hours
            statuses = _append_unique(statuses, "work_outside_personnel_schedule")

    return {
        "display_name": str(raw_record.get("display_name") or ""),
        "arrival_time": _nullable_string(raw_record.get("arrival_time")),
        "departure_time": _nullable_string(raw_record.get("departure_time")),
        "work_hours": work_hours,
        "expected_hours": _nullable_float(raw_record.get("expected_hours")),
        "is_workday": is_workday,
        "effective_is_workday": effective_is_workday,
        "is_late": is_late,
        "late_minutes": late_minutes,
        "is_early_leave": is_early_leave,
        "early_leave_minutes": early_leave_minutes,
        "is_underwork": is_underwork,
        "underwork_hours": underwork_hours,
        "is_overtime": is_overtime,
        "overtime_hours": overtime_hours,
        "is_absent": is_absent,
        "statuses": statuses,
        "employee_issues": employee_issues,
        "technical_issues": technical_issues,
        "personnel_status": (
            personnel_state.status
            if personnel_state.is_non_working or is_remote_personnel_state
            else PERSONNEL_STATUS_NORMAL
        ),
        "personnel_status_label": (
            personnel_state.label
            if personnel_state.is_non_working or is_remote_personnel_state
            else ""
        ),
        "personnel_action_id": (
            personnel_state.action_id
            if personnel_state.is_non_working or is_remote_personnel_state
            else None
        ),
        "is_manually_edited": bool(raw_record.get("manual_edited", False)),
        "manual_edit_payload": (
            raw_record.get("manual_edit_payload")
            if isinstance(raw_record.get("manual_edit_payload"), dict)
            else {}
        ),
        "raw_data": {**raw_record, "non_working_reason": non_working_reason},
    }


def apply_manual_edit_payload_to_defaults(
    defaults: dict[str, Any],
    payload: dict[str, Any],
) -> dict[str, Any]:
    """Apply saved manual fields to a freshly analyzed record payload."""
    if not isinstance(payload, dict):
        return defaults

    result = dict(defaults)
    editable_fields = {
        "arrival_time",
        "departure_time",
        "work_hours",
        "expected_hours",
        "is_workday",
        "effective_is_workday",
        "is_late",
        "late_minutes",
        "is_early_leave",
        "early_leave_minutes",
        "is_underwork",
        "underwork_hours",
        "is_overtime",
        "overtime_hours",
        "is_absent",
    }
    for field in editable_fields:
        if field in payload:
            result[field] = payload[field]

    _normalize_issues_after_manual_payload(result, payload)
    result["raw_data"] = {
        **result.get("raw_data", {}),
        "manual_edited": True,
        "manual_edit_payload": {
            key: value
            for key, value in payload.items()
            if key in editable_fields
        },
    }
    return result


def normalize_attendance_record_manual_issues(
    record: AttendanceRecord,
    payload: dict[str, Any],
) -> None:
    fields = {
        "statuses": list(record.statuses),
        "employee_issues": list(record.employee_issues),
    }
    _normalize_issues_after_manual_payload(fields, payload)
    record.statuses = fields["statuses"]
    record.employee_issues = fields["employee_issues"]


def _normalize_issues_after_manual_payload(
    defaults: dict[str, Any],
    payload: dict[str, Any],
) -> None:
    markers_by_flag = {
        "is_late": ("late", "опозд"),
        "is_early_leave": ("early", "ранн"),
        "is_underwork": ("underwork", "недоработ"),
        "is_absent": ("absence", "absent", "отсутств"),
    }
    for flag, markers in markers_by_flag.items():
        if payload.get(flag) is not False:
            continue
        for field in ("statuses", "employee_issues"):
            defaults[field] = [
                item
                for item in _list_value(defaults.get(field))
                if not any(marker in str(item).lower() for marker in markers)
            ]


def _has_absence_issue(values: list[Any]) -> bool:
    return any(
        marker in str(value).lower()
        for value in values
        for marker in ABSENCE_ISSUE_MARKERS
    )


def _parse_record_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if value is None:
        return None
    return parse_date(str(value))


def _nullable_string(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value)


def _nullable_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _nullable_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _list_value(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _suppress_personnel_non_working_issues(values: list[Any]) -> list[Any]:
    result = []
    for value in values:
        normalized = str(value).lower()
        if any(marker in normalized for marker in SUPPRESSED_EMPLOYEE_ISSUE_MARKERS):
            continue
        result.append(value)
    return result


def _append_unique(values: list[Any], value: str) -> list[Any]:
    if value in values:
        return values
    return [*values, value]


def _monthly_matrix_cell(record: AttendanceRecord | None) -> dict[str, Any]:
    if record is None:
        return {
            "record_id": None,
            "arrival_time": None,
            "departure_time": None,
            "work_hours": None,
            "expected_hours": None,
            "status": "empty",
            "short_label": "",
            "display_text": "",
            "primary_label": MATRIX_STATUS_LABELS["empty"],
            "detail_lines": [],
            "issues": [],
            "is_workday": None,
            "effective_is_workday": None,
            "non_working_reason": "",
            "is_manually_edited": False,
            "manual_edited_at": None,
            "comments_count": 0,
        }

    issues = [
        *record.statuses,
        *record.employee_issues,
        *record.technical_issues,
    ]
    comments_count = getattr(record, "comments_count", 0) or 0
    status = _monthly_matrix_status(record)
    presentation = _monthly_matrix_cell_presentation(
        record=record,
        status=status,
        issues=issues,
        comments_count=comments_count,
    )
    raw_non_working_reason = record.raw_data.get("non_working_reason")
    if _is_remote_calendar_non_working_day(record):
        raw_non_working_reason = ""
    return {
        "record_id": record.id,
        "arrival_time": record.arrival_time,
        "departure_time": record.departure_time,
        "work_hours": record.work_hours,
        "expected_hours": record.expected_hours,
        "status": status,
        **presentation,
        "issues": issues,
        "is_workday": record.is_workday,
        "effective_is_workday": record.effective_is_workday,
        "non_working_reason": raw_non_working_reason
        or _record_non_working_reason(record),
        "is_late": record.is_late,
        "is_early_leave": record.is_early_leave,
        "is_underwork": record.is_underwork,
        "is_overtime": record.is_overtime,
        "is_absent": record.is_absent,
        "personnel_status": record.personnel_status,
        "personnel_status_label": record.personnel_status_label,
        "is_manually_edited": record.is_manually_edited,
        "manual_edited_at": record.manual_edited_at,
        "comments_count": comments_count,
    }


def _monthly_matrix_cell_presentation(
    *,
    record: AttendanceRecord,
    status: str,
    issues: list[Any],
    comments_count: int,
) -> dict[str, Any]:
    primary_label = _matrix_primary_label(record, status)
    short_label = _matrix_short_label(record, status)
    display_text = _matrix_display_text(record, status, short_label)
    detail_lines = _matrix_detail_lines(
        record=record,
        status=status,
        issues=issues,
        comments_count=comments_count,
    )
    return {
        "short_label": short_label,
        "display_text": display_text,
        "primary_label": primary_label,
        "detail_lines": detail_lines,
    }


def _matrix_primary_label(record: AttendanceRecord, status: str) -> str:
    if (
        record.personnel_status != PERSONNEL_STATUS_NORMAL
        and not _is_remote_calendar_non_working_day(record)
    ):
        return (
            PERSONNEL_STATUS_LEGEND.get(record.personnel_status)
            or record.personnel_status_label
            or "Кадровое событие"
        )
    if not record.effective_is_workday:
        return _record_non_working_reason(record) or MATRIX_STATUS_LABELS["non_working"]
    return MATRIX_STATUS_LABELS.get(status, status)


def _matrix_short_label(record: AttendanceRecord, status: str) -> str:
    if (
        record.personnel_status != PERSONNEL_STATUS_NORMAL
        and not _is_remote_calendar_non_working_day(record)
    ):
        return (
            PERSONNEL_STATUS_SHORT_LABELS.get(record.personnel_status)
            or record.personnel_status_label
            or MATRIX_STATUS_SHORT_LABELS.get(status, status)
        )
    if not record.effective_is_workday:
        if not record.is_workday:
            return "Вых"
        return MATRIX_STATUS_SHORT_LABELS["non_working"]
    return MATRIX_STATUS_SHORT_LABELS.get(status, status)


def _matrix_display_text(
    record: AttendanceRecord,
    status: str,
    short_label: str,
) -> str:
    time_part = (
        f"{_format_matrix_time(record.arrival_time)}/"
        f"{_format_matrix_time(record.departure_time)}"
    )

    if not record.effective_is_workday:
        if record.work_hours:
            return f"{short_label} + {_format_export_number(record.work_hours)}ч"
        return short_label

    if record.personnel_status == ACTION_REMOTE:
        if record.arrival_time or record.departure_time or record.work_hours:
            return (
                f"{short_label} · {time_part} · "
                f"{_format_export_number(record.work_hours)}ч"
            )
        return short_label

    if status == "normal":
        return f"{time_part} · {_format_export_number(record.work_hours)}ч"
    if status == "late":
        suffix = f" {record.late_minutes}м" if record.late_minutes else ""
        return f"{short_label}{suffix} · {time_part}"
    if status == "underwork":
        suffix = (
            f" {_format_export_number(record.underwork_hours)}ч"
            if record.underwork_hours
            else ""
        )
        return f"{short_label}{suffix} · {time_part}"
    if status == "overtime":
        suffix = (
            f" +{_format_export_number(record.overtime_hours)}ч"
            if record.overtime_hours
            else ""
        )
        return f"{short_label}{suffix} · {time_part}"
    if status == "technical":
        issue_count = len(record.technical_issues)
        label = f"{short_label} ({issue_count})" if issue_count else short_label
        return f"{label} · {time_part}" if record.arrival_time or record.departure_time else label
    return short_label


def _format_matrix_time(value: Any) -> str:
    if not value:
        return "-"
    text = str(value)
    if "T" in text:
        text = text.split("T", 1)[1]
    if len(text) >= 5 and text[2] == ":":
        return text[:5]
    return text


def _matrix_detail_lines(
    *,
    record: AttendanceRecord,
    status: str,
    issues: list[Any],
    comments_count: int,
) -> list[str]:
    lines = [
        f"Приход: {record.arrival_time or '-'}",
        f"Уход: {record.departure_time or '-'}",
        (
            "Часы: "
            f"{_format_export_number(record.work_hours)} / "
            f"{_format_export_number(record.expected_hours)}"
        ),
        f"Рабочий день: {'да' if record.effective_is_workday else 'нет'}",
    ]
    if record.personnel_status != PERSONNEL_STATUS_NORMAL:
        lines.append(
            "Кадровое событие: "
            f"{record.personnel_status_label or _matrix_primary_label(record, status)}"
        )
    elif not record.effective_is_workday:
        reason = _record_non_working_reason(record)
        if reason:
            lines.append(f"Причина нерабочего дня: {reason}")
    if record.is_manually_edited:
        lines.append("Источник: ручная корректировка EUSRR")

    if record.is_late and record.late_minutes:
        lines.append(f"Опоздание: {record.late_minutes} мин.")
    if record.is_early_leave and record.early_leave_minutes:
        lines.append(f"Ранний уход: {record.early_leave_minutes} мин.")
    if status != "absent" and record.is_underwork and record.underwork_hours:
        lines.append(f"Недоработка: {_format_export_number(record.underwork_hours)} ч.")
    if record.is_overtime and record.overtime_hours:
        lines.append(f"Переработка: {_format_export_number(record.overtime_hours)} ч.")
    if record.is_absent:
        lines.append("Отсутствие: да")
    if record.technical_issues:
        lines.append(
            "Технические проблемы: "
            + "; ".join(str(issue) for issue in record.technical_issues)
        )
    employee_issues = [
        str(issue)
        for issue in record.employee_issues
        if not (status == "absent" and "underwork" in str(issue).lower())
    ]
    if employee_issues:
        lines.append("Проблемы сотрудника: " + "; ".join(employee_issues))
    extra_issues = [
        str(issue)
        for issue in issues
        if issue not in record.technical_issues and issue not in record.employee_issues
        and not (status == "absent" and "underwork" in str(issue).lower())
    ]
    if extra_issues:
        lines.append("Статусы анализа: " + "; ".join(extra_issues))
    if comments_count:
        lines.append(f"Комментарии: {comments_count}")
    return lines


def _monthly_matrix_status(record: AttendanceRecord) -> str:
    if not record.effective_is_workday:
        if record.personnel_status == ACTION_DISMISSED and record.is_overtime:
            return "personnel_issue"
        if record.is_overtime:
            return "overtime"
        return "non_working"
    if record.technical_issues:
        return "technical"
    if _record_has_absence(record):
        return "absent"
    if record.is_underwork:
        return "underwork"
    if record.is_late:
        return "late"
    if record.is_overtime:
        return "overtime"
    return "normal"


def _record_has_absence(record: AttendanceRecord) -> bool:
    return bool(
        record.is_absent
        or _has_absence_issue(
            [
                *record.statuses,
                *record.employee_issues,
            ]
        )
    )


def _non_working_reason(
    *,
    is_workday: bool,
    effective_is_workday: bool,
    personnel_state: EmployeePersonnelState,
) -> str:
    if effective_is_workday:
        return ""
    if personnel_state.is_non_working:
        return personnel_state.label or "Кадровое нерабочее состояние"
    if not is_workday:
        return "Выходной по графику/календарю"
    return "Нерабочий день"


def _record_non_working_reason(record: AttendanceRecord) -> str:
    if record.effective_is_workday:
        return ""
    if record.personnel_status_label and record.personnel_status != ACTION_REMOTE:
        return record.personnel_status_label
    if not record.is_workday:
        return "Выходной по графику/календарю"
    return "Нерабочий день"


def _is_remote_calendar_non_working_day(record: AttendanceRecord) -> bool:
    return bool(
        record.personnel_status == ACTION_REMOTE
        and not record.effective_is_workday
        and not record.is_workday
    )


def _is_valid_workday_record(record: AttendanceRecord) -> bool:
    return bool(record.effective_is_workday and not record.technical_issues)


def _summary_accounted_days(records: list[AttendanceRecord]) -> int:
    return sum(1 for record in records if _is_valid_workday_record(record))


def _summary_unaccounted_days(records: list[AttendanceRecord]) -> int:
    return sum(1 for record in records if record.effective_is_workday and record.technical_issues)


def _summary_worked_hours(records: list[AttendanceRecord]) -> float:
    total = sum(
        record.work_hours or 0
        for record in records
        if _is_valid_workday_record(record)
    )
    return round(total, 2)


def _summary_late_days(records: list[AttendanceRecord]) -> int:
    return sum(1 for record in records if _is_valid_workday_record(record) and record.is_late)


def _summary_early_leave_days(records: list[AttendanceRecord]) -> int:
    return sum(
        1
        for record in records
        if _is_valid_workday_record(record) and record.is_early_leave
    )


def _summary_overtime_days(records: list[AttendanceRecord]) -> int:
    return sum(1 for record in records if _is_valid_workday_record(record) and record.is_overtime)


def _summary_absent_days(records: list[AttendanceRecord]) -> int:
    return sum(
        1
        for record in records
        if _is_valid_workday_record(record) and _record_has_absence(record)
    )


def _employee_display_name(employee: Employee) -> str:
    full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
    return full_name or employee.email or f"#{employee.id}"
