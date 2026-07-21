import io
import json
import zipfile
from datetime import date

import pytest
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook

from finance.models import PayrollDailyWorkEntry, PayrollPeriod, PayrollWorkRecord
from finance.payroll.workbook_import import (
    _period_import_lock_queryset,
    _sanitize_styles,
    parse_workbook,
)

pytestmark = pytest.mark.django_db


def grant(user, codename):
    permission = Permission.objects.get(
        content_type__app_label="finance",
        codename=codename,
    )
    user.user_permissions.add(permission)
    for cache_name in ("_perm_cache", "_user_perm_cache", "_group_perm_cache"):
        if hasattr(user, cache_name):
            delattr(user, cache_name)


def workbook_bytes(*, ivan_first_day=5):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026год"
    for offset in range(7):
        sheet.cell(1, offset + 2, offset + 1)
        sheet.cell(2, offset + 2, "пн")
    sheet.cell(2, 1, "Июль")
    sheet.cell(3, 1, "Иван Иванов")
    sheet.cell(4, 1, "Алексей")
    for offset in range(7):
        sheet.cell(3, offset + 2, ivan_first_day if offset == 0 else 5)
        sheet.cell(4, offset + 2, 4)
    target = io.BytesIO()
    workbook.save(target)
    return target.getvalue()


def upload(payload):
    return SimpleUploadedFile(
        "График.xlsx",
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def make_period(creator):
    return PayrollPeriod.objects.create(
        code="2026-07",
        name="Июль 2026",
        date_from="2026-07-01",
        date_to="2026-07-31",
        pay_date="2026-08-05",
        currency="RUB",
        created_by=creator,
    )


def test_import_period_lock_targets_only_the_period_table():
    query = _period_import_lock_queryset().query

    assert query.select_for_update is True
    assert query.select_for_update_of == ("self",)


def test_invalid_legacy_colour_can_be_repaired_before_parsing():
    payload = workbook_bytes()
    source = zipfile.ZipFile(io.BytesIO(payload))
    target = io.BytesIO()
    with source, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as repaired:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                content = content.replace(b'theme="1"', b'rgb="none"', 1)
            repaired.writestr(item, content)
    period = type(
        "Period",
        (),
        {"date_from": date(2026, 7, 1), "date_to": date(2026, 7, 31)},
    )()

    repaired_payload = _sanitize_styles(target.getvalue())
    with zipfile.ZipFile(io.BytesIO(repaired_payload)) as repaired:
        assert b'rgb="none"' not in repaired.read("xl/styles.xml")
    parsed = parse_workbook(repaired_payload, period)

    assert len(parsed["rows"]) == 2
    assert parsed["rows"][0]["entries"][0]["points"] == 5


def test_api_previews_manual_mapping_and_skips_or_replaces_existing(
    user_factory,
    auth_client_factory,
):
    manager = user_factory(email="workbook.manager@example.test")
    ivan = user_factory(
        email="ivan.workbook@example.test",
        first_name="Иван",
        last_name="Иванов",
    )
    alexey = user_factory(
        email="alexey.one.workbook@example.test",
        first_name="Алексей",
        last_name="Первый",
    )
    alexey_two = user_factory(
        email="alexey.two.workbook@example.test",
        first_name="Алексей",
        last_name="Второй",
    )
    grant(manager, "manage_payroll_inputs")
    period = make_period(manager)
    existing = PayrollDailyWorkEntry.objects.create(
        period=period,
        employee=ivan,
        work_date="2026-07-01",
        target_points="5",
        actual_points="9",
    )
    manual_monthly = PayrollWorkRecord.objects.create(
        period=period,
        employee=alexey,
        target_points="110",
        actual_points="99",
        created_by=manager,
    )
    client = auth_client_factory(manager)
    preview_url = reverse(
        "api:v1:finance-payroll:admin-period-workbook-import-preview",
        kwargs={"pk": period.pk},
    )
    apply_url = reverse(
        "api:v1:finance-payroll:admin-period-workbook-import-apply",
        kwargs={"pk": period.pk},
    )
    payload = workbook_bytes()

    preview_response = client.post(
        preview_url, {"file": upload(payload)}, format="multipart"
    )

    assert preview_response.status_code == 200, preview_response.content
    preview = preview_response.json()
    rows = {row["source_name"]: row for row in preview["rows"]}
    assert rows["Иван Иванов"]["match_status"] == "matched"
    assert rows["Иван Иванов"]["matched_employee_id"] == ivan.pk
    assert rows["Алексей"]["match_status"] == "ambiguous"
    assert set(rows["Алексей"]["candidate_employee_ids"]) == {
        alexey.pk,
        alexey_two.pk,
    }
    # The ambiguous row is counted as a conflict only after the user maps it.
    assert rows["Алексей"]["existing_period_record"] is False
    assert preview["summary"]["existing"] == 1

    mappings = {
        rows["Иван Иванов"]["row_key"]: ivan.pk,
        rows["Алексей"]["row_key"]: alexey.pk,
    }
    skip_mappings = {
        rows["Иван Иванов"]["row_key"]: ivan.pk,
        rows["Алексей"]["row_key"]: None,
    }
    skipped_response = client.post(
        apply_url,
        {
            "file": upload(payload),
            "mode": "skip_existing",
            "mappings": json.dumps(skip_mappings),
            "expected_file_hash": preview["file_hash"],
            "expected_period_lock_version": preview["period_lock_version"],
        },
        format="multipart",
    )

    assert skipped_response.status_code == 200, skipped_response.content
    assert skipped_response.json()["summary"] == {
        "created": 6,
        "replaced": 0,
        "unchanged": 0,
        "skipped": 8,
    }
    existing.refresh_from_db()
    assert existing.actual_points == 9
    assert PayrollWorkRecord.objects.filter(period=period, employee=ivan).exists()
    assert PayrollWorkRecord.objects.filter(period=period, employee=alexey).exists()
    manual_monthly.refresh_from_db()
    assert manual_monthly.actual_points == 99

    second_preview = client.post(
        preview_url, {"file": upload(payload)}, format="multipart"
    ).json()
    replaced_response = client.post(
        apply_url,
        {
            "file": upload(payload),
            "mode": "replace_existing",
            "mappings": json.dumps(mappings),
            "expected_file_hash": second_preview["file_hash"],
            "expected_period_lock_version": second_preview["period_lock_version"],
        },
        format="multipart",
    )

    assert replaced_response.status_code == 200, replaced_response.content
    assert replaced_response.json()["summary"] == {
        "created": 7,
        "replaced": 1,
        "unchanged": 6,
        "skipped": 0,
    }
    existing.refresh_from_db()
    assert existing.actual_points == 5
    manual_monthly.refresh_from_db()
    assert manual_monthly.actual_points == 28
