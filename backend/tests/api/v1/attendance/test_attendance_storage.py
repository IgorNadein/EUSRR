from unittest.mock import patch
from io import BytesIO
from datetime import datetime, timedelta

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from attendance.models import (
    AttendanceAnalysisRun,
    AttendanceAutoSyncSettings,
    AttendanceRecord,
    EmployeeWorkSchedule,
    StandardWorkSchedule,
)
from common.logstorm_client import LogStormClientError
from employees.constants import ACTION_ON_LEAVE, ACTION_RETURNED_FROM_LEAVE
from employees.models import EmployeeAction

pytestmark = pytest.mark.django_db


def _analyze_url():
    return reverse("api:v1:logstorm-attendance-analyze")


def _records_url():
    return reverse("api:v1:attendance-records")


def _monthly_matrix_url():
    return reverse("api:v1:attendance-monthly-matrix")


def _monthly_matrix_export_url():
    return reverse("api:v1:attendance-monthly-matrix-export")


def _work_schedule_url(employee_id):
    return reverse("api:v1:attendance-work-schedule", args=[employee_id])


def _standard_work_schedule_url():
    return reverse("api:v1:attendance-standard-work-schedule")


def _auto_sync_settings_url():
    return reverse("api:v1:attendance-auto-sync-settings")


def _auto_sync_run_now_url():
    return reverse("api:v1:attendance-auto-sync-run-now")


def _record_detail_url(record_id):
    return reverse("api:v1:attendance-record-detail", args=[record_id])


def _record_day_events_url(record_id):
    return reverse("api:v1:attendance-record-day-events", args=[record_id])


def _record_day_event_photo_url(record_id, event_key):
    return reverse(
        "api:v1:attendance-record-day-event-photo",
        args=[record_id, event_key],
    )


def _record_comments_url(record_id):
    return reverse("api:v1:attendance-record-comments", args=[record_id])


def _record_comment_detail_url(record_id, comment_id):
    return reverse(
        "api:v1:attendance-record-comment-detail", args=[record_id, comment_id]
    )


def _payload(employee_id, **overrides):
    payload = {
        "employee_id": employee_id,
        "period_start": "2026-04-20",
        "period_end": "2026-04-21",
    }
    payload.update(overrides)
    return payload


def _logstorm_result():
    return {
        "records": [
            {
                "date": "2026-04-20",
                "employee_id": "42",
                "display_name": "Ivan Petrov",
                "arrival_time": "09:15",
                "departure_time": "18:00",
                "work_hours": 8.75,
                "expected_hours": 9,
                "is_workday": True,
                "is_late": True,
                "late_minutes": 15,
                "is_early_leave": False,
                "early_leave_minutes": None,
                "is_underwork": True,
                "underwork_hours": 0.25,
                "is_overtime": False,
                "overtime_hours": None,
                "is_absent": False,
                "statuses": ["late", "underwork"],
                "employee_issues": ["late"],
                "technical_issues": [],
            },
            {
                "date": "2026-04-21",
                "employee_id": "42",
                "display_name": "Ivan Petrov",
                "arrival_time": None,
                "departure_time": None,
                "work_hours": 0,
                "expected_hours": 9,
                "is_workday": True,
                "is_absent": True,
                "statuses": ["absence"],
                "employee_issues": ["absence"],
                "technical_issues": [],
            },
        ]
    }


def _aware_datetime(year, month, day):
    return timezone.make_aware(datetime(year, month, day))


def _move_auto_hired_before_test_period(employee):
    employee.actions.filter(action="hired").update(date=_aware_datetime(2026, 1, 1))


class _PhotoResponse:
    content = b"fake-image"
    headers = {"Content-Type": "image/jpeg"}


def test_analyze_saves_run_and_records(auth_client_factory, user_factory):
    staff = user_factory(staff=True)
    employee = user_factory(first_name="Ivan", last_name="Petrov")
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    assert AttendanceAnalysisRun.objects.count() == 1
    assert AttendanceRecord.objects.count() == 2

    run = AttendanceAnalysisRun.objects.get()
    assert run.employee == employee
    assert run.triggered_by == staff
    assert run.period_start.isoformat() == "2026-04-20"
    assert run.period_end.isoformat() == "2026-04-21"
    assert run.request_payload["employee_id"] == str(employee.id)

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    assert record.analysis_run == run
    assert record.arrival_time == "09:15"
    assert record.is_late is True
    assert record.effective_is_workday is True
    assert record.personnel_status == "normal"
    assert record.statuses == ["late", "underwork"]
    assert record.raw_data["employee_id"] == "42"


def test_work_schedule_endpoint_returns_default_for_self(
    auth_client_factory,
    user_factory,
):
    employee = user_factory()
    client = auth_client_factory(employee)

    response = client.get(_work_schedule_url(employee.id))

    assert response.status_code == 200
    assert response.data["employee_id"] == employee.id
    assert response.data["is_default"] is True
    assert response.data["start_time"] == "08:00"
    assert response.data["workdays"] == EmployeeWorkSchedule.DEFAULT_WORKDAYS


def test_staff_can_save_employee_work_schedule(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    response = client.patch(
        _work_schedule_url(employee.id),
        {
            "start_time": "09:00",
            "end_time": "18:00",
            "expected_hours": 8,
            "workdays": ["Monday", "Wednesday", "Friday"],
            "is_active": True,
        },
        format="json",
    )

    assert response.status_code == 200
    assert response.data["is_default"] is False
    assert response.data["start_time"] == "09:00:00"
    assert response.data["workdays"] == ["Monday", "Wednesday", "Friday"]

    schedule = EmployeeWorkSchedule.objects.get(employee=employee)
    assert schedule.updated_by == staff
    assert schedule.expected_hours == 8


def test_staff_can_read_default_standard_work_schedule(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    client = auth_client_factory(staff)

    response = client.get(_standard_work_schedule_url())

    assert response.status_code == 200
    assert response.data["id"] is None
    assert response.data["is_default"] is True
    assert response.data["start_time"] == "08:00"
    assert response.data["end_time"] == "17:00"
    assert response.data["expected_hours"] == 9
    assert response.data["workdays"] == EmployeeWorkSchedule.DEFAULT_WORKDAYS


def test_staff_can_save_standard_work_schedule(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    client = auth_client_factory(staff)

    response = client.patch(
        _standard_work_schedule_url(),
        {
            "start_time": "07:30",
            "end_time": "16:30",
            "expected_hours": 8,
            "workdays": ["Monday", "Tuesday", "Wednesday", "Thursday"],
        },
        format="json",
    )

    assert response.status_code == 200
    assert response.data["is_default"] is False
    assert response.data["start_time"] == "07:30:00"
    assert response.data["end_time"] == "16:30:00"
    assert response.data["workdays"] == [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
    ]

    schedule = StandardWorkSchedule.objects.get()
    assert schedule.updated_by == staff
    assert schedule.expected_hours == 8

    get_response = client.get(_standard_work_schedule_url())
    assert get_response.status_code == 200
    assert get_response.data["id"] == schedule.id
    assert get_response.data["start_time"] == "07:30:00"


def test_non_staff_cannot_manage_standard_work_schedule(
    auth_client_factory,
    user_factory,
):
    user = user_factory()
    client = auth_client_factory(user)

    get_response = client.get(_standard_work_schedule_url())
    patch_response = client.patch(
        _standard_work_schedule_url(),
        {"start_time": "09:00"},
        format="json",
    )

    assert get_response.status_code == 403
    assert patch_response.status_code == 403
    assert not StandardWorkSchedule.objects.exists()


def test_staff_can_read_default_auto_sync_settings(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    client = auth_client_factory(staff)

    response = client.get(_auto_sync_settings_url())

    assert response.status_code == 200
    assert response.data["enabled"] is False
    assert response.data["frequency_minutes"] == 1440
    assert response.data["lookback_days"] == 3
    assert response.data["last_status"] == "idle"
    assert AttendanceAutoSyncSettings.objects.count() == 1


def test_staff_can_update_auto_sync_settings(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    client = auth_client_factory(staff)

    response = client.patch(
        _auto_sync_settings_url(),
        {
            "enabled": True,
            "frequency_minutes": 15,
            "lookback_days": 7,
        },
        format="json",
    )

    assert response.status_code == 200
    assert response.data["enabled"] is True
    assert response.data["frequency_minutes"] == 15
    assert response.data["lookback_days"] == 7
    assert response.data["next_run_at"] is not None
    settings = AttendanceAutoSyncSettings.objects.get()
    assert settings.updated_by == staff


def test_non_staff_cannot_manage_auto_sync_settings(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    client = auth_client_factory(user)

    get_response = client.get(_auto_sync_settings_url())
    patch_response = client.patch(
        _auto_sync_settings_url(),
        {"enabled": True},
        format="json",
    )
    run_response = client.post(_auto_sync_run_now_url())

    assert get_response.status_code == 403
    assert patch_response.status_code == 403
    assert run_response.status_code == 403


def test_non_staff_cannot_save_other_employee_work_schedule(
    auth_client_factory,
    user_factory,
):
    user = user_factory()
    employee = user_factory()
    client = auth_client_factory(user)

    response = client.patch(
        _work_schedule_url(employee.id),
        {"start_time": "09:00"},
        format="json",
    )

    assert response.status_code == 403
    assert not EmployeeWorkSchedule.objects.filter(employee=employee).exists()


def test_analyze_uses_saved_work_schedule_when_payload_omits_schedule(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    EmployeeWorkSchedule.objects.create(
        employee=employee,
        start_time="10:00",
        end_time="19:00",
        expected_hours=8,
        workdays=["Tuesday"],
    )
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ) as analyze:
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    schedule = analyze.call_args.kwargs["schedule"]
    assert schedule == {
        "start_time": "10:00",
        "end_time": "19:00",
        "expected_hours": 8,
        "workdays": ["Tuesday"],
        "date_overrides": [],
    }

    run = AttendanceAnalysisRun.objects.get()
    assert run.schedule_payload == schedule


def test_analyze_uses_standard_work_schedule_when_payload_omits_schedule(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    StandardWorkSchedule.objects.create(
        start_time="07:00",
        end_time="16:00",
        expected_hours=8,
        workdays=["Monday", "Tuesday", "Wednesday", "Thursday"],
    )
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ) as analyze:
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    schedule = analyze.call_args.kwargs["schedule"]
    assert schedule == {
        "start_time": "07:00",
        "end_time": "16:00",
        "expected_hours": 8,
        "workdays": ["Monday", "Tuesday", "Wednesday", "Thursday"],
        "date_overrides": [],
    }

    run = AttendanceAnalysisRun.objects.get()
    assert run.schedule_payload == schedule


def _auto_sync_result(employee, target_date):
    return {
        "records": [
            {
                "date": target_date.isoformat(),
                "employee_id": str(employee.id),
                "display_name": str(employee),
                "arrival_time": "08:00",
                "departure_time": "17:00",
                "work_hours": 9,
                "expected_hours": 9,
                "is_workday": True,
                "is_absent": False,
                "statuses": [],
                "employee_issues": [],
                "technical_issues": [],
            }
        ]
    }


def test_auto_sync_disabled_settings_do_not_run_analysis(user_factory):
    from attendance.services import run_attendance_auto_sync

    user_factory()
    AttendanceAutoSyncSettings.objects.create(enabled=False)

    with patch("common.logstorm_attendance.analyze_employee_attendance") as analyze:
        settings = run_attendance_auto_sync(force=False)

    analyze.assert_not_called()
    assert settings.last_status == "idle"


def test_auto_sync_updates_due_active_employees_and_uses_schedules(user_factory):
    from attendance.services import run_attendance_auto_sync

    employee = user_factory(first_name="Ivan", last_name="Petrov")
    other = user_factory(first_name="Anna", last_name="Sidorova")
    user_factory(active=False)
    EmployeeWorkSchedule.objects.create(
        employee=employee,
        start_time="10:00",
        end_time="19:00",
        expected_hours=8,
        workdays=["Monday"],
    )
    StandardWorkSchedule.objects.create(
        start_time="07:30",
        end_time="16:30",
        expected_hours=8,
        workdays=["Tuesday"],
    )
    now = timezone.now()
    AttendanceAutoSyncSettings.objects.create(
        enabled=True,
        frequency_minutes=15,
        lookback_days=3,
        next_run_at=now - timedelta(minutes=1),
    )
    today = timezone.localdate()

    def analyze_result(*, employee, **kwargs):
        return _auto_sync_result(employee, today)

    with patch(
        "common.logstorm_attendance.analyze_employee_attendance",
        side_effect=analyze_result,
    ) as analyze:
        settings = run_attendance_auto_sync(force=False)

    assert analyze.call_count == 2
    calls_by_employee_id = {
        call.kwargs["employee"].id: call.kwargs
        for call in analyze.call_args_list
    }
    employee_call = calls_by_employee_id[employee.id]
    other_call = calls_by_employee_id[other.id]
    assert employee_call["period_start"] == today - timedelta(days=2)
    assert employee_call["period_end"] == today
    assert employee_call["schedule"]["start_time"] == "10:00"
    assert other_call["schedule"]["start_time"] == "07:30"
    assert settings.last_status == "success"
    assert settings.last_success_count == 2
    assert settings.last_error_count == 0
    assert settings.next_run_at > settings.last_finished_at
    assert AttendanceRecord.objects.filter(employee__in=[employee, other]).count() == 2


def test_auto_sync_continues_after_employee_error(user_factory):
    from attendance.services import run_attendance_auto_sync

    first = user_factory(first_name="Broken", last_name="User")
    second = user_factory(first_name="Good", last_name="User")
    AttendanceAutoSyncSettings.objects.create(
        enabled=True,
        frequency_minutes=60,
        lookback_days=1,
        next_run_at=timezone.now() - timedelta(minutes=1),
    )
    today = timezone.localdate()

    def analyze_result(*, employee, **kwargs):
        if employee == first:
            raise LogStormClientError("logstorm down")
        return _auto_sync_result(employee, today)

    with patch(
        "common.logstorm_attendance.analyze_employee_attendance",
        side_effect=analyze_result,
    ):
        settings = run_attendance_auto_sync(force=False)

    assert settings.last_status == "partial"
    assert settings.last_success_count == 1
    assert settings.last_error_count == 1
    assert "logstorm down" in settings.last_error
    assert AttendanceRecord.objects.filter(employee=second, date=today).exists()


def test_auto_sync_skips_parallel_running_state(user_factory):
    from attendance.services import run_attendance_auto_sync

    user_factory()
    AttendanceAutoSyncSettings.objects.create(
        enabled=True,
        last_status=AttendanceAutoSyncSettings.STATUS_RUNNING,
        next_run_at=timezone.now() - timedelta(minutes=1),
    )

    with patch("common.logstorm_attendance.analyze_employee_attendance") as analyze:
        settings = run_attendance_auto_sync(force=False)

    analyze.assert_not_called()
    assert settings.last_status == "running"


def test_staff_can_run_auto_sync_now(auth_client_factory, user_factory):
    staff = user_factory(staff=True)
    client = auth_client_factory(staff)
    settings = AttendanceAutoSyncSettings.objects.create(
        enabled=False,
        last_status=AttendanceAutoSyncSettings.STATUS_SUCCESS,
    )

    with patch(
        "api.v1.attendance.views.run_attendance_auto_sync",
        return_value=settings,
    ) as run_sync:
        response = client.post(_auto_sync_run_now_url())

    assert response.status_code == 200
    run_sync.assert_called_once_with(force=True)
    assert response.data["last_status"] == "success"


def test_analyze_updates_existing_daily_record(auth_client_factory, user_factory):
    staff = user_factory(staff=True)
    employee = user_factory()
    _move_auto_hired_before_test_period(employee)
    client = auth_client_factory(staff)

    first_result = _logstorm_result()
    second_result = {
        "records": [
            {
                **first_result["records"][0],
                "arrival_time": "08:55",
                "is_late": False,
                "late_minutes": None,
                "statuses": [],
                "employee_issues": [],
            }
        ]
    }

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=first_result,
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=second_result,
    ):
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    assert AttendanceAnalysisRun.objects.count() == 2
    assert (
        AttendanceRecord.objects.filter(employee=employee, date="2026-04-20").count()
        == 1
    )

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    assert record.arrival_time == "08:55"
    assert record.is_late is False
    assert record.statuses == []


def test_records_endpoint_returns_saved_records(auth_client_factory, user_factory):
    staff = user_factory(staff=True)
    employee = user_factory()
    other = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")
        client.post(_analyze_url(), _payload(other.id), format="json")

    response = client.get(
        _records_url(),
        {
            "employee_id": employee.id,
            "date_from": "2026-04-20",
            "date_to": "2026-04-21",
        },
    )

    assert response.status_code == 200
    results = response.json()["results"]
    assert len(results) == 2
    assert {item["employee"] for item in results} == {employee.id}
    assert {item["date"] for item in results} == {"2026-04-20", "2026-04-21"}


def test_user_can_read_own_attendance_record_detail(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(user.id), format="json")

    record = AttendanceRecord.objects.get(employee=user, date="2026-04-20")

    response = client.get(_record_detail_url(record.id))

    assert response.status_code == 200
    assert response.json()["id"] == record.id
    assert response.json()["employee"] == user.id


def test_staff_can_load_attendance_day_events(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    with patch(
        "api.v1.attendance.views.LogStormClient.get_attendance_day_events",
        return_value=[
            {
                "event_key": "abc",
                "time": "2026-04-20T09:15:00+03:00",
                "time_label": "09:15:00",
                "caption": "Успешный вход",
                "device": "door-1",
                "device_name": "Вход",
                "serial_no": 10,
                "has_photo": True,
                "photo_url": "/attendance/events/photos/abc/",
            },
            {
                "event_key": "def",
                "time": "2026-04-20T18:00:00+03:00",
                "time_label": "18:00:00",
                "caption": "Успешный выход",
                "device": "door-2",
                "device_name": "Выход",
                "serial_no": 11,
                "has_photo": False,
                "photo_url": None,
            },
        ],
    ) as logstorm_events:
        response = client.get(_record_day_events_url(record.id))

    assert response.status_code == 200
    logstorm_events.assert_called_once_with(
        employee_id=str(employee.id),
        record_date=record.date,
    )
    events = response.json()
    assert [event["caption"] for event in events] == ["Успешный вход", "Успешный выход"]
    assert events[0]["photo_url"].endswith(
        f"/api/v1/attendance/records/{record.id}/day-events/abc/photo/"
    )
    assert events[1]["photo_url"] is None


def test_user_cannot_load_other_employee_attendance_day_events(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    user = user_factory(staff=False)
    employee = user_factory()
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")

    response = client.get(_record_day_events_url(record.id))

    assert response.status_code == 404


def test_attendance_day_events_logstorm_error_returns_502(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    with patch(
        "api.v1.attendance.views.LogStormClient.get_attendance_day_events",
        side_effect=LogStormClientError("boom"),
    ):
        response = client.get(_record_day_events_url(record.id))

    assert response.status_code == 502
    assert response.json()["error"] == "logstorm_unavailable"


def test_staff_can_proxy_attendance_day_event_photo(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    with (
        patch(
            "api.v1.attendance.views.LogStormClient.get_attendance_day_events",
            return_value=[{"event_key": "abc", "has_photo": True}],
        ),
        patch(
            "api.v1.attendance.views.LogStormClient.get_attendance_event_photo",
            return_value=_PhotoResponse(),
        ) as logstorm_photo,
    ):
        response = client.get(_record_day_event_photo_url(record.id, "abc"))

    assert response.status_code == 200
    assert response.content == b"fake-image"
    assert response["Content-Type"] == "image/jpeg"
    logstorm_photo.assert_called_once_with("abc")


def test_attendance_day_event_photo_rejects_event_outside_record_day(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    with (
        patch(
            "api.v1.attendance.views.LogStormClient.get_attendance_day_events",
            return_value=[{"event_key": "other", "has_photo": True}],
        ),
        patch(
            "api.v1.attendance.views.LogStormClient.get_attendance_event_photo",
        ) as logstorm_photo,
    ):
        response = client.get(_record_day_event_photo_url(record.id, "abc"))

    assert response.status_code == 404
    logstorm_photo.assert_not_called()


def test_monthly_matrix_endpoint_returns_excel_like_month(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory(first_name="Ivan", last_name="Petrov")
    other = user_factory(first_name="Anna", last_name="Sidorova")
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")
        client.post(_analyze_url(), _payload(other.id), format="json")

    response = client.get(
        _monthly_matrix_url(),
        {
            "employee_ids": f"{employee.id},{other.id}",
            "month": "2026-04",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["month"] == "2026-04"
    assert payload["month_label"] == "Апрель 2026"
    assert [item["id"] for item in payload["employees"]] == [employee.id, other.id]
    assert len(payload["rows"]) == 30

    target_row = next(row for row in payload["rows"] if row["date"] == "2026-04-20")
    employee_cell = target_row["cells"][str(employee.id)]
    assert employee_cell["arrival_time"] == "09:15"
    assert employee_cell["departure_time"] == "18:00"
    assert employee_cell["status"] == "underwork"
    assert employee_cell["short_label"] == "НД"
    assert employee_cell["display_text"] == "НД 0.25ч · 09:15/18:00"
    assert "Недоработка: 0.25 ч." in employee_cell["detail_lines"]

    summary_by_key = {item["key"]: item for item in payload["summary"]}
    assert summary_by_key["late_days"]["values"][str(employee.id)] == 1
    assert summary_by_key["absent_days"]["values"][str(employee.id)] == 1
    assert summary_by_key["worked_hours"]["values"][str(employee.id)] == 8.75


def test_staff_can_export_monthly_matrix_xlsx_from_saved_records(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory(first_name="Ivan", last_name="Petrov")
    other = user_factory(first_name="Anna", last_name="Sidorova")
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")
        client.post(_analyze_url(), _payload(other.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    client.post(
        _record_comments_url(record.id),
        {"text": "Проверить причину опоздания"},
        format="json",
    )

    with patch("api.v1.attendance.views.analyze_employee_attendance") as analyze:
        response = client.get(
            _monthly_matrix_export_url(),
            {
                "employee_ids": f"{employee.id},{other.id}",
                "period_start": "2026-04-20",
                "period_end": "2026-05-05",
            },
        )

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"] == (
        'attachment; filename="attendance-2026-04-20_2026-05-05.xlsx"'
    )
    analyze.assert_not_called()

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Апрель 2026", "Май 2026", "Легенда"]

    april = workbook["Апрель 2026"]
    assert april["A1"].value == "Дата"
    assert april["B1"].value == "День"
    assert april["C1"].value == "Petrov Ivan"
    assert april["D1"].value == "Sidorova Anna"
    assert april["A21"].value == "2026-04-20"
    assert april["C21"].value == "НД 0.25ч · 09:15/18:00"
    assert april["C21"].hyperlink.target.endswith(
        f"/attendance?record={record.id}&events=1"
    )
    assert "Недоработка: 0.25 ч." in april["C21"].comment.text
    assert "Проверить причину опоздания" in april["C21"].comment.text
    assert april["C33"].value == 2
    assert april["C35"].value == 8.75
    assert workbook["Легенда"]["A2"].value == "OK"


def test_monthly_matrix_prioritizes_absence_when_no_times_and_no_personnel_reason(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)
    result = {
        "records": [
            {
                "date": "2026-04-20",
                "employee_id": str(employee.id),
                "display_name": "Absent User",
                "arrival_time": None,
                "departure_time": None,
                "work_hours": 0,
                "expected_hours": 9,
                "is_workday": True,
                "is_absent": True,
                "is_underwork": True,
                "underwork_hours": 9,
                "statuses": ["absence", "underwork"],
                "employee_issues": ["absence", "underwork"],
                "technical_issues": [],
            }
        ]
    }

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=result,
    ):
        response = client.post(
            _analyze_url(),
            _payload(employee.id, period_start="2026-04-20", period_end="2026-04-20"),
            format="json",
        )
    assert response.status_code == 200

    matrix_response = client.get(
        _monthly_matrix_url(),
        {
            "employee_ids": str(employee.id),
            "month": "2026-04",
        },
    )
    matrix_row = next(
        row for row in matrix_response.json()["rows"] if row["date"] == "2026-04-20"
    )
    matrix_cell = matrix_row["cells"][str(employee.id)]
    assert matrix_cell["status"] == "absent"
    assert matrix_cell["display_text"] == "Н"
    assert matrix_cell["primary_label"] == "Отсутствие"


def test_non_staff_cannot_export_monthly_matrix_xlsx(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    client = auth_client_factory(user)

    response = client.get(
        _monthly_matrix_export_url(),
        {
            "employee_ids": str(user.id),
            "period_start": "2026-04-20",
            "period_end": "2026-04-21",
        },
    )

    assert response.status_code == 403


def test_monthly_matrix_xlsx_export_rejects_invalid_period(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    response = client.get(
        _monthly_matrix_export_url(),
        {
            "employee_ids": str(employee.id),
            "period_start": "2026-04-21",
            "period_end": "2026-04-20",
        },
    )

    assert response.status_code == 400


def test_monthly_matrix_xlsx_export_includes_empty_month_without_records(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory(first_name="Empty", last_name="Month")
    client = auth_client_factory(staff)

    response = client.get(
        _monthly_matrix_export_url(),
        {
            "employee_ids": str(employee.id),
            "period_start": "2026-04-20",
            "period_end": "2026-04-21",
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    april = workbook["Апрель 2026"]
    assert april["C1"].value == "Month Empty"
    assert april["C21"].value is None


def test_monthly_matrix_xlsx_export_keeps_full_month_but_limits_records_to_period(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory(first_name="Ivan", last_name="Petrov")
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    response = client.get(
        _monthly_matrix_export_url(),
        {
            "employee_ids": str(employee.id),
            "period_start": "2026-04-21",
            "period_end": "2026-04-21",
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    april = workbook["Апрель 2026"]
    assert april.max_row == 39
    assert april["A20"].value == "2026-04-19"
    assert april["A21"].value == "2026-04-20"
    assert april["A22"].value == "2026-04-21"
    assert april["C21"].value is None
    assert april["C22"].value == "Н"
    assert april["C33"].value == 1
    assert april["C35"].value == 0


def test_user_can_read_own_attendance_records(auth_client_factory, user_factory):
    user = user_factory(staff=False)
    other = user_factory()
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(user.id), format="json")
        staff_client.post(_analyze_url(), _payload(other.id), format="json")

    response = client.get(_records_url())

    assert response.status_code == 200
    results = response.json()["results"]
    assert len(results) == 2
    assert {item["employee"] for item in results} == {user.id}


def test_user_cannot_read_other_employee_attendance_records(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    other = user_factory()
    client = auth_client_factory(user)

    response = client.get(_records_url(), {"employee_id": other.id})
    assert response.status_code == 403


def test_staff_can_update_attendance_record(auth_client_factory, user_factory):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    patch_payload = {
        "arrival_time": "09:00",
        "departure_time": "18:05",
        "work_hours": 9.08,
        "is_late": False,
        "late_minutes": None,
        "is_underwork": False,
        "underwork_hours": None,
    }
    with patch(
        "api.v1.attendance.views.LogStormClient.update_attendance_override",
        return_value={"id": 1, "patch": patch_payload},
    ) as logstorm_update:
        response = client.patch(
            _record_detail_url(record.id),
            patch_payload,
            format="json",
        )

    assert response.status_code == 200
    logstorm_update.assert_called_once()
    record.refresh_from_db()
    assert record.arrival_time == "09:00"
    assert record.work_hours == 9.08
    assert record.is_late is False
    assert record.is_manually_edited is True
    assert record.manual_edit_payload["arrival_time"] == "09:00"
    assert "late" not in record.statuses
    assert response.json()["arrival_time"] == "09:00"

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record.refresh_from_db()
    assert record.arrival_time == "09:00"
    assert record.is_manually_edited is True
    assert record.manual_edit_payload["arrival_time"] == "09:00"


def test_user_cannot_update_attendance_record(auth_client_factory, user_factory):
    user = user_factory(staff=False)
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(user.id), format="json")

    record = AttendanceRecord.objects.get(employee=user, date="2026-04-20")
    response = client.patch(
        _record_detail_url(record.id),
        {"arrival_time": "09:00"},
        format="json",
    )

    assert response.status_code == 403


def test_user_can_analyze_own_attendance(auth_client_factory, user_factory):
    user = user_factory(staff=False)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        response = client.post(_analyze_url(), _payload(user.id), format="json")

    assert response.status_code == 200
    assert AttendanceRecord.objects.filter(employee=user).count() == 2


def test_user_cannot_analyze_other_employee_attendance(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    other = user_factory()
    client = auth_client_factory(user)

    response = client.post(_analyze_url(), _payload(other.id), format="json")

    assert response.status_code == 403


def test_personnel_non_working_state_is_saved_and_suppresses_absence(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    _move_auto_hired_before_test_period(employee)
    client = auth_client_factory(staff)
    leave_action = EmployeeAction.objects.create(
        employee=employee,
        action=ACTION_ON_LEAVE,
        date=_aware_datetime(2026, 4, 20),
    )

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    absent_record = AttendanceRecord.objects.get(employee=employee, date="2026-04-21")
    assert absent_record.personnel_status == ACTION_ON_LEAVE
    assert absent_record.personnel_action == leave_action
    assert absent_record.effective_is_workday is False
    assert absent_record.is_absent is False
    assert absent_record.employee_issues == []
    assert absent_record.statuses == []

    matrix_response = client.get(
        _monthly_matrix_url(),
        {
            "employee_ids": str(employee.id),
            "month": "2026-04",
        },
    )
    matrix_row = next(
        row for row in matrix_response.json()["rows"] if row["date"] == "2026-04-21"
    )
    matrix_cell = matrix_row["cells"][str(employee.id)]
    assert matrix_cell["short_label"] == "ОТП"
    assert matrix_cell["display_text"] == "ОТП"
    assert matrix_cell["primary_label"] == "Отпуск"


def test_calendar_non_working_day_suppresses_absence_without_personnel_state(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    client = auth_client_factory(staff)
    weekend_absence_result = {
        "records": [
            {
                "date": "2026-04-19",
                "employee_id": str(employee.id),
                "display_name": "Weekend User",
                "arrival_time": None,
                "departure_time": None,
                "work_hours": 0,
                "expected_hours": 9,
                "is_workday": False,
                "is_absent": True,
                "statuses": ["День критического отсутствия (100% отсутствуют)"],
                "employee_issues": ["absence"],
                "technical_issues": [
                    "День критического отсутствия (100% отсутствуют)"
                ],
            }
        ]
    }

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=weekend_absence_result,
    ):
        response = client.post(
            _analyze_url(),
            _payload(employee.id, period_start="2026-04-19", period_end="2026-04-19"),
            format="json",
        )

    assert response.status_code == 200
    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-19")
    assert record.is_workday is False
    assert record.effective_is_workday is False
    assert record.is_absent is False
    assert record.statuses == []
    assert record.employee_issues == []
    assert record.technical_issues == []

    records_response = client.get(
        _records_url(),
        {
            "employee_id": employee.id,
            "date_from": "2026-04-19",
            "date_to": "2026-04-19",
        },
    )
    assert records_response.status_code == 200
    assert records_response.json()["results"][0]["non_working_reason"] == (
        "Выходной по графику/календарю"
    )

    matrix_response = client.get(
        _monthly_matrix_url(),
        {
            "employee_ids": str(employee.id),
            "month": "2026-04",
        },
    )
    matrix_row = next(
        row for row in matrix_response.json()["rows"] if row["date"] == "2026-04-19"
    )
    matrix_cell = matrix_row["cells"][str(employee.id)]
    assert matrix_cell["status"] == "non_working"
    assert matrix_cell["short_label"] == "Вых"
    assert matrix_cell["primary_label"] == "Выходной по графику/календарю"
    assert matrix_cell["issues"] == []


def test_personnel_non_working_state_counts_work_as_overtime(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    _move_auto_hired_before_test_period(employee)
    client = auth_client_factory(staff)
    EmployeeAction.objects.create(
        employee=employee,
        action=ACTION_ON_LEAVE,
        date=_aware_datetime(2026, 4, 20),
    )

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    worked_record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    assert worked_record.personnel_status == ACTION_ON_LEAVE
    assert worked_record.effective_is_workday is False
    assert worked_record.is_late is False
    assert worked_record.is_underwork is False
    assert worked_record.is_overtime is True
    assert worked_record.overtime_hours == worked_record.work_hours
    assert worked_record.employee_issues == []
    assert worked_record.statuses == ["work_outside_personnel_schedule"]


def test_personnel_state_is_recalculated_when_actions_change(
    auth_client_factory,
    user_factory,
):
    staff = user_factory(staff=True)
    employee = user_factory()
    _move_auto_hired_before_test_period(employee)
    client = auth_client_factory(staff)
    EmployeeAction.objects.create(
        employee=employee,
        action=ACTION_ON_LEAVE,
        date=_aware_datetime(2026, 4, 20),
    )

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    first_record = AttendanceRecord.objects.get(employee=employee, date="2026-04-21")
    assert first_record.personnel_status == ACTION_ON_LEAVE
    assert first_record.is_absent is False

    EmployeeAction.objects.create(
        employee=employee,
        action=ACTION_RETURNED_FROM_LEAVE,
        date=_aware_datetime(2026, 4, 21),
    )

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        response = client.post(_analyze_url(), _payload(employee.id), format="json")

    assert response.status_code == 200
    updated_record = AttendanceRecord.objects.get(employee=employee, date="2026-04-21")
    assert updated_record.personnel_status == "normal"
    assert updated_record.effective_is_workday is True
    assert updated_record.is_absent is True
    assert updated_record.employee_issues == ["absence"]


def test_attendance_record_comments_use_communications(
    auth_client_factory, user_factory
):
    staff = user_factory(staff=True, first_name="Admin", last_name="User")
    employee = user_factory()
    client = auth_client_factory(staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")

    create_response = client.post(
        _record_comments_url(record.id),
        {"text": "Проверить причину опоздания"},
        format="json",
    )
    assert create_response.status_code == 201
    comment = create_response.json()
    assert comment["record"] == record.id
    assert comment["text"] == "Проверить причину опоздания"
    assert comment["author"]["id"] == staff.id

    list_response = client.get(_record_comments_url(record.id))
    assert list_response.status_code == 200
    assert [item["text"] for item in list_response.json()] == [
        "Проверить причину опоздания"
    ]

    records_response = client.get(_records_url(), {"employee_id": employee.id})
    assert records_response.status_code == 200
    target = next(
        item for item in records_response.json()["results"] if item["id"] == record.id
    )
    assert target["comments_count"] == 1


def test_attendance_record_comment_notifies_employee_and_admins(
    auth_client_factory,
    monkeypatch,
    user_factory,
):
    author = user_factory(staff=True, first_name="Admin", last_name="Author")
    other_staff = user_factory(staff=True)
    superuser = user_factory(superuser=True)
    employee = user_factory()
    inactive_staff = user_factory(staff=True, active=False)
    client = auth_client_factory(author)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    sent_notifications = []

    def capture_notification(**kwargs):
        sent_notifications.append(kwargs)

    monkeypatch.setattr(
        "communications.notifications.handlers._send_notification",
        capture_notification,
    )

    response = client.post(
        _record_comments_url(record.id),
        {"text": "Проверить причину опоздания"},
        format="json",
    )

    assert response.status_code == 201
    recipient_ids = {payload["recipient"].id for payload in sent_notifications}
    assert recipient_ids == {employee.id, other_staff.id, superuser.id}
    assert author.id not in recipient_ids
    assert inactive_staff.id not in recipient_ids
    assert sent_notifications

    for payload in sent_notifications:
        assert payload["verb"] == "commented"
        assert payload["action_url"] == f"/attendance?record={record.id}&comments=1"
        assert payload["target"] == record
        assert payload["data"]["object_type"] == "AttendanceRecord"
        assert payload["data"]["object_id"] == record.id
        assert payload["data"]["attendance_record_id"] == record.id
        assert payload["data"]["employee_id"] == employee.id
        assert payload["data"]["record_date"] == "2026-04-20"


def test_attendance_record_comment_delete_is_author_only(
    auth_client_factory,
    user_factory,
):
    author = user_factory(staff=True)
    other_staff = user_factory(staff=True)
    employee = user_factory()
    author_client = auth_client_factory(author)
    other_client = auth_client_factory(other_staff)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        author_client.post(_analyze_url(), _payload(employee.id), format="json")

    record = AttendanceRecord.objects.get(employee=employee, date="2026-04-20")
    comment_id = author_client.post(
        _record_comments_url(record.id),
        {"text": "Комментарий автора"},
        format="json",
    ).json()["id"]

    forbidden_response = other_client.delete(
        _record_comment_detail_url(record.id, comment_id)
    )
    assert forbidden_response.status_code == 403

    delete_response = author_client.delete(
        _record_comment_detail_url(record.id, comment_id)
    )
    assert delete_response.status_code == 204
    assert author_client.get(_record_comments_url(record.id)).json() == []


def test_user_can_comment_own_attendance_record(auth_client_factory, user_factory):
    user = user_factory(staff=False)
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(user.id), format="json")

    record = AttendanceRecord.objects.get(employee=user, date="2026-04-20")

    create_response = client.post(
        _record_comments_url(record.id),
        {"text": "Мой комментарий"},
        format="json",
    )

    assert create_response.status_code == 201
    assert create_response.json()["author"]["id"] == user.id
    assert client.get(_record_comments_url(record.id)).json()[0]["text"] == (
        "Мой комментарий"
    )


def test_own_attendance_record_comment_notifies_admins_but_not_author(
    auth_client_factory,
    monkeypatch,
    user_factory,
):
    user = user_factory(staff=False)
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(user.id), format="json")

    record = AttendanceRecord.objects.get(employee=user, date="2026-04-20")
    sent_notifications = []

    def capture_notification(**kwargs):
        sent_notifications.append(kwargs)

    monkeypatch.setattr(
        "communications.notifications.handlers._send_notification",
        capture_notification,
    )

    response = client.post(
        _record_comments_url(record.id),
        {"text": "Мой комментарий"},
        format="json",
    )

    assert response.status_code == 201
    recipient_ids = {payload["recipient"].id for payload in sent_notifications}
    assert recipient_ids == {staff.id}
    assert user.id not in recipient_ids
    assert sent_notifications[0]["action_url"] == (
        f"/attendance?record={record.id}&comments=1"
    )


def test_user_cannot_comment_other_employee_attendance_record(
    auth_client_factory,
    user_factory,
):
    user = user_factory(staff=False)
    other = user_factory()
    staff = user_factory(staff=True)
    staff_client = auth_client_factory(staff)
    client = auth_client_factory(user)

    with patch(
        "api.v1.attendance.views.analyze_employee_attendance",
        return_value=_logstorm_result(),
    ):
        staff_client.post(_analyze_url(), _payload(other.id), format="json")

    record = AttendanceRecord.objects.get(employee=other, date="2026-04-20")

    response = client.post(
        _record_comments_url(record.id),
        {"text": "Чужая запись"},
        format="json",
    )

    assert response.status_code == 404
