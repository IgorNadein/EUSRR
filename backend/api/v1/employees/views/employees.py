"""EmployeeViewSet — CRUD сотрудников, профиль, навыки, LDAP-info, Excel-экспорт."""

from __future__ import annotations

import logging
import traceback

from common.emails import send_templated_mail
from django.conf import settings
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Exists, OuterRef, Prefetch, Q, Subquery
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.crypto import get_random_string
from employees.constants import ACTION_DISMISSED
from employees.models import (Department, EmployeeAction, EmployeeDepartment,
                              LdapSyncState, Position, RoleAssignment, Skill)
from employees.utils import _to_bool
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ...permissions import AdminOrActionOrModelPerms, IsSelfOrStaff
from ..serializers import (EmployeeListSerializer, EmployeeSerializer,
                           SkillSerializer)
from ._helpers import Employee

logger = logging.getLogger(__name__)


class EmployeeViewSet(viewsets.ModelViewSet):
    """
    /api/v1/employees/

    Доступ:
      - GET list/retrieve          — только аутентифицированные пользователи.
      - POST (create)              — только staff/superuser.
      - PATCH/PUT/DELETE {id}      — staff/superuser ИЛИ пользователи с модельными правами.
      - GET/PATCH /employees/me/   — только аутентифицированные; PATCH правит профиль текущего пользователя.
      - POST {id}/add_skill|remove_skill — требуется employees.manage_employee_skills.

    Поиск: last_name, first_name, patronymic, email, phone_number

    LDAP синхронизация происходит через Django сигналы (signals_ldap.py).
    """

    serializer_class = EmployeeSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["last_name", "first_name",
                     "patronymic", "email", "phone_number"]
    ordering_fields = ["last_name", "first_name", "created_at", "id"]
    ordering = ["last_name", "first_name"]
    required_perms_by_action = {
        "add_skill": "employees.manage_employee_skills",
        "remove_skill": "employees.manage_employee_skills",
        "ldap_info": "employees.view_ldap_info",
    }

    def get_permissions(self):
        if self.action == "create":
            return [IsAuthenticated(), AdminOrActionOrModelPerms()]
        if self.action in {
            "update",
            "partial_update",
            "destroy",
            "add_skill",
            "remove_skill",
            "ldap_info",
        }:
            return [IsAuthenticated(), (IsSelfOrStaff | AdminOrActionOrModelPerms)()]
        return [IsAuthenticated()]

    def get_queryset(self):
        last_action_code_sq = Subquery(
            EmployeeAction.objects.filter(employee_id=OuterRef("pk"))
            .order_by("-date")
            .values("action")[:1]
        )
        dep_links_prefetch = Prefetch(
            "departments_links",
            queryset=EmployeeDepartment.objects.filter(is_active=True).select_related(
                "department", "role"
            ),
            to_attr="dept_links",
        )

        prefetches = [
            "skills",
            dep_links_prefetch,
            Prefetch("actions", queryset=EmployeeAction.objects.order_by("-date")),
        ]

        qs = (
            Employee.objects.select_related("position")
            .prefetch_related(*prefetches)
            .annotate(last_action_code=last_action_code_sq)
            .order_by(*self.ordering)
        )

        qp = self.request.query_params

        # по отделу
        dep = qp.get("department")
        if dep:
            try:
                dep_id = int(dep)
            except (TypeError, ValueError):
                dep_id = None
            if dep_id:
                member_ids = EmployeeDepartment.objects.filter(
                    department_id=dep_id
                ).values("employee_id")
                head_ids = Department.objects.filter(
                    id=dep_id).values("head_id")
                role_assignment_ids = RoleAssignment.objects.filter(
                    role__department_id=dep_id, is_active=True
                ).values("employee_id")

                qs = qs.filter(
                    Q(id__in=member_ids)
                    | Q(id__in=head_ids)
                    | Q(id__in=role_assignment_ids)
                ).distinct()

                qs = qs.annotate(
                    _is_dept_member=Exists(
                        EmployeeDepartment.objects.filter(
                            employee_id=OuterRef("pk"),
                            department_id=dep_id,
                            is_active=True,
                        )
                    ),
                    _is_dept_head=Exists(
                        Department.objects.filter(
                            id=dep_id, head_id=OuterRef("pk"))
                    ),
                    _has_role_assignment=Exists(
                        RoleAssignment.objects.filter(
                            employee_id=OuterRef("pk"),
                            role__department_id=dep_id,
                            is_active=True,
                        )
                    ),
                )

                self.request._department_filter_id = dep_id

        # по должности
        position = qp.get("position")
        if position:
            qs = qs.filter(position_id=position)

        # по навыкам (any-of)
        skill_ids = qp.getlist("skill")
        if skill_ids:
            qs = qs.filter(skills__in=skill_ids).distinct()

        # статусы
        email_verified = _to_bool(qp.get("email_verified"))
        if email_verified is not None:
            qs = qs.filter(email_verified=email_verified)

        active = _to_bool(qp.get("active"))
        if active is not None:
            qs = qs.filter(is_active=active)

        actually = _to_bool(qp.get("actually_active"))
        if actually is True:
            qs = qs.filter(
                Q(email_verified=True)
                & (
                    Q(last_action_code__isnull=True, is_active=True)
                    | ~Q(last_action_code=ACTION_DISMISSED)
                )
            )
        elif actually is False:
            qs = qs.exclude(
                Q(email_verified=True)
                & (
                    Q(last_action_code__isnull=True, is_active=True)
                    | ~Q(last_action_code=ACTION_DISMISSED)
                )
            )

        created_at_gte = qp.get("created_at__gte")
        if created_at_gte:
            qs = qs.filter(created_at__gte=created_at_gte)

        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return EmployeeListSerializer
        return EmployeeSerializer

    def perform_create(self, serializer):
        """Создание сотрудника (административное, без пароля).
        
        Создание LDAP пользователей с паролем - через RegisterAPIView.
        """
        instance = serializer.save()
        instance.set_unusable_password()
        instance.save(update_fields=["password"])

        # Навыки
        skills_ids = self.request.data.get("skills_ids", [])
        if skills_ids:
            instance.skills.set(skills_ids)

    def perform_update(self, serializer):
        """Обновление сотрудника."""
        instance = serializer.instance
        old_email = instance.email

        # Передаем данные для синхронизации с LDAP (через сигналы)
        instance._ldap_changes = dict(self.request.data)
        if 'avatar' in self.request.FILES:
            instance._ldap_avatar = self.request.FILES['avatar']

        serializer.save()

        # Если email изменился - сбрасываем верификацию
        new_email = instance.email
        if new_email and new_email.lower() != old_email.lower():
            instance.email_verified = False
            instance.email_activation_code = get_random_string(6, "0123456789")
            instance.save(update_fields=["email_verified", "email_activation_code"])

            try:
                send_templated_mail(
                    subject="Подтверждение нового email",
                    to=[instance.email],
                    template_base="emails/registration_verify_code",
                    context={"code": instance.email_activation_code, "user": instance},
                )
            except Exception:
                pass

    @action(detail=False, methods=["get", "patch"])
    def me(self, request):
        """GET — профиль текущего пользователя; PATCH — частичное обновление."""
        instance: Employee = request.user  # type: ignore

        if request.method == "GET":
            instance = (
                Employee.objects.select_related("position")
                .prefetch_related(
                    "skills",
                    Prefetch(
                        "departments_links",
                        queryset=EmployeeDepartment.objects.filter(
                            is_active=True
                        ).select_related("department", "role"),
                        to_attr="dept_links",
                    ),
                    Prefetch(
                        "actions", queryset=EmployeeAction.objects.order_by("-date")
                    ),
                )
                .get(pk=request.user.pk)
            )
            ctx = self.get_serializer_context()
            ctx["include_actions"] = True
            ctx["include_action_history"] = True
            data = self.get_serializer(instance, context=ctx).data
            return Response(data, status=200)

        # PATCH — используем стандартный механизм (миксин управляет LDAP sync)
        serializer = self.get_serializer(
            instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=200)

    @action(detail=True, methods=["post"])
    def add_skill(self, request, pk=None):
        """body: { "skill_id": 3 } ИЛИ { "skill_name": "Python" }"""
        emp = self.get_object()
        sid = request.data.get("skill_id")
        sname = (request.data.get("skill_name") or "").strip()

        sk = None
        if sid:
            sk = Skill.objects.filter(pk=sid).first()
        if not sk and sname:
            sk = Skill.objects.filter(
                name__iexact=sname
            ).first() or Skill.objects.create(name=sname)
        if not sk:
            return Response({"detail": "Навык не найден/не указан"}, status=400)

        emp.skills.add(sk)
        return Response(
            {"ok": True, "skill": {"id": sk.id, "name": sk.name}}, status=200
        )

    @action(detail=True, methods=["post"])
    def remove_skill(self, request, pk=None):
        """body: { "skill_id": 3 } ИЛИ { "skill_name": "Python" }"""
        emp = self.get_object()
        sid = request.data.get("skill_id")
        sname = (request.data.get("skill_name") or "").strip()

        sk = None
        if sid:
            sk = Skill.objects.filter(pk=sid).first()
        if not sk and sname:
            sk = Skill.objects.filter(name__iexact=sname).first()
        if not sk:
            return Response({"detail": "Навык не найден"}, status=404)

        emp.skills.remove(sk)
        return Response(
            {"ok": True, "removed": {"id": sk.id, "name": sk.name}}, status=200
        )

    @action(detail=True, methods=["get"], url_path="ldap-info")
    def ldap_info(self, request, pk=None):
        """GET /api/v1/employees/{id}/ldap-info/ — LDAP информация о сотруднике."""
        from employees.ldap.orm_models import LdapUser

        emp = self.get_object()
        force_refresh = request.query_params.get(
            'force_refresh', '').lower() == 'true'

        if not force_refresh and emp.username:
            return Response(
                {
                    "sAMAccountName": emp.username,
                    "cached": True,
                },
                status=status.HTTP_200_OK,
            )

        try:
            ldap_sync = LdapSyncState.objects.get(
                model='employee',
                object_pk=str(emp.pk)
            )
            if not ldap_sync.ldap_dn:
                return Response(
                    {"detail": "У этого сотрудника нет связи с LDAP"},
                    status=status.HTTP_404_NOT_FOUND,
                )
        except LdapSyncState.DoesNotExist:
            return Response(
                {"detail": "У этого сотрудника нет связи с LDAP"},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            ldap_user = LdapUser.objects.get(dn=ldap_sync.ldap_dn)

            sam_account_name = ldap_user.sam_account_name
            if not sam_account_name:
                return Response(
                    {"detail": "Не удалось получить LDAP информацию"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            emp.username = sam_account_name
            emp.save(update_fields=["username"])

            return Response(
                {
                    "sAMAccountName": sam_account_name,
                    "cached": False,
                },
                status=status.HTTP_200_OK,
            )

        except LdapUser.DoesNotExist:
            return Response(
                {"detail": "Не удалось получить LDAP информацию"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            logger.error(
                f"Error fetching LDAP info for employee {emp.id}: {e}", exc_info=True
            )
            return Response(
                {"detail": f"Ошибка при запросе LDAP информации: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """GET /api/v1/employees/export-excel/ — экспорт в Excel."""
        from datetime import datetime

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        queryset = self.filter_queryset(self.get_queryset())
        queryset = (
            queryset.select_related("position")
            .prefetch_related(
                "skills", "departments_links__department", "departments_links__role"
            )
            .order_by("last_name", "first_name")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        headers = [
            "ID",
            "Фамилия",
            "Имя",
            "Отчество",
            "Email",
            "Телефон",
            "Должность",
            "Отделы",
            "Дата рождения",
            "Дата регистрации",
            "Активен",
            "Email подтвержден",
            "Навыки",
            "Telegram",
            "WhatsApp",
            "WeChat",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_num, emp in enumerate(queryset, 2):
            departments = emp.departments_links.filter(is_active=True)
            dept_names = ", ".join(
                [
                    f"{d.department.name}" +
                    (f" ({d.role.name})" if d.role else "")
                    for d in departments
                ]
            )

            skills = ", ".join([s.name for s in emp.skills.all()])

            def safe_phone_str(phone_field):
                """Безопасная конвертация PhoneNumber в строку"""
                if not phone_field:
                    return ""
                try:
                    from phonenumbers import PhoneNumberFormat, format_number

                    return format_number(phone_field, PhoneNumberFormat.INTERNATIONAL)
                except Exception:
                    try:
                        return str(phone_field)
                    except Exception:
                        return ""

            phone_str = safe_phone_str(emp.phone_number)
            whatsapp_str = safe_phone_str(emp.whatsapp)

            row_data = [
                emp.id,
                emp.last_name or "",
                emp.first_name or "",
                emp.patronymic or "",
                emp.email or "",
                phone_str,
                emp.position.name if emp.position else "",
                dept_names,
                emp.birth_date.strftime("%d.%m.%Y") if emp.birth_date else "",
                emp.created_at.strftime(
                    "%d.%m.%Y %H:%M") if emp.created_at else "",
                "Да" if emp.is_active else "Нет",
                "Да" if emp.email_verified else "Нет",
                skills,
                emp.telegram or "",
                whatsapp_str,
                emp.wechat or "",
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"employees_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        if getattr(self, "action", None) in {"retrieve", "me"}:
            ctx["include_actions"] = True
            ctx["include_action_history"] = True
        return ctx
