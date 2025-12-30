# flake8: noqa
from __future__ import annotations
# backend/api/v1/employees/views.py

import traceback
from datetime import timedelta
from typing import Any, Dict, List, Optional

from common.emails import send_templated_mail
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import FieldError
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import (Case, Count, Exists, F, IntegerField, OuterRef,
                              Prefetch, Q, Subquery, Value, When)
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.crypto import get_random_string
from employees.constants import ACTION_DISMISSED
from employees.ldap.directory_service import (DirectoryDepartmentDTO,
                                              DirectoryService,
                                              DirectoryUserDTO)
from employees.ldap.errors import (DirectoryDbError, DirectoryLdapError,
                                   DirectoryServiceError)
from employees.ldap.infrastructure.connections import _conn
from employees.models import (Department, DepartmentPermission, DepartmentRole,
                              DeptPerm, EmployeeAction, EmployeeDepartment,
                              LdapSyncState, Position, RoleAssignment, Skill)
from employees.utils import (_build_links_for_dept, _detect_phone_field,
                             _ensure_department_permissions,
                             _head_choices_for_dept, _normalize_phone,
                             _perm_choices_synced, _to_bool,
                             _validate_head_active)
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError

from ..permissions import (AdminOrActionOrModelPerms, AdminOrDeptAllowed,
                           IsSelfOrStaff, has_dept_perm)
from .serializers import (AddMemberInput, DepartmentBriefSerializer,
                          DepartmentRoleSerializer, DepartmentSerializer,
                          EmailSerializer, EmailVerifySerializer,
                          EmployeeActionSerializer, EmployeeBriefSerializer,
                          EmployeeListSerializer, EmployeeSerializer,
                          GroupSerializer, PositionSerializer,
                          RegisterSerializer, RemoveMemberInput, SetHeadInput,
                          SetMemberRoleInput, SkillSerializer)

Employee = get_user_model()


PHONE_FIELD = _detect_phone_field()


def _is_ldap_enabled():
    """Проверяет, включена ли интеграция с LDAP."""
    return getattr(settings, "LDAP_ENABLED", False)


def _ldap_try(fn):
    """Выполняет функцию, которая работает с LDAP, и обрабатывает ошибки.
    
    Если LDAP отключен, функция не выполняется и возвращается None.
    """
    if not _is_ldap_enabled():
        return None
    
    try:
        fn()
        return None
    except (DirectoryLdapError, DirectoryServiceError, DirectoryDbError) as e:
        return Response(
            {"detail": f"LDAP sync failed: {e}"}, status=status.HTTP_502_BAD_GATEWAY
        )


def _with_ldap_service(operation_name="LDAP operation"):
    """Декоратор для методов, которые используют DirectoryService.
    
    Пропускает выполнение LDAP-операций если LDAP отключен.
    При ошибках LDAP возвращает Response с кодом 502.
    
    Args:
        operation_name: Название операции для логирования
    """
    def decorator(func):
        def wrapper(*args, **kwargs):
            if not _is_ldap_enabled():
                # LDAP отключен - пропускаем операцию
                return None
            
            try:
                return func(*args, **kwargs)
            except (DirectoryLdapError, DirectoryServiceError, DirectoryDbError) as e:
                return Response(
                    {"detail": f"{operation_name} failed: {e}"},
                    status=status.HTTP_502_BAD_GATEWAY
                )
        return wrapper
    return decorator


class HistoryActionMixin:
    """
    Добавляет GET /{basename}/{pk}/history/
    Параметры (необяз.): ?from=ISO ?to=ISO ?user=<id|email> ?type=+|~|-
    """

    history_diff_fields = None  # список полей для diff; если None — попытаемся угадать

    @action(detail=True, methods=["get"], permission_classes=[IsAuthenticated])
    def history(self, request, pk=None):
        obj = self.get_object()
        qs = obj.history.select_related("history_user").order_by(
            "-history_date", "-history_id"
        )

        q_from = request.query_params.get("from")
        q_to = request.query_params.get("to")
        q_user = request.query_params.get("user")
        q_type = request.query_params.get("type")

        if q_from:
            try:
                qs = qs.filter(history_date__gte=q_from)
            except:
                pass
        if q_to:
            try:
                qs = qs.filter(history_date__lte=q_to)
            except:
                pass
        if q_user:
            qs = qs.filter(
                Q(history_user__id__iexact=q_user)
                | Q(history_user__email__iexact=q_user)
            )
        if q_type in {"+", "~", "-"}:
            qs = qs.filter(history_type=q_type)

        items = list(qs)
        results = []
        for i, cur in enumerate(items):
            prev = items[i + 1] if i + 1 < len(items) else None
            changes = {}
            # какие поля сравнивать
            if self.history_diff_fields is not None:
                fields = self.history_diff_fields
            else:
                # берём только реальные field.name модели (без M2M)
                fields = [f.name for f in obj._meta.fields]

            for name in fields:
                new = getattr(cur, name, None)
                old = getattr(prev, name, None) if prev else None
                if old != new:
                    changes[name] = {"old": old, "new": new}

            results.append(
                {
                    "history_id": cur.history_id,
                    "history_date": cur.history_date,
                    "history_type": cur.history_type,  # "+", "~", "-"
                    "history_user": getattr(cur.history_user, "email", None),
                    "changes": changes,
                }
            )
        return Response({"count": len(results), "results": results})


class ResendEmailAPIView(APIView):
    """POST /api/v1/auth/resend-email/  body: {"email": "..."}"""

    throttle_scope = "anon"
    permission_classes = [AllowAny]

    def post(self, request):
        ser = EmailSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        email = ser.validated_data["email"]

        user = Employee.objects.filter(email__iexact=email).first()
        if not user:
            return Response({"ok": False, "error": "user_not_found"}, status=404)
        if user.email_verified:
            return Response({"ok": False, "error": "already_verified"}, status=400)

        user.email_activation_code = get_random_string(6, "0123456789")
        user.save(update_fields=["email_activation_code"])

        send_templated_mail(
            subject="Подтверждение регистрации",
            to=[user.email],
            template_base="emails/registration_verify_code",
            context={"code": user.email_activation_code, "user": user},
        )
        return Response({"ok": True}, status=200)


class VerifyEmailAPIView(APIView):
    throttle_scope = "anon"
    permission_classes = [AllowAny]

    def post(self, request):
        """Подтверждает email и активирует пользователя.
        
        В режиме с LDAP активирует запись в LDAP.
        В режиме без LDAP просто активирует пользователя в БД.
        """
        ser = EmailVerifySerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        email = ser.validated_data["email"].strip().lower()
        code = ser.validated_data["code"].strip()

        user = Employee.objects.filter(email__iexact=email).first()
        if not user:
            return Response({"ok": False, "error": "user_not_found"}, status=404)
        if not code:
            return Response({"ok": False, "error": "empty_code"}, status=400)

        created = getattr(user, "created_at", None) or getattr(
            user, "date_joined", None
        )
        if created and not user.email_verified:
            if timezone.now() - created > timedelta(minutes=5):
                user.delete()
                return Response({"ok": False, "error": "expired"}, status=400)

        if not user.verify_email(code):
            return Response({"ok": False, "error": "invalid_code"}, status=400)

        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # Режим с LDAP: активируем запись в LDAP
            try:
                from employees.models import LdapSyncState
                svc = DirectoryService()
                
                # Проверяем наличие LDAP-идентификаторов в LdapSyncState
                sync_state = LdapSyncState.objects.filter(
                    model="employee",
                    object_pk=str(user.pk)
                ).first()
                
                has_ldap = sync_state and (sync_state.ldap_dn or sync_state.ldap_guid)
                
                if has_ldap:
                    # Запись существует - активируем через DirectoryService
                    user = svc.update_user(user, {"is_active": True})
                else:
                    # LDAP запись не найдена - активируем только в БД
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(
                        f"User {email} has no LDAP sync state, activating in DB only"
                    )
                    user.is_active = True
                    user.save(update_fields=["is_active"])
                
                # Убеждаемся, что БД тоже активна
                if not user.is_active:
                    user.is_active = True
                    user.save(update_fields=["is_active"])
            except DirectoryLdapError as e:
                # Если LDAP недоступен, всё равно активируем в БД
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    f"LDAP error during activation for {email}: {e}, activating in DB"
                )
                user.is_active = True
                user.save(update_fields=["is_active"])
            except DirectoryDbError as e:
                return Response(
                    {"ok": False, "error": "db_error", "detail": str(e)}, status=500
                )
            except DirectoryServiceError as e:
                # При ошибке сервиса тоже активируем в БД
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    f"Service error during activation for {email}: {e}, activating in DB"
                )
                user.is_active = True
                user.save(update_fields=["is_active"])
        else:
            # Режим без LDAP: просто активируем пользователя в БД
            if not user.is_active:
                user.is_active = True
                user.save(update_fields=["is_active"])

        return Response({"ok": True, "user_id": user.id}, status=200)


class RegisterAPIView(APIView):
    """Регистрация: создаём учётку в LDAP (disabled) с паролем, в БД — set_unusable_password."""

    throttle_scope = "anon"
    permission_classes = [AllowAny]
    parser_classes = (JSONParser, FormParser, MultiPartParser)

    @transaction.atomic
    def post(self, request):
        import logging
        logger = logging.getLogger(__name__)
        
        # Логируем входящие данные для диагностики
        logger.warning(f"[REGISTER] Received data: {request.data}")
        logger.warning(f"[REGISTER] Content-Type: {request.content_type}")
        
        # 0) хотя бы один контакт
        if not (
            request.data.get("telegram")
            or request.data.get("whatsapp")
            or request.data.get("wechat")
        ):
            logger.warning("[REGISTER] No contact provided")
            return Response(
                {
                    "detail": "Заполните хотя бы одно из полей: WhatsApp, WeChat или Telegram"
                },
                status=400,
            )

        ser = RegisterSerializer(data=request.data)
        if not ser.is_valid():
            logger.warning(f"[REGISTER] Validation errors: {ser.errors}")
        ser.is_valid(raise_exception=True)
        v = ser.validated_data

        email = v["email"].strip().lower()
        password = v["password"]
        phone_norm = _normalize_phone(
            v.get("phone_number") or request.data.get("phone")
        )
        if not phone_norm:
            return Response({"ok": False, "error": "invalid_phone"}, status=400)

        existing_phone = Employee.objects.filter(phone_number=phone_norm).first()
        if existing_phone:
            logger.warning(
                "[REGISTER] Phone %s already used by user id=%s",
                phone_norm,
                existing_phone.id,
            )
            return Response(
                {
                    "ok": False,
                    "error": "phone_taken",
                    "detail": "Номер телефона уже зарегистрирован."
                    " Войдите в существующий аккаунт или используйте другой номер.",
                    "phone_number": [
                        "Этот номер телефона уже привязан к другому аккаунту."
                    ],
                },
                status=400,
            )

        user = Employee.objects.filter(email__iexact=email).first()
        if user:
            if user.email_verified:
                # Email уже верифицирован - нельзя регистрироваться
                return Response({"ok": False, "error": "email_taken"}, status=400)
            else:
                # Есть неверифицированный пользователь - повторная отправка кода
                return Response(
                    {"ok": True, "pending_verification": True, "user_id": user.id},
                    status=200,
                )

        avatar_file = v.get("avatar") or getattr(request, "FILES", {}).get("avatar")
        avatar_bytes = None
        avatar_name = None
        if avatar_file:
            try:
                avatar_bytes = (
                    avatar_file.read() if hasattr(avatar_file, "read") else None
                )
                avatar_name = getattr(avatar_file, "name", None) or "avatar.jpg"
            except Exception:
                avatar_bytes = None
                avatar_name = None

        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # Режим с LDAP: создаём disabled учётку в LDAP + пароль
            svc = DirectoryService()
            dto = DirectoryUserDTO(
                first_name=v["first_name"],
                last_name=v["last_name"],
                email=email,
                phone_e164=phone_norm,
                department_dn=None,
                group_cns=[],
                initial_password=password,  # пароль идёт только в LDAP
                avatar_bytes=avatar_bytes,
                is_active=False,  # disabled до верификации
            )
            try:
                emp = svc.create_user(dto)
            except DirectoryLdapError as e:
                return Response(
                    {"ok": False, "error": "ldap_error", "detail": str(e)}, status=502
                )
            except DirectoryDbError as e:
                return Response(
                    {"ok": False, "error": "db_error", "detail": str(e)}, status=500
                )
        else:
            # Режим без LDAP: создаём пользователя напрямую в БД
            emp = Employee.objects.create(
                first_name=v["first_name"],
                last_name=v["last_name"],
                email=email,
                phone_number=phone_norm,
                is_active=False,  # не активен до верификации email
                is_ldap_managed=False,
            )
            # Устанавливаем пароль в БД
            emp.set_password(password)

        # 2) Заполняем доп.поля БД
        if avatar_bytes:
            try:
                emp.avatar.save(avatar_name, ContentFile(avatar_bytes), save=False)
            except Exception:
                pass
        
        emp.telegram = v.get("telegram", "")
        emp.whatsapp = v.get("whatsapp", "")
        emp.wechat = v.get("wechat", "")
        emp.birth_date = v["birth_date"]
        emp.gender = v["gender"]  # обязательное поле
        
        if v.get("patronymic"):
            emp.patronymic = v["patronymic"]
        
        pos_id = v.get("position")
        if pos_id and Position.objects.filter(pk=pos_id).exists():
            emp.position_id = pos_id
        
        emp.save()
        
        skills_ids = v.get("skills") or []
        if skills_ids:
            emp.skills.set(Skill.objects.filter(pk__in=skills_ids))

        # 3) Отправляем код верификации
        emp.email_activation_code = get_random_string(6, "0123456789")
        emp.save(update_fields=["email_activation_code"])
        send_templated_mail(
            subject="Подтверждение регистрации",
            to=[emp.email],
            template_base="emails/registration_verify_code",
            context={"code": emp.email_activation_code, "user": emp},
        )

        return Response(
            {
                "id": emp.id,
                "email": emp.email,
                "email_verified": emp.email_verified,
                "is_active": emp.is_active,
            },
            status=status.HTTP_201_CREATED,
        )


class DepartmentViewSet(viewsets.ModelViewSet):
    """
    CRUD отделов + действия:
      - POST /departments/{id}/set_head
      - POST /departments/{id}/set_member_role
      - POST /departments/{id}/add_member
      - POST /departments/{id}/remove_member
    Права:
      - update/partial_update/destroy → manage_department
      - set_head                   → change_department_head
      - set_member_role           → assign_department_role     (назначение/снятие роли)
      - add_member/remove_member  → manage_department          (управление участниками)
      - create                    → staff/superuser
      - чтение                    → аутентифицированным
    """

    queryset = Department.objects.select_related("head").prefetch_related("roles").all()
    serializer_class = DepartmentSerializer

    # пермишены (скоуп-право по отделу)
    class ManagePerm(AdminOrDeptAllowed):
        """Право на общее управление отделом."""

        required_code = DeptPerm.MANAGE

    class ChangeHeadPerm(AdminOrDeptAllowed):
        """Право на смену руководителя отдела."""

        required_code = DeptPerm.CHANGE_HEAD

    class AssignRolePerm(AdminOrDeptAllowed):
        """Право на назначение ролей участникам отдела."""

        required_code = DeptPerm.ASSIGN_ROLE

    def get_permissions(self):
        if self.action in {"update", "partial_update", "destroy"}:
            return [self.ManagePerm()]
        if self.action == "set_head":
            return [self.ChangeHeadPerm()]
        if self.action in {"set_member_role"}:
            return [self.AssignRolePerm()]
        if self.action in {"add_member", "remove_member"}:
            return [self.ManagePerm()]
        if self.action == "create":
            return [AdminOrActionOrModelPerms()]
        if self.action in {
            "members",
            "user_perms",
            "ui_context",
            "list",
            "retrieve",
            "my_departments",
        }:
            return [IsAuthenticated()]
        return [AdminOrActionOrModelPerms()]

    # --- поиск и сортировка, как в старой реализации/тестах ---
    ordering = ["name"]
    ordering_fields = ["name", "id"]

    def get_queryset(self):
        qs = super().get_queryset()

        # ---------- аннотации для employees_count ----------
        # Активные линкы данного отдела
        active_links = EmployeeDepartment.objects.filter(
            department_id=OuterRef("pk"), is_active=True
        )

        # Подсчёт distinct сотрудников по активным линкам
        active_count_subq = (
            active_links.values("department_id")
            .annotate(c=Count("employee_id", distinct=True))
            .values("c")[:1]
        )

        qs = qs.annotate(
            active_count=Coalesce(
                Subquery(active_count_subq, output_field=IntegerField()),
                Value(0),
            ),
            head_in_active=Exists(active_links.filter(employee_id=OuterRef("head_id"))),
        ).annotate(
            employees_count=F("active_count")
            + Case(
                When(
                    Q(head_id__isnull=False) & Q(head_in_active=False),
                    then=Value(1),
                ),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
        # ---------- /аннотации ----------

        # Поиск и сортировка — как и раньше
        search = (
            self.request.query_params.get("search")
            or self.request.query_params.get("q")
            or ""
        ).strip()
        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(description__icontains=search)
                | Q(head__first_name__icontains=search)
                | Q(head__last_name__icontains=search)
                | Q(head__patronymic__icontains=search)
            ).distinct()

        ordering = self.request.query_params.get("ordering")
        if ordering in {"name", "-name", "id", "-id"}:
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by("name")
        return qs

    # --- частичное изменение: изменение head требует отдельного кода ---

    def partial_update(self, request, *args, **kwargs) -> Response:
        """Частичный апдейт отдела через сервисный слой (LDAP → DB).

        Логика:
        • Если меняется head — сперва права + валидация, затем DirectoryService.set_head().
        • Поля name/description — через DirectoryService.update_department().
        • Не вызываем super().partial_update() во избежание двойной записи.

        Args:
            request: DRF Request.
            *args: Прочие позиционные аргументы.
            **kwargs: Прочие именованные аргументы.

        Returns:
            Response: Сериализованные данные отдела после применения изменений.

        Raises:
            (внутренне обрабатываются и мапятся в HTTP-коды)
        """
        instance = self.get_object()
        data: Dict[str, Any] = request.data
        ldap_enabled = _is_ldap_enabled()
        svc = DirectoryService() if ldap_enabled else None
        print(
            "[DEBUG:partial_update] start",
            {
                "dept_id": instance.id,
                "dept_name": instance.name,
                "current_head_id": instance.head_id,
                "raw_keys": list(data.keys()),
                "raw_payload": dict(data),
                "ldap_enabled": ldap_enabled,
            },
        )

        # --- HEAD ---
        if any(k in data for k in ("head", "head_id")):
            raw_desired = data.get("head_id", data.get("head", None))

            try:
                if raw_desired is None:
                    pass
                if isinstance(raw_desired, str) and raw_desired.strip().lower() in {
                    "",
                    "null",
                    "none",
                }:
                    pass
                try:
                    desired_head_id = int(raw_desired)
                except (TypeError, ValueError):
                    raise ValueError("head_id должен быть целым числом или null")
            except ValueError as e:
                print("[ERROR:partial_update] parse head_id failed:", repr(e))
                return Response(
                    {"head_id": [str(e)]}, status=status.HTTP_400_BAD_REQUEST
                )
            print(
                "[DEBUG:partial_update] head branch",
                {
                    "desired_head_id": desired_head_id,
                    "current_head_id": instance.head_id,
                },
            )

            # --- права ---
            perm = self.ChangeHeadPerm()
            has_perm = perm.has_permission(
                request, self
            ) and perm.has_object_permission(request, self, instance)
            print(
                "[DEBUG:partial_update] perm check:",
                has_perm,
                "user_id=",
                getattr(request.user, "id", None),
            )
            if not has_perm:
                print("[ERROR:partial_update] forbidden head change")
                return Response(
                    {"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN
                )

            # Если указали того же руководителя — пропустим лишнюю работу
            if (desired_head_id or None) == (instance.head_id or None):
                pass  # идем дальше — возможно пришли name/description
            else:
                # --- валидация кандидата ---
                new_head = None
                if desired_head_id is not None:
                    employee_model = (
                        type(instance)._meta.get_field("head").remote_field.model
                    )
                    new_head = get_object_or_404(employee_model, id=desired_head_id)

                    # Требовать verified, если новый head — не сам текущий head-пользователь
                    require_verified = not (
                        instance.head_id
                        and instance.head_id == getattr(request.user, "id", None)
                    )
                    ok, errs = _validate_head_active(
                        instance,
                        desired_head_id,
                        require_email_verified=require_verified,
                    )
                    print(
                        "[DEBUG:partial_update] head candidate validation:",
                        ok,
                        errs if not ok else None,
                    )
                    if not ok:
                        return Response(errs, status=status.HTTP_400_BAD_REQUEST)

                # --- LDAP → DB ---
                if ldap_enabled:
                    # Режим с LDAP: синхронизируем через DirectoryService
                    try:
                        print(
                            "[DEBUG:partial_update] calling DirectoryService.set_head",
                            {
                                "dept_id": instance.id,
                                "new_head_id": getattr(new_head, "id", None),
                            },
                        )
                        instance = svc.set_head(instance, new_head)
                        print(
                            "[DEBUG:partial_update] set_head OK",
                            {
                                "result_head_id": instance.head_id,
                                "head_appointed_at": getattr(
                                    instance, "head_appointed_at", None
                                ),
                            },
                        )
                    except (
                        DirectoryLdapError,
                        DirectoryDbError,
                        DirectoryServiceError,
                    ) as e:
                        code = (
                            status.HTTP_502_BAD_GATEWAY
                            if isinstance(e, DirectoryLdapError)
                            else (
                                status.HTTP_400_BAD_REQUEST
                                if isinstance(e, DirectoryServiceError)
                                else status.HTTP_500_INTERNAL_SERVER_ERROR
                            )
                        )
                        print(
                            "[ERROR:partial_update] set_head FAILED:",
                            {
                                "exc_type": type(e).__name__,
                                "status": code,
                                "message": str(e),
                                "trace": traceback.format_exc(),
                            },
                        )
                        return Response({"detail": str(e)}, status=code)
                else:
                    # Режим без LDAP: обновляем напрямую в БД
                    instance.head = new_head
                    if new_head:
                        instance.head_appointed_at = timezone.now()
                    else:
                        instance.head_appointed_at = None
                    instance.save(update_fields=["head", "head_appointed_at"])
                    print("[DEBUG:partial_update] set_head OK (no LDAP)")

        # --- NAME / DESCRIPTION ---
        changes: Dict[str, Any] = {}
        for k in ("name", "description"):
            if k in data:
                changes[k] = data.get(k)

        print("[DEBUG:partial_update] changes branch:", changes)
        if changes:
            if ldap_enabled:
                # Режим с LDAP: синхронизируем через DirectoryService
                try:
                    print(
                        "[DEBUG:partial_update] calling DirectoryService.update_department",
                        {
                            "dept_id": instance.id,
                            "changes": changes,
                        },
                    )
                    instance = svc.update_department(instance, changes)
                    print("[DEBUG:partial_update] update_department OK")
                except DirectoryLdapError as e:
                    print(
                        "[ERROR:partial_update] update_department LDAP FAILED:",
                        {"message": str(e), "trace": traceback.format_exc()},
                    )
                    return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)
                except DirectoryDbError as e:
                    print(
                        "[ERROR:partial_update] update_department DB FAILED:",
                        {"message": str(e), "trace": traceback.format_exc()},
                    )
                    return Response(
                        {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            else:
                # Режим без LDAP: обновляем напрямую в БД
                for k, v in changes.items():
                    setattr(instance, k, v)
                instance.save(update_fields=list(changes.keys()))
                print("[DEBUG:partial_update] update_department OK (no LDAP)")
        resp = self.get_serializer(instance).data
        print("[DEBUG:partial_update] done OK, response keys:", list(resp.keys()))
        return Response(self.get_serializer(instance).data, status=status.HTTP_200_OK)

    # -------- actions --------

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def set_head(self, request, pk: str | None = None) -> Response:
        """Назначение/снятие руководителя отдела через сервисный слой (LDAP → DB).

        Логика:
          • Проверка прав через ChangeHeadPerm.
          • Валидация кандидата (email_verified ослабляется для текущего head).
          • Применение через DirectoryService.set_head().

        Args:
            request: DRF Request.
            pk: Идентификатор отдела из URL.

        Returns:
            Response: Данные отдела после операции.

        Raises:
            ValidationError: При невалидном вводе (поднимается сериализатором).
            DirectoryServiceError: Бизнес-ошибка сервиса (400).
            DirectoryLdapError: Ошибка взаимодействия с LDAP (502).
            DirectoryDbError: Ошибка записи в БД (500).
        """
        print("[DEBUG:set_head] start pk=", pk, "raw_data=", request.data)
        dept = self.get_object()
        print(
            "[DEBUG:set_head] dept:",
            {"id": dept.id, "name": dept.name, "current_head_id": dept.head_id},
        )

        payload = SetHeadInput(data=request.data)
        payload.is_valid(raise_exception=True)
        head_id = payload.validated_data.get("head_id")
        print("[DEBUG:set_head] validated head_id=", head_id)

        employee_model = Department._meta.get_field("head").remote_field.model
        new_head = (
            get_object_or_404(employee_model, id=head_id)
            if head_id is not None
            else None
        )
        print(
            "[DEBUG:set_head] resolved new_head:",
            (
                None
                if new_head is None
                else {"id": new_head.id, "email": getattr(new_head, "email", None)}
            ),
        )

        # --- права ---
        print("[DEBUG:set_head] checking permissions via ChangeHeadPerm")
        perm = self.ChangeHeadPerm()
        has_perm = perm.has_permission(request, self) and perm.has_object_permission(
            request, self, dept
        )
        print(
            "[DEBUG:set_head] permissions:",
            has_perm,
            "user_id=",
            getattr(request.user, "id", None),
        )
        if not has_perm:
            print("[ERROR:set_head] permission denied")
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

        # --- валидация кандидата ---
        if head_id is not None:
            require_verified = not (
                dept.head_id and dept.head_id == getattr(request.user, "id", None)
            )
            print(
                "[DEBUG:set_head] validating candidate:",
                {"candidate_id": head_id, "require_email_verified": require_verified},
            )
            ok, errs = _validate_head_active(
                dept, head_id, require_email_verified=require_verified
            )
            print(
                "[DEBUG:set_head] validation result:",
                ok,
                "errors:",
                errs if not ok else None,
            )
            if not ok:
                print("[ERROR:set_head] candidate validation failed")
                return Response(errs, status=status.HTTP_400_BAD_REQUEST)

        # --- применение ---
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: применяем через DirectoryService
            svc = DirectoryService()
            try:
                print(
                    "[DEBUG:set_head] applying via DirectoryService.set_head:",
                    {
                        "dept_id": dept.id,
                        "old_head_id": dept.head_id,
                        "new_head_id": getattr(new_head, "id", None),
                    },
                )
                dept = svc.set_head(dept, new_head)
                print(
                    "[DEBUG:set_head] apply OK:",
                    {
                        "dept_id": dept.id,
                        "result_head_id": dept.head_id,
                        "head_appointed_at": getattr(dept, "head_appointed_at", None),
                    },
                )
            except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                status_code = (
                    status.HTTP_502_BAD_GATEWAY
                    if isinstance(e, DirectoryLdapError)
                    else (
                        status.HTTP_400_BAD_REQUEST
                        if isinstance(e, DirectoryServiceError)
                        else status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                )
                print(
                    "[ERROR:set_head] apply FAILED:",
                    {"type": type(e).__name__, "status": status_code, "message": str(e)},
                )
                return Response({"detail": str(e)}, status=status_code)
        else:
            # Режим без LDAP: обновляем напрямую в БД
            dept.head = new_head
            if new_head:
                dept.head_appointed_at = timezone.now()
            else:
                dept.head_appointed_at = None
            dept.save(update_fields=["head", "head_appointed_at"])
            print("[DEBUG:set_head] apply OK (no LDAP)")

        resp = self.get_serializer(dept).data
        print(
            "[DEBUG:set_head] response payload prepared:", {"keys": list(resp.keys())}
        )
        return Response(resp, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def set_member_role(self, request, pk=None):
        """
        Назначает/снимает РОЛЬ сотруднику в контексте отдела.
        
        Если сотрудник — член отдела, обновляет EmployeeDepartment.role.
        Если нет — создаёт/обновляет RoleAssignment (новая логика: роли можно
        назначать любому сотруднику компании).

        Тело:
            { "employee_id": <int>, "role_id": <int|null> }
        Права:
            AssignRolePerm (DeptPerm.ASSIGN_ROLE)
        Ответ:
            200 {"employee_id":..., "role_id": <int|null>, "is_active": <bool>, "via_assignment": <bool>}
            400 если роль не принадлежит отделу
        """
        dept = self.get_object()
        payload = SetMemberRoleInput(data=request.data)
        if not payload.is_valid():
            return Response(payload.errors, status=400)

        emp_id = payload.validated_data["employee_id"]
        role_id = payload.validated_data.get("role_id")

        # проверяем роль, если указана, что она принадлежит этому отделу
        employee_model = Department._meta.get_field("head").remote_field.model
        employee = get_object_or_404(employee_model, id=emp_id)
        role = None
        if role_id is not None:
            role = get_object_or_404(DepartmentRole, id=role_id)
            if role.department_id != dept.id:
                return Response(
                    {"role_id": ["Role does not belong to this department."]},
                    status=400,
                )

        # Проверяем, является ли сотрудник членом отдела
        link = EmployeeDepartment.objects.filter(
            employee_id=emp_id, department_id=dept.id
        ).first()
        
        via_assignment = False
        
        if link:
            # Сотрудник — член отдела: обновляем роль в линке (старая логика)
            svc = DirectoryService() if _is_ldap_enabled() else None
            
            if svc:
                try:
                    svc.set_member_role(dept, employee, role)
                except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                    return Response(
                        {"detail": str(e)},
                        status=(
                            502
                            if isinstance(e, DirectoryLdapError)
                            else 400 if isinstance(e, DirectoryServiceError) else 500
                        ),
                    )
            else:
                link.role = role
                link.save(update_fields=["role"])
            
            # Также создаём/обновляем RoleAssignment для консистентности
            if role:
                RoleAssignment.objects.update_or_create(
                    employee_id=emp_id,
                    role=role,
                    defaults={"is_active": True, "assigned_by": request.user}
                )
            else:
                # Снятие роли — деактивируем все назначения этой роли для сотрудника
                RoleAssignment.objects.filter(
                    employee_id=emp_id,
                    role__department=dept,
                    is_active=True
                ).update(is_active=False)
        else:
            # Сотрудник НЕ член отдела: используем только RoleAssignment
            via_assignment = True
            
            if role:
                # Назначаем роль через RoleAssignment
                from employees.ldap.services.department_service import DepartmentService
                from employees.ldap.services.group_service import GroupService
                from employees.ldap.services.user_service import UserService
                
                if _is_ldap_enabled():
                    try:
                        group_service = GroupService()
                        user_service = UserService(group_service)
                        dept_service = DepartmentService(group_service, user_service)
                        dept_service.assign_role(employee, role, request.user)
                    except Exception as e:
                        return Response({"detail": str(e)}, status=400)
                else:
                    RoleAssignment.objects.update_or_create(
                        employee=employee,
                        role=role,
                        defaults={"is_active": True, "assigned_by": request.user}
                    )
            else:
                # Снятие всех ролей отдела для этого сотрудника
                RoleAssignment.objects.filter(
                    employee_id=emp_id,
                    role__department=dept,
                    is_active=True
                ).update(is_active=False)

        return Response(
            {
                "employee_id": emp_id,
                "role_id": (role.id if role else None),
                "is_active": link.is_active if link else True,
                "via_assignment": via_assignment,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["get"], url_path="members")
    def members(self, request, pk=None):
        """
        GET /api/v1/departments/{id}/members/
        Список участников отдела в формате, удобном для фронта-шаблона.
        """
        dept = self.get_object()
        links = _build_links_for_dept(dept, EmployeeBriefSerializer)
        return Response({"count": len(links), "results": links}, status=200)

    @action(detail=True, methods=["get"], url_path="user-perms")
    def user_perms(self, request, pk=None):
        """
        GET /api/v1/departments/{id}/user-perms/
        Флаги прав текущего пользователя в рамках отдела.
        """
        dept = self.get_object()
        uid = getattr(request.user, "id", None)
        data = {
            "is_head": bool(uid and uid == dept.head_id),
            "can_manage": has_dept_perm(request.user, dept.id, DeptPerm.MANAGE),
            "can_change_head": has_dept_perm(
                request.user, dept.id, DeptPerm.CHANGE_HEAD
            ),
            "can_assign_roles": has_dept_perm(
                request.user, dept.id, DeptPerm.ASSIGN_ROLE
            ),
        }
        return Response(data, status=200)

    @action(detail=True, methods=["get"], url_path="ui-context")
    def ui_context(self, request, pk=None):
        """
        GET /api/v1/departments/{id}/ui-context/
        BFF-агрегатор для страницы отдела: возвращает все необходимые данные одной пачкой.
        Поля:
          - dept: DepartmentSerializer
          - roles: DepartmentRoleSerializer[]
          - links: см. /members
          - head_choices: [{id,name}]
          - dept_perm_choices: [{id,code,name}]
          - user_perms: {is_head, can_manage, can_change_head, can_assign_roles}
        """
        dept = self.get_object()
        dept_data = self.get_serializer(dept).data

        roles_qs = (
            DepartmentRole.objects.filter(department_id=dept.id)
            .prefetch_related("scoped_permissions")
            .order_by("name", "id")
        )
        roles_data = DepartmentRoleSerializer(roles_qs, many=True).data

        links = _build_links_for_dept(dept, EmployeeBriefSerializer)
        perm_choices = _perm_choices_synced()
        head_choices = _head_choices_for_dept(dept, EmployeeBriefSerializer)
        user_perms = {
            "is_head": (
                request.user.id == dept.head_id
                if getattr(request.user, "id", None)
                else False
            ),
            "can_manage": has_dept_perm(request.user, dept.id, DeptPerm.MANAGE),
            "can_change_head": has_dept_perm(
                request.user, dept.id, DeptPerm.CHANGE_HEAD
            ),
            "can_assign_roles": has_dept_perm(
                request.user, dept.id, DeptPerm.ASSIGN_ROLE
            ),
        }

        payload = {
            "dept": dept_data,
            "roles": roles_data,
            "links": links,
            "head_choices": head_choices,
            "dept_perm_choices": perm_choices,
            "user_perms": user_perms,
        }
        return Response(payload, status=200)

    @action(detail=True, methods=["post"], url_path="add_member")
    @transaction.atomic
    def add_member(self, request, pk: int | None = None):
        """Добавляет сотрудника в отдел: MOVE в OU → активирует линк (LDAP → DB)."""
        dept = self.get_object()
        payload = AddMemberInput(data=request.data)
        payload.is_valid(raise_exception=True)

        emp_id = payload.validated_data["employee_id"]

        employee_model = Department._meta.get_field("head").remote_field.model
        employee = get_object_or_404(employee_model, id=emp_id)

        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: добавляем через DirectoryService
            svc = DirectoryService()
            try:
                svc.add_member(dept, employee)
                link = (
                    EmployeeDepartment.objects.filter(
                        employee_id=emp_id, department_id=dept.id
                    )
                    .only("role_id", "is_active")
                    .first()
                )

                return Response(
                    {
                        "employee_id": emp_id,
                        "is_active": True,
                        "role_id": getattr(link, "role_id", None) if link else None,
                    },
                    status=status.HTTP_200_OK,
                )
            except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                return Response(
                    {"detail": str(e)},
                    status=(
                        status.HTTP_502_BAD_GATEWAY
                        if isinstance(e, DirectoryLdapError)
                        else (
                            status.HTTP_400_BAD_REQUEST
                            if isinstance(e, DirectoryServiceError)
                            else status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                    ),
                )
        else:
            # Режим без LDAP: создаём/активируем линк напрямую
            link, created = EmployeeDepartment.objects.get_or_create(
                employee_id=emp_id,
                department_id=dept.id,
                defaults={"is_active": True}
            )
            if not created and not link.is_active:
                link.is_active = True
                link.save(update_fields=["is_active"])
            
            return Response(
                {
                    "employee_id": emp_id,
                    "is_active": True,
                    "role_id": link.role_id,
                },
                status=status.HTTP_200_OK,
            )

    @action(detail=True, methods=["post"], url_path="remove_member")
    @transaction.atomic
    def remove_member(self, request, pk: int | None = None):
        """Удаляет члена отдела: MOVE в Users OU → удаляет линк (LDAP → DB)."""
        dept = self.get_object()

        payload = RemoveMemberInput(data=request.data)
        payload.is_valid(raise_exception=True)
        emp_id: int = payload.validated_data["employee_id"]

        employee_model = Department._meta.get_field("head").remote_field.model
        employee = get_object_or_404(employee_model, id=emp_id)
        # защита от удаления руководителя — ДО вызова сервиса
        if dept.head_id == emp_id:
            return Response(
                {"detail": "Нельзя удалить руководителя отдела."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: удаляем через DirectoryService
            svc = DirectoryService()
            try:
                svc.remove_member(dept, employee)
                return Response(
                    {"employee_id": emp_id, "removed": True},
                    status=status.HTTP_200_OK,
                )
            except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                return Response(
                    {"detail": str(e)},
                    status=(
                        502
                        if isinstance(e, DirectoryLdapError)
                        else 400 if isinstance(e, DirectoryServiceError) else 500
                    ),
                )
        else:
            # Режим без LDAP: деактивируем линк напрямую
            try:
                link = EmployeeDepartment.objects.get(
                    employee_id=emp_id,
                    department_id=dept.id
                )
                link.is_active = False
                link.save(update_fields=["is_active"])
                return Response(
                    {"employee_id": emp_id, "removed": True},
                    status=status.HTTP_200_OK,
                )
            except EmployeeDepartment.DoesNotExist:
                return Response(
                    {"detail": "Employee is not a member of this department."},
                    status=status.HTTP_404_NOT_FOUND,
                )

    @action(detail=False, methods=["get"], url_path="my-departments")
    def my_departments(self, request) -> Response:
        """Вернёт список отделов, доступных текущему пользователю.

        Логика:
            Только отделы, где пользователь является руководителем
              или имеет активную связь в ``EmployeeDepartment`` (is_active=True).

        Args:
            request: Текущий HTTP-запрос DRF.

        Returns:
            Response: JSON-массив отделов в формате ``DepartmentSerializer`` (many=True),
            отсортированный по названию и id.

        Raises:
            NotAuthenticated: Если пользователь не авторизован (обрабатывается классом
                разрешений в ``get_permissions``).
        """
        user = request.user
        qs = self.get_queryset()

        active_link_exists = EmployeeDepartment.objects.filter(
            department_id=OuterRef("pk"),
            employee_id=user.id,
            is_active=True,
        )
        user_qs = qs.filter(Q(head_id=user.id) | Exists(active_link_exists)).distinct()

        user_qs = user_qs.order_by("name", "id")
        data = DepartmentBriefSerializer(user_qs, many=True).data
        return Response(data)

    def create(self, request, *args, **kwargs):
        """Создание отдела: сначала LDAP OU → затем запись Department (если LDAP включен)."""
        print("[DEBUG] Начало создания отдела. Данные запроса:", request.data)
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        head = None
        head_id = ser.validated_data.get("head") or ser.validated_data.get("head_id")
        if head_id:
            print("[DEBUG] Найден head_id:", head_id)
            employee_model = Department._meta.get_field("head").remote_field.model
            head = get_object_or_404(employee_model, id=head_id)

        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: создаём через DirectoryService
            dto = DirectoryDepartmentDTO(
                name=ser.validated_data["name"],
                description=ser.validated_data.get("description", ""),
                head=head,
            )
            svc = DirectoryService()
            try:
                print("[DEBUG] Создание отдела в DirectoryService:", dto)
                dept = svc.create_department(dto)
                print("[DEBUG] Отдел успешно создан:", dept)
                return Response(
                    self.get_serializer(dept).data, status=status.HTTP_201_CREATED
                )
            except DirectoryLdapError as e:
                print("[ERROR] Ошибка LDAP при создании отдела:", str(e))
                return Response({"detail": str(e)}, status=502)
            except DirectoryDbError as e:
                print("[ERROR] Ошибка базы данных при создании отдела:", str(e))
                return Response({"detail": str(e)}, status=500)
        else:
            # Режим без LDAP: создаём напрямую в БД
            dept = Department.objects.create(
                name=ser.validated_data["name"],
                description=ser.validated_data.get("description", ""),
                head=head,
            )
            print("[DEBUG] Отдел создан в БД (без LDAP):", dept)
            return Response(
                self.get_serializer(dept).data, status=status.HTTP_201_CREATED
            )

    def destroy(self, request, *args, **kwargs):
        """Удаляет отдел: сначала в LDAP → затем из БД (если LDAP включен)."""
        print("[DEBUG] Начало удаления отдела.")
        dept = self.get_object()
        print("[DEBUG] Удаляемый отдел:", dept)
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: удаляем через DirectoryService
            svc = DirectoryService()
            try:
                print("[DEBUG] Удаление отдела в DirectoryService.")
                svc.delete_department(dept)
                print("[DEBUG] Отдел успешно удалён.")
                return Response(status=204)
            except DirectoryLdapError as e:
                print("[ERROR] Ошибка LDAP при удалении отдела:", str(e))
                return Response({"detail": str(e)}, status=502)
            except DirectoryDbError as e:
                print("[ERROR] Ошибка базы данных при удалении отдела:", str(e))
                return Response({"detail": str(e)}, status=500)
        else:
            # Режим без LDAP: удаляем напрямую из БД
            dept.delete()
            print("[DEBUG] Отдел удалён из БД (без LDAP).")
            return Response(status=204)


class EmployeeViewSet(viewsets.ModelViewSet):
    """
    /api/v1/employees/

    Доступ:
      - GET list/retrieve          — только аутентифицированные пользователи.
      - POST (create)              — только staff/superuser (в методе create есть явная проверка).
      - PATCH/PUT/DELETE {id}      — staff/superuser ИЛИ пользователи с модельными правами
                                     на Employee (через AdminOrActionOrModelPerms).
      - GET/PATCH /employees/me/   — только аутентифицированные; PATCH правит профиль текущего пользователя.
      - POST {id}/add_skill|remove_skill — требуется спец-пермишен
                                     "employees.manage_employee_skills" ИЛИ staff/superuser
                                     (через AdminOrActionOrModelPerms).

    Поиск: last_name, first_name, patronymic, email, phone_number

    Фильтры:
      ?department=<id>            — участники отдела + head
      ?position=<id>              — по должности
      ?skill=<id>&skill=<id>      — любые из навыков (OR)
      ?email_verified=true|false
      ?active=true|false
      ?actually_active=true|false — реально активные: email_verified=True и не DISMISSED
                                    (или нет действий, но is_active=True)
      ?created_at__gte=<iso_date> — зарегистрированные после указанной даты

    Сортировка:
      ?ordering=last_name|first_name|created_at|id  (+ '-' для DESC)

    Примечания:
      - В detail/retrieve и в me мы добавляем в контекст сериализатора флаги
        include_actions и include_action_history — в ответе придут кадровые
        события сотрудника и их история.
      - Создание /employees/ доступно ТОЛЬКО staff/superuser, даже если у пользователя
        есть модельный пермишен add_employee (это отмечено явно в create()).
    """

    serializer_class = EmployeeSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["last_name", "first_name", "patronymic", "email", "phone_number"]
    ordering_fields = ["last_name", "first_name", "created_at", "id"]
    ordering = ["last_name", "first_name"]
    required_perms_by_action = {
        "add_skill": "employees.manage_employee_skills",
        "remove_skill": "employees.manage_employee_skills",
    }

    def get_permissions(self):
        # create: только staff/superuser (перепроверяете и в методе create)
        if self.action == "create":
            return [IsAuthenticated(), AdminOrActionOrModelPerms()]

        # update/partial_update/destroy: владелец ИЛИ админ/модельные пермишены
        if self.action in {
            "update",
            "partial_update",
            "destroy",
            "add_skill",
            "remove_skill",
        }:
            return [IsAuthenticated(), (IsSelfOrStaff | AdminOrActionOrModelPerms)()]

        # list / retrieve / me -> только аутентификация
        return [IsAuthenticated()]

    def get_queryset(self):
        last_action_code_sq = Subquery(
            EmployeeAction.objects.filter(employee_id=OuterRef("pk"))
            .order_by("-date")
            .values("action")[:1]
        )
        dep_links_prefetch = Prefetch(
            "departments_links",
            queryset=EmployeeDepartment.objects.filter(is_active=True).select_related(
                "department", "role"
            ),
            to_attr="dept_links",
        )

        prefetches = [
            "skills",
            dep_links_prefetch,
            Prefetch("actions", queryset=EmployeeAction.objects.order_by("-date")),
        ]

        qs = (
            Employee.objects.select_related("position")
            .prefetch_related(*prefetches)
            .annotate(last_action_code=last_action_code_sq)
            .order_by(*self.ordering)
        )

        qp = self.request.query_params

        # по отделу: связи + руководитель + сотрудники с ролями через RoleAssignment
        dep = qp.get("department")
        if dep:
            try:
                dep_id = int(dep)
            except (TypeError, ValueError):
                dep_id = None
            if dep_id:
                # Члены отдела
                member_ids = EmployeeDepartment.objects.filter(
                    department_id=dep_id
                ).values("employee_id")
                # Руководитель
                head_ids = Department.objects.filter(id=dep_id).values("head_id")
                # Сотрудники с ролями через RoleAssignment (не члены отдела)
                role_assignment_ids = RoleAssignment.objects.filter(
                    role__department_id=dep_id,
                    is_active=True
                ).values("employee_id")
                
                qs = qs.filter(
                    Q(id__in=member_ids) | Q(id__in=head_ids) | Q(id__in=role_assignment_ids)
                ).distinct()
                
                # Аннотируем тип связи с отделом для каждого сотрудника
                # is_dept_member: True если член отдела, False если только роль
                qs = qs.annotate(
                    _is_dept_member=Exists(
                        EmployeeDepartment.objects.filter(
                            employee_id=OuterRef("pk"),
                            department_id=dep_id,
                            is_active=True
                        )
                    ),
                    _is_dept_head=Exists(
                        Department.objects.filter(
                            id=dep_id,
                            head_id=OuterRef("pk")
                        )
                    ),
                    _has_role_assignment=Exists(
                        RoleAssignment.objects.filter(
                            employee_id=OuterRef("pk"),
                            role__department_id=dep_id,
                            is_active=True
                        )
                    )
                )
                
                # Сохраняем dep_id в request для использования в сериализаторе
                self.request._department_filter_id = dep_id

        # по должности
        position = qp.get("position")
        if position:
            qs = qs.filter(position_id=position)

        # по навыкам (any-of)
        skill_ids = qp.getlist("skill")
        if skill_ids:
            qs = qs.filter(skills__in=skill_ids).distinct()

        # статусы
        email_verified = _to_bool(qp.get("email_verified"))
        if email_verified is not None:
            qs = qs.filter(email_verified=email_verified)

        active = _to_bool(qp.get("active"))
        if active is not None:
            qs = qs.filter(is_active=active)

        # реально активен
        actually = _to_bool(qp.get("actually_active"))
        if actually is True:
            qs = qs.filter(
                Q(email_verified=True)
                & (
                    Q(last_action_code__isnull=True, is_active=True)
                    | ~Q(last_action_code=ACTION_DISMISSED)
                )
            )
        elif actually is False:
            qs = qs.exclude(
                Q(email_verified=True)
                & (
                    Q(last_action_code__isnull=True, is_active=True)
                    | ~Q(last_action_code=ACTION_DISMISSED)
                )
            )

        # фильтр по дате создания
        created_at_gte = qp.get("created_at__gte")
        if created_at_gte:
            qs = qs.filter(created_at__gte=created_at_gte)

        return qs

    def get_serializer_class(self):
        # для списка отдаём облегчённый, но с нужными полями
        if self.action == "list":
            return EmployeeListSerializer
        return EmployeeSerializer

    def create(self, request, *args, **kwargs):
        """Создание пользователя админом: пароль уходит ТОЛЬКО в LDAP, локальный — unusable."""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Only staff can create users."}, status=403)

        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        vd = dict(ser.validated_data)

        if not (vd.get("whatsapp") or vd.get("telegram") or vd.get("wechat")):
            return Response(
                {
                    "detail": "Заполните хотя бы одно из полей: WhatsApp, WeChat или Telegram"
                },
                status=400,
            )

        email = vd.pop("email", None)
        phone_number = vd.pop("phone_number", None)
        password = request.data.get("password")
        if not email or not phone_number or not password:
            return Response(
                {"detail": "email, phone_number и password обязательны"}, status=400
            )

        avatar_bytes = None
        avatar_file = vd.pop("avatar", None)
        if avatar_file and hasattr(avatar_file, "read"):
            try:
                avatar_bytes = avatar_file.read()
            except Exception:
                avatar_bytes = None

        # Проверяем, включен ли LDAP
        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # Режим с LDAP: создаём через DirectoryService
            dto = DirectoryUserDTO(
                first_name=vd.get("first_name", ""),
                last_name=vd.get("last_name", ""),
                email=email.lower(),
                phone_e164=phone_number,
                department_dn=vd.pop("department_dn", None),
                group_cns=vd.pop("group_cns", []) or [],
                initial_password=password,  # ← только LDAP
                avatar_bytes=avatar_bytes,
                is_active=vd.get("is_active", True),
            )

            try:
                user = DirectoryService().create_user(dto)
                # DB-only поля
                for k in (
                    "patronymic",
                    "birth_date",
                    "telegram",
                    "whatsapp",
                    "wechat",
                    "position",
                ):
                    if k in vd:
                        setattr(user, k + ("_id" if k == "position" else ""), vd[k])
                user.save()
                if user.position_id:
                    try:
                        DirectoryService().assign_position(user, user.position)
                    except Exception:
                        # можно залогировать warning; создание пользователя не роняем
                        pass
                skills = vd.get("skills_ids") or []
                if skills:
                    user.skills.set(skills)
                out = self.get_serializer(user)
                return Response(
                    out.data, status=201, headers=self.get_success_headers(out.data)
                )
            except DirectoryLdapError as e:
                return Response({"detail": str(e)}, status=502)
            except DirectoryDbError as e:
                return Response({"detail": str(e)}, status=500)
        else:
            # Режим без LDAP: создаём напрямую в БД
            try:
                with transaction.atomic():
                    user = Employee.objects.create(
                        first_name=vd.get("first_name", ""),
                        last_name=vd.get("last_name", ""),
                        email=email.lower(),
                        phone_number=phone_number,
                        is_active=vd.get("is_active", True),
                        is_ldap_managed=False,
                    )
                    # Устанавливаем пароль
                    user.set_password(password)

                    # DB-only поля
                    for k in (
                        "patronymic",
                        "birth_date",
                        "telegram",
                        "whatsapp",
                        "wechat",
                        "position",
                        "gender",
                    ):
                        if k in vd:
                            setattr(user, k + ("_id" if k == "position" else ""), vd[k])

                    # Аватар
                    if avatar_bytes:
                        user.photo.save(
                            f"avatar_{user.id}.jpg",
                            ContentFile(avatar_bytes),
                            save=False,
                        )

                    user.save()

                    # Навыки
                    skills = vd.get("skills_ids") or []
                    if skills:
                        user.skills.set(skills)

                out = self.get_serializer(user)
                return Response(
                    out.data, status=201, headers=self.get_success_headers(out.data)
                )
            except Exception as e:
                return Response(
                    {"detail": f"Ошибка создания пользователя: {str(e)}"},
                    status=500,
                )

    @action(detail=False, methods=["get", "patch"])
    def me(self, request):
        """GET — профиль текущего пользователя; PATCH — частичное обновление своего профиля.

        Returns:
            Response: Данные профиля.
        """
        instance: Employee = request.user  # type: ignore

        if request.method == "GET":
            instance = (
                Employee.objects.select_related("position")
                .prefetch_related(
                    "skills",
                    Prefetch(
                        "departments_links",
                        queryset=EmployeeDepartment.objects.filter(
                            is_active=True
                        ).select_related("department", "role"),
                        to_attr="dept_links",
                    ),
                    Prefetch(
                        "actions", queryset=EmployeeAction.objects.order_by("-date")
                    ),
                )
                .get(pk=request.user.pk)
            )
            ctx = self.get_serializer_context()
            ctx["include_actions"] = True
            ctx["include_action_history"] = True
            data = self.get_serializer(instance, context=ctx).data
            return Response(data, status=200)

        # PATCH
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        
        try:
            old_email = instance.email  # Сохраняем старый email
            
            # Логируем для отладки
            logger.info(f"[ME PATCH] START - user_id={instance.id}, email={instance.email}")
            logger.info(f"[ME PATCH] request.data keys: {list(request.data.keys())}")
            logger.info(f"[ME PATCH] request.FILES keys: {list(request.FILES.keys())}")
            logger.info(f"[ME PATCH] Content-Type: {request.content_type}")
            
            # ДЕТАЛЬНАЯ ДИАГНОСТИКА АВАТАРА
            if 'avatar' in request.data:
                avatar_data = request.data['avatar']
                avatar_type = type(avatar_data).__name__
                logger.info(f"[ME PATCH] ⚠️ avatar В request.data:")
                logger.info(f"  - type: {avatar_type}")
                logger.info(f"  - value: {repr(avatar_data)[:200]}")
                logger.info(f"  - is empty string: {avatar_data == ''}")
                logger.info(f"  - is None: {avatar_data is None}")
                logger.info(f"  - bool(avatar_data): {bool(avatar_data)}")
                if hasattr(avatar_data, '__len__'):
                    logger.info(f"  - length: {len(avatar_data) if avatar_data else 0}")
            else:
                logger.info(f"[ME PATCH] ✓ avatar НЕТ в request.data")
                
            if 'avatar' in request.FILES:
                avatar_file = request.FILES['avatar']
                logger.info(f"[ME PATCH] ⚠️ avatar В request.FILES:")
                logger.info(f"  - name: {avatar_file.name}")
                logger.info(f"  - size: {avatar_file.size}")
                logger.info(f"  - content_type: {avatar_file.content_type}")
            else:
                logger.info(f"[ME PATCH] ✓ avatar НЕТ в request.FILES")
            
            # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: удаляем пустое поле avatar ДО валидации
            data_for_serializer = request.data
            if 'avatar' in request.data:
                avatar_value = request.data.get('avatar')
                logger.info(f"[ME PATCH] � Проверяем avatar: type={type(avatar_value)}, value={repr(avatar_value)[:100]}")
                
                # Если avatar - пустая строка, создаем копию данных без него
                if avatar_value == '':
                    logger.warning(f"[ME PATCH] 🔧 Удаляем пустое поле avatar из данных")
                    data_for_serializer = {k: v for k, v in request.data.items() if k != 'avatar'}
            
            logger.info("[ME PATCH] Step 1: Validating serializer...")
            logger.info(f"[ME PATCH] Step 1: Передаем в сериализатор data с ключами: {list(data_for_serializer.keys())}")
            
            try:
                ser = self.get_serializer(instance, data=data_for_serializer, partial=True)
                ser.is_valid(raise_exception=True)
                logger.info("[ME PATCH] Step 1: ✓ Serializer validated successfully")
            except ValidationError as ve:
                logger.error(f"[ME PATCH] Step 1: ❌ Serializer validation FAILED:")
                logger.error(f"  - error detail: {ve.detail}")
                logger.error(f"  - error detail type: {type(ve.detail)}")
                if hasattr(ve.detail, 'items'):
                    for field, errors in ve.detail.items():
                        logger.error(f"  - field '{field}': {errors}")
                raise
                
            vd = dict(ser.validated_data)
            
            logger.info(f"[ME PATCH] Step 2: validated_data keys: {list(vd.keys())}")
            if 'avatar' in vd:
                avatar_obj = vd.get('avatar')
                logger.info(f"[ME PATCH] Step 2: avatar В validated_data:")
                logger.info(f"  - type: {type(avatar_obj)}")
                logger.info(f"  - hasattr read: {hasattr(avatar_obj, 'read')}")
                logger.info(f"  - repr: {repr(avatar_obj)[:200]}")
                if avatar_obj:
                    logger.info(f"  - name: {getattr(avatar_obj, 'name', None)}")
                    logger.info(f"  - size: {getattr(avatar_obj, 'size', None)}")
            else:
                logger.info(f"[ME PATCH] Step 2: ✓ avatar НЕТ в validated_data")

            # Проверяем изменение email
            logger.info("[ME PATCH] Step 3: Checking email change...")
            new_email = vd.get("email")
            email_changed = new_email and new_email.lower() != old_email.lower()
            logger.info(f"[ME PATCH] Step 3: email_changed={email_changed}")

            # Бизнес-валидация каналов связи
            logger.info("[ME PATCH] Step 4: Validating contact channels...")
            if all(k not in vd for k in ("whatsapp", "telegram", "wechat")):
                new_whatsapp = vd.get("whatsapp", instance.whatsapp)
                new_telegram = vd.get("telegram", instance.telegram)
                new_wechat = vd.get("wechat", instance.wechat)
                if not (new_whatsapp or new_telegram or new_wechat):
                    logger.warning("[ME PATCH] Step 4: Contact validation failed")
                    return Response(
                        {
                            "detail": "Должен быть указан хотя бы один канал связи (WhatsApp/Telegram/WeChat)."
                        },
                        status=400,
                    )
            logger.info("[ME PATCH] Step 4: Contact validation passed")

            ldap_enabled = _is_ldap_enabled()
            logger.info(f"[ME PATCH] Step 5: LDAP enabled={ldap_enabled}")

            # LDAP-часть
            ldap_keys = {
                "first_name",
                "last_name",
                "email",
                "phone_number",
                "is_active",
            }
            ldap_changes = {
                k: vd.pop(k)
                for k in list(vd.keys())
                if k in ldap_keys
            }
            logger.info(f"[ME PATCH] Step 6: ldap_changes keys: {list(ldap_changes.keys())}")

            svc_changes = dict(ldap_changes)
            pos_key_present = ("position" in request.data) or (
                "position_id" in request.data
            )
            if pos_key_present:
                pos_raw = (
                    request.data.get("position")
                    if "position" in request.data
                    else request.data.get("position_id")
                )
                svc_changes["position"] = pos_raw  # None допустим
                vd.pop("position", None)
                vd.pop("position_id", None)
                logger.info(f"[ME PATCH] Step 6: position={pos_raw}")

            move_to_department_dn = request.data.get("department_dn")
            group_cns = request.data.get("group_cns")
            logger.info(f"[ME PATCH] Step 7: move_to_department_dn={move_to_department_dn}, group_cns={group_cns}")
            
            # Проверяем аватар в validated_data (base64) или в FILES (FormData)
            logger.info("[ME PATCH] Step 8: Processing avatar...")
            avatar_file = ser.validated_data.get("avatar") or request.FILES.get("avatar")
            logger.info(f"[ME PATCH] Step 8: avatar_file={avatar_file}, type={type(avatar_file) if avatar_file else None}")
            if avatar_file and hasattr(avatar_file, "read"):
                try:
                    logger.info("[ME PATCH] Step 8: Reading avatar bytes...")
                    svc_changes["avatar_bytes"] = avatar_file.read()
                    logger.info(f"[ME PATCH] Step 8: avatar_bytes read successfully, length={len(svc_changes['avatar_bytes'])}")
                except Exception as e:
                    logger.error(f"[ME PATCH] Step 8: Error reading avatar: {e}", exc_info=True)

            logger.info(f"[ME PATCH] Step 9: svc_changes keys: {list(svc_changes.keys())}")

            # Проверяем, есть ли у пользователя ldap_dn
            has_ldap_dn = False
            if ldap_enabled:
                from employees.models import LdapSyncState
                has_ldap_dn = (
                    LdapSyncState.objects.filter(
                        model='employee',
                        object_pk=str(instance.pk),
                        ldap_dn__isnull=False
                    ).exists()
                    or (hasattr(instance, 'ldap_dn') and instance.ldap_dn)
                )
                logger.info(f"[ME PATCH] Step 9.5: LDAP enabled, user has ldap_dn: {has_ldap_dn}")

            if ldap_enabled and has_ldap_dn and (svc_changes or move_to_department_dn or group_cns is not None):
                # Режим с LDAP: обновляем через DirectoryService (только для пользователей с ldap_dn)
                logger.info("[ME PATCH] Step 10: LDAP mode - updating via DirectoryService...")
                logger.info(f"[ME PATCH] Step 10: DirectoryService params - svc_changes={list(svc_changes.keys())}, move_to_department_dn={move_to_department_dn}, group_cns={group_cns}")
                svc = DirectoryService()
                try:
                    logger.info(f"[ME PATCH] Step 10: Calling svc.update_user(instance.id={instance.id}, changes={svc_changes}, group_cns={group_cns}, move_to_department_dn={move_to_department_dn})")
                    instance = svc.update_user(
                        instance,
                        changes=svc_changes,
                        group_cns=group_cns if group_cns is not None else None,
                        move_to_department_dn=move_to_department_dn,
                    )
                    logger.info(f"[ME PATCH] Step 10: DirectoryService update successful - updated instance.id={instance.id}")
                    if 'avatar_bytes' in svc_changes:
                        logger.info(f"[ME PATCH] Step 10: Avatar updated in LDAP, new path={instance.avatar.name if instance.avatar else 'None'}")
                except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                    logger.error(f"[ME PATCH] Step 10: DirectoryService error: type={type(e).__name__}, message={e}", exc_info=True)
                    return Response(
                        {"detail": str(e)},
                        status=502 if isinstance(e, DirectoryLdapError) else 500,
                    )
            elif (not ldap_enabled or not has_ldap_dn) and svc_changes:
                # Режим без LDAP или пользователь без ldap_dn: обновляем напрямую в БД
                logger.info(f"[ME PATCH] Step 10: LDAP disabled or no ldap_dn, updating DB directly with: {list(svc_changes.keys())}")
                for k, v in svc_changes.items():
                    if k != "position" and k != "avatar_bytes":
                        setattr(instance, k, v)
                        logger.info(f"[ME PATCH] Step 10: Set {k} = {v}")
                    elif k == "position":
                        instance.position_id = v
                        logger.info(f"[ME PATCH] Step 10: Set position_id = {v}")
                    elif k == "avatar_bytes" and v:
                        logger.info(f"[ME PATCH] Step 10: Saving avatar, size={len(v)}")
                        filename = f"avatar_{instance.id}.jpg"
                        instance.avatar.save(
                            filename,
                            ContentFile(v),
                            save=False,
                        )
                        logger.info(f"[ME PATCH] Step 10: Avatar saved as {filename}, path={instance.avatar.name if instance.avatar else 'None'}")
                logger.info("[ME PATCH] Step 10: Saving instance to DB...")
                instance.save()
                logger.info(f"[ME PATCH] Step 10: Instance saved, avatar path={instance.avatar.name if instance.avatar else 'None'}")

            # DB-only
            logger.info(f"[ME PATCH] Step 11: Processing DB-only fields, vd keys: {list(vd.keys())}")
            
            # КРИТИЧНО: удаляем avatar из vd, т.к. он уже обработан выше
            if 'avatar' in vd:
                logger.info("[ME PATCH] Step 11: Removing avatar from vd (already processed)")
                vd.pop('avatar')
            
            if vd:
                logger.info("[ME PATCH] Step 11: Validating and saving DB-only fields...")
                ser_db = self.get_serializer(instance, data=vd, partial=True)
                try:
                    ser_db.is_valid(raise_exception=True)
                    ser_db.save()
                    instance = ser_db.instance
                    data = ser_db.data
                    logger.info("[ME PATCH] Step 11: DB-only fields saved successfully")
                except ValidationError as ve:
                    logger.warning(f"[ME PATCH] Step 11: Validation error: {ve.detail}")
                    return Response(ve.detail, status=400)
            else:
                logger.info("[ME PATCH] Step 11: No DB-only fields to save")
                data = self.get_serializer(instance).data
            
            logger.info(f"[ME PATCH] Step 12: Final response data keys: {list(data.keys())}")
            if 'avatar' in data:
                avatar_preview = data['avatar'][:100] if data['avatar'] else None
                logger.info(f"[ME PATCH] Step 12: avatar in response: {avatar_preview}...")

            # Сброс email_verified при изменении email
            if email_changed:
                logger.info("[ME PATCH] Step 13: Processing email change...")
                from django.utils.crypto import get_random_string
                from common.emails import send_templated_mail

                instance.email_verified = False
                instance.email_activation_code = get_random_string(6, "0123456789")
                instance.save(update_fields=["email_verified", "email_activation_code"])

                # Отправляем код на новый email
                try:
                    send_templated_mail(
                        subject="Подтверждение нового email",
                        to=[instance.email],
                        template_base="emails/registration_verify_code",
                        context={"code": instance.email_activation_code, "user": instance},
                    )
                    logger.info("[ME PATCH] Step 13: Email verification sent")
                except Exception as email_err:
                    logger.warning(f"[ME PATCH] Step 13: Failed to send email: {email_err}")

                # Обновляем данные в ответе
                data["email_verified"] = False

            logger.info("[ME PATCH] SUCCESS - returning response")
            return Response(data, status=200)
            
        except Exception as e:
            logger.error(f"[ME PATCH] FATAL ERROR: {e}", exc_info=True)
            logger.error(f"[ME PATCH] Traceback:\n{traceback.format_exc()}")
            return Response(
                {"detail": f"Internal server error: {str(e)}"},
                status=500
            )

    @action(detail=True, methods=["post"])
    def add_skill(self, request, pk=None):
        """
        body: { "skill_id": 3 } ИЛИ { "skill_name": "Python" }
        """
        emp = self.get_object()
        sid = request.data.get("skill_id")
        sname = (request.data.get("skill_name") or "").strip()

        sk = None
        if sid:
            sk = Skill.objects.filter(pk=sid).first()
        if not sk and sname:
            sk = Skill.objects.filter(
                name__iexact=sname
            ).first() or Skill.objects.create(name=sname)
        if not sk:
            return Response({"detail": "Навык не найден/не указан"}, status=400)

        emp.skills.add(sk)
        return Response(
            {"ok": True, "skill": {"id": sk.id, "name": sk.name}}, status=200
        )

    @action(detail=True, methods=["post"])
    def remove_skill(self, request, pk=None):
        """
        body: { "skill_id": 3 } ИЛИ { "skill_name": "Python" }
        """
        emp = self.get_object()
        sid = request.data.get("skill_id")
        sname = (request.data.get("skill_name") or "").strip()

        sk = None
        if sid:
            sk = Skill.objects.filter(pk=sid).first()
        if not sk and sname:
            sk = Skill.objects.filter(name__iexact=sname).first()
        if not sk:
            return Response({"detail": "Навык не найден"}, status=404)

        emp.skills.remove(sk)
        return Response(
            {"ok": True, "removed": {"id": sk.id, "name": sk.name}}, status=200
        )

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        GET /api/v1/employees/export-excel/
        
        Экспортирует всех сотрудников в Excel формат.
        
        Параметры:
            - Применяются все фильтры из queryset (department, position, active и т.д.)
        
        Возвращает:
            - Excel файл (.xlsx) с данными сотрудников
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Получаем отфильтрованный queryset
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.select_related('position').prefetch_related(
            'skills',
            'departments_links__department',
            'departments_links__role'
        ).order_by('last_name', 'first_name')
        
        # Создаём Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"
        
        # Заголовки
        headers = [
            'ID',
            'Фамилия',
            'Имя',
            'Отчество',
            'Email',
            'Телефон',
            'Должность',
            'Отделы',
            'Дата рождения',
            'Дата регистрации',
            'Активен',
            'Email подтвержден',
            'Навыки',
            'Telegram',
            'WhatsApp',
            'WeChat'
        ]
        
        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Записываем данные
        for row_num, emp in enumerate(queryset, 2):
            # Получаем отделы
            departments = emp.departments_links.filter(is_active=True)
            dept_names = ", ".join([
                f"{d.department.name}" + (f" ({d.role.name})" if d.role else "")
                for d in departments
            ])
            
            # Получаем навыки
            skills = ", ".join([s.name for s in emp.skills.all()])
            
            # Конвертируем PhoneNumber поля в строки
            def safe_phone_str(phone_field):
                """Безопасная конвертация PhoneNumber в строку"""
                if not phone_field:
                    return ''
                try:
                    from phonenumbers import format_number, PhoneNumberFormat
                    return format_number(phone_field, PhoneNumberFormat.INTERNATIONAL)
                except Exception:
                    try:
                        return str(phone_field)
                    except Exception:
                        return ''
            
            phone_str = safe_phone_str(emp.phone_number)
            whatsapp_str = safe_phone_str(emp.whatsapp)
            
            # Данные строки
            row_data = [
                emp.id,
                emp.last_name or '',
                emp.first_name or '',
                emp.patronymic or '',
                emp.email or '',
                phone_str,
                emp.position.name if emp.position else '',
                dept_names,
                emp.birth_date.strftime('%d.%m.%Y') if emp.birth_date else '',
                emp.created_at.strftime('%d.%m.%Y %H:%M') if emp.created_at else '',
                'Да' if emp.is_active else 'Нет',
                'Да' if emp.email_verified else 'Нет',
                skills,
                emp.telegram or '',
                whatsapp_str,
                emp.wechat or ''
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Автоподбор ширины колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Максимум 50
            ws.column_dimensions[column].width = adjusted_width
        
        # Закрепляем первую строку (заголовки)
        ws.freeze_panes = 'A2'
        
        # Создаём HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"employees_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Сохраняем в response
        wb.save(response)
        
        return response

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        if getattr(self, "action", None) in {"retrieve", "me"}:
            ctx["include_actions"] = True
            ctx["include_action_history"] = True
        return ctx

    def partial_update(self, request, *args, **kwargs):
        """Частичное обновление: сначала LDAP-совместимые поля → затем только DB-only.

        Returns:
            Response: Обновлённые данные сотрудника.
        """
        emp = self.get_object()
        old_email = emp.email  # Сохраняем старый email для проверки

        print("=" * 80)
        print(
            "[EMP PATCH] pk=%s, actor=%s, method=%s"
            % (emp.pk, getattr(request.user, "id", None), request.method)
        )
        print(f"[EMP PATCH] request.data keys: {list(request.data.keys())}")
        print(f"[EMP PATCH] request.FILES keys: {list(request.FILES.keys())}")
        print(f"[EMP PATCH] Content-Type: {request.content_type}")
        if "avatar" in request.data:
            avatar_data = request.data["avatar"]
            avatar_len = len(str(avatar_data)) if avatar_data else 0
            print(
                "[EMP PATCH] avatar in request.data: type=%s, length=%s"
                % (type(avatar_data), avatar_len)
            )
        if "avatar" in request.FILES:
            avatar_file_req = request.FILES["avatar"]
            print(
                "[EMP PATCH] avatar in request.FILES: name=%s, size=%s"
                % (avatar_file_req.name, avatar_file_req.size)
            )

        data = request.data.copy()
        if hasattr(data, "_mutable") and not data._mutable:
            data._mutable = True
        avatar_raw = data.get("avatar")
        if avatar_raw in ("", None):
            data.pop("avatar", None)
            print("[EMP PATCH] removed empty avatar value from payload")

        ser = self.get_serializer(emp, data=data, partial=True)
        try:
            ser.is_valid(raise_exception=True)
        except ValidationError as exc:
            print("[EMP PATCH] serializer errors: %s" % exc.detail)
            raise
        vd = dict(ser.validated_data)
        ldap_enabled = _is_ldap_enabled()

        print(f"[EMP PATCH] validated_data keys: {list(vd.keys())}")
        if "avatar" in vd:
            avatar_vd = vd["avatar"]
            print(
                (
                    "[EMP PATCH] avatar in validated_data: type=%s, "
                    "hasattr read=%s"
                )
                % (type(avatar_vd), hasattr(avatar_vd, "read"))
            )
            if hasattr(avatar_vd, "name"):
                print(
                    "[EMP PATCH] avatar object info: name=%s, size=%s"
                    % (
                        getattr(avatar_vd, "name", None),
                        getattr(avatar_vd, "size", None),
                    )
                )

        # Проверяем изменение email
        new_email = vd.get("email")
        email_changed = new_email and new_email.lower() != old_email.lower()

        # --- LDAP часть ---
        ldap_keys = {
            "first_name",
            "last_name",
            "email",
            "phone_number",
            "is_active",
        }
        ldap_changes = {
            k: vd.pop(k)
            for k in list(vd.keys())
            if k in ldap_keys
        }

        svc_changes = dict(ldap_changes)
        pos_key_present = ("position" in data) or ("position_id" in data)
        if pos_key_present:
            pos_raw = (
                data.get("position")
                if "position" in data
                else data.get("position_id")
            )
            svc_changes["position"] = pos_raw  # None допустим
            vd.pop("position", None)
            vd.pop("position_id", None)

        move_to_department_dn = data.get("department_dn")
        group_cns = data.get("group_cns")  # список строк/None

        # avatar → bytes (если разрешено править аватар в LDAP)
        avatar_file = ser.validated_data.get("avatar")
        avatar_file_type = type(avatar_file) if avatar_file else None
        print(
            "[EMP PATCH] avatar_file extracted: %s, type=%s"
            % (avatar_file, avatar_file_type)
        )
        if avatar_file and hasattr(avatar_file, "read"):
            try:
                svc_changes["avatar_bytes"] = avatar_file.read()
                if hasattr(avatar_file, "seek"):
                    avatar_file.seek(0)
                print(
                    "[EMP PATCH] avatar_bytes prepared for LDAP, length=%s"
                    % len(svc_changes["avatar_bytes"])
                )
            except Exception as exc:
                print(f"[EMP PATCH] Error reading avatar for LDAP: {exc}")

        # Проверяем, есть ли у пользователя ldap_dn
        has_ldap_dn = False
        if ldap_enabled:
            from employees.models import LdapSyncState
            has_ldap_dn = (
                LdapSyncState.objects.filter(
                    model='employee',
                    object_pk=str(emp.pk),
                    ldap_dn__isnull=False
                ).exists()
                or (hasattr(emp, 'ldap_dn') and emp.ldap_dn)
            )
            print(f"[EMP PATCH] LDAP enabled, user has ldap_dn: {has_ldap_dn}")

        if ldap_enabled and has_ldap_dn:
            # LDAP mode: sync with DirectoryService (только для пользователей с ldap_dn)
            svc = DirectoryService()
            if svc_changes or move_to_department_dn or group_cns is not None:
                try:
                    emp = svc.update_user(
                        emp,
                        changes=svc_changes,
                        group_cns=group_cns if group_cns is not None else None,
                        move_to_department_dn=move_to_department_dn,
                    )
                except (
                    DirectoryLdapError,
                    DirectoryDbError,
                    DirectoryServiceError,
                ) as e:
                    return Response(
                        {"detail": str(e)},
                        status=(
                            502 if isinstance(e, DirectoryLdapError) else 500
                        ),
                    )
        else:
            # Non-LDAP mode or user without ldap_dn: direct DB updates
            # Restore ldap_changes back to vd for DB-only save
            vd.update(ldap_changes)
            
            # Handle position update in non-LDAP mode
            if pos_key_present:
                pos_raw = (
                    request.data.get("position")
                    if "position" in request.data
                    else request.data.get("position_id")
                )
                if pos_raw is None:
                    vd["position"] = None
                elif isinstance(pos_raw, int):
                    try:
                        vd["position"] = Position.objects.get(id=pos_raw)
                    except Position.DoesNotExist:
                        return Response(
                            {"detail": f"Position {pos_raw} not found"},
                            status=400,
                        )
                elif isinstance(pos_raw, dict) and "id" in pos_raw:
                    try:
                        vd["position"] = Position.objects.get(id=pos_raw["id"])
                    except Position.DoesNotExist:
                        return Response(
                            {"detail": f"Position {pos_raw['id']} not found"},
                            status=400,
                        )

            # Handle avatar in non-LDAP mode
            if avatar_file and hasattr(avatar_file, "read"):
                try:
                    from django.core.files.base import ContentFile

                    avatar_bytes = avatar_file.read()
                    print(
                        "[EMP PATCH] Non-LDAP avatar bytes len=%s"
                        % (len(avatar_bytes) if avatar_bytes else 0)
                    )
                    emp.avatar.save(
                        avatar_file.name,
                        ContentFile(avatar_bytes),
                        save=False,
                    )
                    vd.pop("avatar", None)
                    print(
                        (
                            "[EMP PATCH] Avatar saved to model field, "
                            "current path=%s"
                        )
                        % (emp.avatar.name if emp.avatar else None)
                    )
                except Exception as exc:
                    print(
                        "[EMP PATCH] Error while saving avatar in DB mode: %s"
                        % exc
                    )

        # --- DB-only часть ---
        # Обновляем только оставшиеся поля, чтобы не перетирать работу сервиса
        
        # КРИТИЧНО: удаляем avatar из vd, т.к. он уже обработан выше
        if 'avatar' in vd:
            print("[EMP PATCH] Removing avatar from vd (already processed)")
            vd.pop('avatar')
        
        if vd:
            ser_db = self.get_serializer(emp, data=vd, partial=True)
            try:
                ser_db.is_valid(raise_exception=True)
                ser_db.save()
                emp = ser_db.instance
                data = ser_db.data
            except ValidationError as exc:
                print("[EMP PATCH] DB serializer validation error: %s" % exc.detail)
                return Response(exc.detail, status=400)
        else:
            data = self.get_serializer(emp).data

        print(f"[EMP PATCH] Final response keys: {list(data.keys())}")
        if isinstance(data, dict) and data.get("avatar"):
            print(
                "[EMP PATCH] avatar preview in response: %s..."
                % data["avatar"][:80]
            )

        # Сброс email_verified при изменении email
        if email_changed:
            from django.utils.crypto import get_random_string
            from common.emails import send_templated_mail

            emp.email_verified = False
            emp.email_activation_code = get_random_string(6, "0123456789")
            emp.save(update_fields=["email_verified", "email_activation_code"])

            # Отправляем код на новый email
            try:
                send_templated_mail(
                    subject="Подтверждение нового email",
                    to=[emp.email],
                    template_base="emails/registration_verify_code",
                    context={"code": emp.email_activation_code, "user": emp},
                )
            except Exception:
                # Не падаем если не удалось отправить письмо
                pass

            # Обновляем данные в ответе
            data["email_verified"] = False

        return Response(data, status=200)


class PositionViewSet(HistoryActionMixin, viewsets.ModelViewSet):
    """
    /api/v1/positions/
      GET list/retrieve   — аутентифицированные
      POST/PUT/PATCH/DEL  — staff/superuser ИЛИ пользователь с model perms
      Экшены:
        POST /{id}/set-groups
        POST /{id}/add-groups
        POST /{id}/remove-groups
        GET  /{id}/permissions
    """

    queryset = Position.objects.all().prefetch_related("groups")
    serializer_class = PositionSerializer
    pagination_class = None
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "id"]
    ordering = ["name"]
    history_diff_fields = ["name", "description"]
    permission_classes = [AdminOrActionOrModelPerms]
    required_perms_by_action = {
        "set_groups": "employees.assign_position_groups",
        "add_groups": "employees.assign_position_groups",
        "remove_groups": "employees.assign_position_groups",
    }

    def _validate_groups_payload(self, request):
        ids = request.data.get("groups")
        if not isinstance(ids, list):
            return None, Response(
                {"detail": "Поле 'groups' должно быть списком id"}, status=400
            )
        qs = Group.objects.filter(id__in=ids)
        if qs.count() != len(set(ids)):
            return None, Response(
                {"detail": "Некоторые группы не найдены"}, status=400
            )
        return qs, None

    def get_permissions(self):
        """Возвращает пермишены для текущего экшена.

        Returns:
            list: Инスタнсы DRF-permission'ов.
        """
        if self.action in {"list", "retrieve", "permissions"}:
            return [IsAuthenticated()]
        if self.action == "create":
            return [AdminOrActionOrModelPerms()]
        return [AdminOrActionOrModelPerms()]

    @action(detail=True, methods=["post"])
    def set_groups(self, request, pk=None):
        pos = self.get_object()
        qs, err = self._validate_groups_payload(request)
        if err:
            return err
        pos.groups.set(qs)
        err2 = _ldap_try(lambda: DirectoryService().reconcile_position(pos))
        if err2:
            return err2
        return Response(
            {
                "ok": True,
                "group_ids": list(pos.groups.values_list("id", flat=True)),
            }
        )

    @action(detail=True, methods=["post"])
    def add_groups(self, request, pk=None):
        pos = self.get_object()
        qs, err = self._validate_groups_payload(request)
        if err:
            return err
        pos.groups.add(*qs)
        err2 = _ldap_try(lambda: DirectoryService().reconcile_position(pos))
        if err2:
            return err2
        return Response(
            {
                "ok": True,
                "group_ids": list(pos.groups.values_list("id", flat=True)),
            }
        )

    @action(detail=True, methods=["post"])
    def remove_groups(self, request, pk=None):
        pos = self.get_object()
        qs, err = self._validate_groups_payload(request)
        if err:
            return err
        pos.groups.remove(*qs)
        err2 = _ldap_try(lambda: DirectoryService().reconcile_position(pos))
        if err2:
            return err2
        return Response(
            {
                "ok": True,
                "group_ids": list(pos.groups.values_list("id", flat=True)),
            }
        )

    @action(detail=True, methods=["get"])
    def permissions(self, request, pk=None):
        pos = self.get_object()
        perms = (
            Permission.objects.filter(group__positions=pos)
            .select_related("content_type")
            .distinct()
        )
        data = [
            {
                "id": p.id,
                "codename": f"{p.content_type.app_label}.{p.codename}",
                "name": p.name,
                "app": p.content_type.app_label,
                "model": p.content_type.model,
            }
            for p in perms
        ]
        return Response({"count": len(data), "results": data}, status=200)

    def perform_create(self, serializer):
        pos = serializer.save()
        err = _ldap_try(lambda: DirectoryService().reconcile_position(pos))
        if err:
            # позиция создана в БД, но LDAP не синкнулся — возвращаем 502
            raise Exception(
                err.data["detail"]
            )  # или пропустите исключение, если не хотите рвать транзакцию

    def perform_update(self, serializer):
        pos = serializer.save()
        err = _ldap_try(lambda: DirectoryService().reconcile_position(pos))
        if err:
            # позиция обновлена в БД; LDAP можно синкнуть повторно позднее
            pass

    def perform_destroy(self, instance):
        # best-effort: не блокируем удаление позиции из БД из-за LDAP
        _ldap_try(lambda: DirectoryService().delete_position_group(instance))
        return super().perform_destroy(instance)


class DepartmentRoleViewSet(viewsets.ModelViewSet):
    """
        Роли отдела:
            - list/retrieve с фильтром ?department=<id>
            - create/update/destroy → требуется право DeptPerm.ASSIGN_ROLE
                в рамках отдела роли
      - GET  /department-roles/perm_choices/
      - GET  /department-roles/{id}/perms/
      - POST /department-roles/{id}/set_perms  (ids или codes; полная замена)
    """

    queryset = (
        DepartmentRole.objects.select_related("department")
        .prefetch_related("scoped_permissions")
        .all()
    )
    serializer_class = DepartmentRoleSerializer

    class AssignRolePerm(AdminOrDeptAllowed):
        """Право на назначение ролей участникам отдела."""

        required_code = DeptPerm.ASSIGN_ROLE

    # стабильная сортировка для листинга
    ordering_fields = ("name", "id")
    ordering = ("name", "id")

    def get_permissions(self):
        if self.action in {
            "create",
            "update",
            "partial_update",
            "destroy",
            "set_perms",
            "assign",
            "revoke",
        }:
            return [self.AssignRolePerm()]
        return [IsAuthenticated()]

    def get_queryset(self):
        """
        Фильтрация по отделу и стабильный порядок.
        """
        qs = super().get_queryset()
        dept = self.request.query_params.get("department")
        if dept:
            qs = qs.filter(department_id=dept)

        ord_param = self.request.query_params.get("ordering")
        if ord_param in {"name", "-name", "id", "-id"}:
            # стабильный тай-брейк по id
            qs = qs.order_by(
                ord_param, "id" if not ord_param.startswith("-") else "-id"
            )
        else:
            qs = qs.order_by(*self.ordering)
        return qs

    def create(self, request, *args, **kwargs):
        """Создание роли: сначала группа в LDAP → затем запись в БД (если LDAP включен)."""
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        
        dept_id = ser.validated_data.get("department")
        if isinstance(dept_id, Department):
            dept = dept_id
        else:
            dept = get_object_or_404(Department, id=dept_id)
        
        name = ser.validated_data["name"]
        codes = ser.validated_data.pop("scoped_permission_codes", None)
        perms = ser.validated_data.pop("scoped_permissions", None)
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: создаём через DepartmentService
            from employees.ldap.services.department_service import DepartmentService
            from employees.ldap.services.group_service import GroupService
            from employees.ldap.services.user_service import UserService
            
            group_service = GroupService()
            user_service = UserService(group_service)
            dept_service = DepartmentService(group_service, user_service)
            
            # Подготавливаем scoped_permissions (codes имеют приоритет над ids)
            scoped_permissions = None
            if codes is not None:
                scoped_permissions = list(
                    DepartmentPermission.objects.filter(code__in=codes)
                )
            elif perms is not None:
                scoped_permissions = list(perms)
            
            try:
                role = dept_service.create_role(
                    department=dept,
                    name=name,
                    scoped_permissions=scoped_permissions,
                )
                return Response(
                    self.get_serializer(role).data, status=status.HTTP_201_CREATED
                )
            except DirectoryLdapError as e:
                return Response(
                    {"detail": f"LDAP error: {e}"}, status=status.HTTP_502_BAD_GATEWAY
                )
            except DirectoryDbError as e:
                return Response(
                    {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            # Режим без LDAP: создаём напрямую через сериализатор
            role = DepartmentRole.objects.create(
                department=dept,
                name=name,
            )
            # codes имеют приоритет над ids (как в сериализаторе)
            if codes is not None:
                qs = DepartmentPermission.objects.filter(code__in=codes)
                role.scoped_permissions.set(list(qs))
            elif perms is not None:
                role.scoped_permissions.set(perms)
            
            return Response(
                self.get_serializer(role).data, status=status.HTTP_201_CREATED
            )

    def update(self, request, *args, **kwargs):
        """Обновление роли: сначала LDAP → затем БД (если LDAP включен)."""
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        ser = self.get_serializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: обновляем через DepartmentService
            from employees.ldap.services.department_service import DepartmentService
            from employees.ldap.services.group_service import GroupService
            from employees.ldap.services.user_service import UserService
            
            group_service = GroupService()
            user_service = UserService(group_service)
            dept_service = DepartmentService(group_service, user_service)
            
            changes = {}
            if "name" in ser.validated_data:
                changes["name"] = ser.validated_data["name"]
            if "scoped_permissions" in ser.validated_data:
                changes["scoped_permissions"] = ser.validated_data["scoped_permissions"]
            if "scoped_permission_codes" in ser.validated_data:
                changes["scoped_permission_codes"] = ser.validated_data["scoped_permission_codes"]
            
            try:
                role = dept_service.update_role(instance, changes)
                return Response(self.get_serializer(role).data)
            except DirectoryLdapError as e:
                return Response(
                    {"detail": f"LDAP error: {e}"}, status=status.HTTP_502_BAD_GATEWAY
                )
            except DirectoryDbError as e:
                return Response(
                    {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            # Режим без LDAP: обновляем через сериализатор
            self.perform_update(ser)
            return Response(ser.data)

    def partial_update(self, request, *args, **kwargs):
        """Частичное обновление роли."""
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Удаление роли: сначала группа из LDAP → затем запись из БД (если LDAP включен)."""
        instance = self.get_object()
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: удаляем через DepartmentService
            from employees.ldap.services.department_service import DepartmentService
            from employees.ldap.services.group_service import GroupService
            from employees.ldap.services.user_service import UserService
            
            group_service = GroupService()
            user_service = UserService(group_service)
            dept_service = DepartmentService(group_service, user_service)
            
            try:
                dept_service.delete_role(instance)
                return Response(status=status.HTTP_204_NO_CONTENT)
            except DirectoryLdapError as e:
                return Response(
                    {"detail": f"LDAP error: {e}"}, status=status.HTTP_502_BAD_GATEWAY
                )
            except DirectoryDbError as e:
                return Response(
                    {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            # Режим без LDAP: удаляем напрямую
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"])
    def perm_choices(self, request):
        """
        Возвращает справочник скоуп-прав для ролей отдела из DeptPerm.CHOICES.
        Если каких-то записей нет в БД, создаёт их (идемпотентно).
        """
        data = _ensure_department_permissions()
        return Response({"count": len(data), "results": data}, status=200)

    @action(detail=True, methods=["get"])
    def perms(self, request, pk=None):
        """
        Возвращает список прав, назначенных данной роли (отсортирован по коду).
        """
        role = self.get_object()
        data = [
            {"id": p.id, "code": p.code, "name": p.name}
            for p in role.scoped_permissions.order_by("code")
        ]
        return Response({"count": len(data), "results": data}, status=200)

    @action(detail=True, methods=["post"])
    def set_perms(self, request, pk=None):
        """
        Полностью заменяет набор прав у роли.
        Тело запроса:
          - либо "permission_ids": [1,2,3]
          - либо "permission_codes": ["manage_department", ...]
        """
        role = self.get_object()

        ids = request.data.get("permission_ids") or []
        codes = request.data.get("permission_codes") or []

        if isinstance(ids, list) and ids:
            # валидация набора id
            ids_int = {int(i) for i in ids if str(i).isdigit()}
            qs = DepartmentPermission.objects.filter(id__in=ids_int)
            if qs.count() != len(ids_int):
                return Response(
                    {"detail": "Некоторые permission_ids не найдены."}, status=400
                )
        elif isinstance(codes, list) and codes:
            # валидация набора кодов
            codes_set = set(codes)
            qs = DepartmentPermission.objects.filter(code__in=codes_set)
            if qs.count() != len(codes_set):
                return Response(
                    {"detail": "Некоторые permission_codes не найдены."}, status=400
                )
        else:
            qs = DepartmentPermission.objects.none()

        role.scoped_permissions.set(list(qs))
        ser = self.get_serializer(role)
        return Response(ser.data, status=200)

    @action(detail=True, methods=["get"])
    def assignments(self, request, pk=None):
        """
        Возвращает список назначений роли (RoleAssignment).
        
        Query params:
          - ?active=true/false — фильтр по is_active (по умолчанию только активные)
        """
        role = self.get_object()
        qs = RoleAssignment.objects.filter(role=role).select_related(
            "employee", "assigned_by"
        )
        
        active = request.query_params.get("active", "true").lower()
        if active == "true":
            qs = qs.filter(is_active=True)
        elif active == "false":
            qs = qs.filter(is_active=False)
        
        data = [
            {
                "id": a.id,
                "employee_id": a.employee_id,
                "employee_name": str(a.employee) if a.employee else None,
                "assigned_at": a.assigned_at.isoformat() if a.assigned_at else None,
                "assigned_by_id": a.assigned_by_id,
                "assigned_by_name": str(a.assigned_by) if a.assigned_by else None,
                "is_active": a.is_active,
            }
            for a in qs.order_by("-assigned_at")
        ]
        return Response({"count": len(data), "results": data}, status=200)

    @action(detail=True, methods=["post"])
    def assign(self, request, pk=None):
        """
        Назначает роль сотруднику (не требует членства в отделе).
        
        Тело запроса:
          - employee_id: int — ID сотрудника
        
        Требуется право DeptPerm.ASSIGN_ROLE в отделе роли.
        При включенном LDAP: сначала добавляет в группу LDAP, потом в БД.
        """
        role = self.get_object()
        employee_id = request.data.get("employee_id")
        
        if not employee_id:
            return Response(
                {"detail": "employee_id is required."}, status=400
            )
        
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {"detail": "Сотрудник не найден."}, status=404
            )
        
        assigned_by = request.user if request.user.is_authenticated else None
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: через DepartmentService (сначала LDAP, потом БД)
            try:
                from employees.ldap.services.department_service import DepartmentService
                from employees.ldap.services.group_service import GroupService
                from employees.ldap.services.user_service import UserService
                
                group_service = GroupService()
                user_service = UserService(group_service)
                dept_service = DepartmentService(group_service, user_service)
                
                assignment = dept_service.assign_role(employee, role, assigned_by)
                
                return Response({
                    "id": assignment.id,
                    "employee_id": assignment.employee_id,
                    "role_id": assignment.role_id,
                    "assigned_at": assignment.assigned_at.isoformat() if assignment.assigned_at else None,
                    "is_active": assignment.is_active,
                }, status=201)
            except DirectoryLdapError as e:
                return Response(
                    {"detail": f"LDAP error: {e}"}, status=status.HTTP_502_BAD_GATEWAY
                )
            except DirectoryDbError as e:
                return Response(
                    {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            # Режим без LDAP: создаём RoleAssignment напрямую
            assignment, created = RoleAssignment.objects.update_or_create(
                employee=employee,
                role=role,
                defaults={
                    "is_active": True,
                    "assigned_by": assigned_by,
                }
            )
            return Response({
                "id": assignment.id,
                "employee_id": assignment.employee_id,
                "role_id": assignment.role_id,
                "assigned_at": assignment.assigned_at.isoformat() if assignment.assigned_at else None,
                "is_active": assignment.is_active,
            }, status=201)

    @action(detail=True, methods=["post"])
    def revoke(self, request, pk=None):
        """
        Отзывает роль у сотрудника.
        
        Тело запроса:
          - employee_id: int — ID сотрудника
        
        Требуется право DeptPerm.ASSIGN_ROLE в отделе роли.
        При включенном LDAP: сначала удаляет из группы LDAP, потом из БД.
        """
        role = self.get_object()
        employee_id = request.data.get("employee_id")
        
        if not employee_id:
            return Response(
                {"detail": "employee_id is required."}, status=400
            )
        
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {"detail": "Сотрудник не найден."}, status=404
            )
        
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # Режим с LDAP: через DepartmentService (сначала LDAP, потом БД)
            try:
                from employees.ldap.services.department_service import DepartmentService
                from employees.ldap.services.group_service import GroupService
                from employees.ldap.services.user_service import UserService
                
                group_service = GroupService()
                user_service = UserService(group_service)
                dept_service = DepartmentService(group_service, user_service)
                
                dept_service.revoke_role(employee, role)
                
                return Response(status=204)
            except DirectoryLdapError as e:
                return Response(
                    {"detail": f"LDAP error: {e}"}, status=status.HTTP_502_BAD_GATEWAY
                )
        else:
            # Режим без LDAP: деактивируем RoleAssignment напрямую
            RoleAssignment.objects.filter(
                employee=employee,
                role=role
            ).update(is_active=False)
            return Response(status=204)


class SkillViewSet(viewsets.ModelViewSet):
    """
    /api/v1/skills/
      - GET list/retrieve      — IsAuthenticated
      - POST (create)          — IsAuthenticated (любой авторизованный может создавать новые навыки)
      - PATCH/PUT              — staff/superuser ИЛИ perm employees.change_skill
      - DELETE                 — staff/superuser ИЛИ perm employees.delete_skill
    Экшены:
      - POST /skills/bulk_create   — как create (любой авторизованный)
      - POST /skills/merge         — объединить source->target (perm employees.change_skill)
        body: {
          "source_id": 1,
          "target_id": 2,
          "reassign": true,         # переназначить сотрудникам source -> target (по умолч. true)
          "delete_source": true     # удалить исходный навык после переназначения (по умолч. true)
        }
    """

    queryset = Skill.objects.all().order_by("name")
    serializer_class = SkillSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name", "id"]
    ordering = ["name"]
    pagination_class = None
    required_perms_by_action = {
        "merge": "employees.change_skill",
    }

    def get_permissions(self):
        if self.action in ("create", "bulk_create", "list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), AdminOrActionOrModelPerms()]

    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        names = request.data.get("names")
        if not isinstance(names, list) or not names:
            return Response(
                {"detail": "Поле 'names' должно быть непустым списком строк"},
                status=400,
            )

        # нормализация: trim + отбраковка пустых, дедуп (case-insensitive)
        cleaned = []
        seen = set()
        for n in names:
            if not isinstance(n, str):
                continue
            s = n.strip()
            if not s:
                continue
            key = s.lower()
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(s)

        # отфильтруем уже существующие (по точному совпадению имени)
        existing = set(
            Skill.objects.filter(name__in=cleaned).values_list("name", flat=True)
        )
        to_create = [Skill(name=n) for n in cleaned if n not in existing]
        Skill.objects.bulk_create(to_create, ignore_conflicts=True)

        created = Skill.objects.filter(
            name__in=[n for n in cleaned if n not in existing]
        ).order_by("name")
        return Response(
            {
                "created_count": created.count(),
                "created": SkillSerializer(created, many=True).data,
            },
            status=201,
        )

    @action(detail=False, methods=["post"])
    @transaction.atomic
    def merge(self, request):
        """
        Заменяет навык source на target, опционально переносит всех сотрудников и удаляет source.
        Требуется право employees.change_skill.
        """
        sid = request.data.get("source_id")
        tid = request.data.get("target_id")
        reassign = bool(request.data.get("reassign", True))
        delete_source = bool(request.data.get("delete_source", True))

        if not sid or not tid:
            return Response({"detail": "source_id и target_id обязательны"}, status=400)
        try:
            sid = int(sid)
            tid = int(tid)
        except (TypeError, ValueError):
            return Response(
                {"detail": "source_id и target_id должны быть числами"}, status=400
            )
        if sid == tid:
            return Response(
                {"detail": "source_id и target_id не должны совпадать"}, status=400
            )

        try:
            source = Skill.objects.get(pk=sid)
            target = Skill.objects.get(pk=tid)
        except Skill.DoesNotExist:
            return Response({"detail": "Skill не найден"}, status=404)

        moved_count = 0
        if reassign:
            # переназначить всем сотрудникам: source -> target
            qs = Employee.objects.filter(skills=source).only("id").distinct()
            for emp in qs:
                emp.skills.add(target)
                emp.skills.remove(source)
                moved_count += 1

        # удалить исходный навык при необходимости
        if delete_source:
            source.delete()

        return Response(
            {
                "ok": True,
                "source_id": sid,
                "target_id": tid,
                "reassigned_employees": moved_count,
                "source_deleted": bool(delete_source),
            },
            status=200,
        )


class EmployeeActionViewSet(HistoryActionMixin, viewsets.ModelViewSet):
    """
    /api/v1/employee-actions/
      - GET list/retrieve  — IsAuthenticated
      - POST create        — staff/superuser ИЛИ perm employees.add_employeeaction
      - PATCH/PUT update   — staff/superuser ИЛИ perm employees.change_employeeaction
      - DELETE destroy     — staff/superuser ИЛИ perm employees.delete_employeeaction

    Фильтры:
      ?employee=<id>   — по сотруднику
      ?date_from=ISO   — >=
      ?date_to=ISO     — <=
    """

    serializer_class = EmployeeActionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["date", "id"]
    ordering = ["-date"]
    pagination_class = None
    history_diff_fields = ["action", "date", "comment", "extra", "employee_id"]

    def get_queryset(self):
        qs = EmployeeAction.objects.select_related("employee").order_by(*self.ordering)
        qp = self.request.query_params
        emp = qp.get("employee")
        if emp:
            try:
                emp_id = int(emp)
                qs = qs.filter(employee_id=emp_id)
            except (TypeError, ValueError):
                return EmployeeAction.objects.none()
        df = qp.get("date_from")
        dt = qp.get("date_to")
        if df:
            try:
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if dt:
            try:
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass
        return qs

    # --- права на запись через модельные пермишены ---
    def get_permissions(self):
        if self.action == "create":
            self.required_perm_code = "employees.add_employeeaction"
            return [IsAuthenticated(), AdminOrActionOrModelPerms()]
        if self.action in ("update", "partial_update"):
            self.required_perm_code = "employees.change_employeeaction"
            return [IsAuthenticated(), AdminOrActionOrModelPerms()]
        if self.action == "destroy":
            self.required_perm_code = "employees.delete_employeeaction"
            return [IsAuthenticated(), AdminOrActionOrModelPerms()]
        self.required_perm_code = None
        return super().get_permissions()

    # --- бизнес-эффекты после записи события ---
    def _apply_effects(self, action_obj: EmployeeAction):
        emp = action_obj.employee
        ldap_enabled = _is_ldap_enabled()
        
        if action_obj.action == ACTION_DISMISSED:
            # ВАЖНО: сначала деактивируем сотрудника (чтобы get_base_dn_for_employee работал правильно)
            if emp.is_active:
                emp.is_active = False
                emp.save(update_fields=["is_active"])
            
            # деактивируем связи с отделами в БД
            EmployeeDepartment.objects.filter(employee=emp, is_active=True).update(
                is_active=False, date_to=timezone.now().date()
            )
            
            # синхронизируем с LDAP (disable учётной записи)
            if ldap_enabled:
                try:
                    svc = DirectoryService()
                    # Проверяем, есть ли у пользователя ldap_dn
                    from employees.models import LdapSyncState
                    has_ldap_dn = LdapSyncState.objects.filter(
                        model='employee',
                        object_pk=str(emp.pk),
                        ldap_dn__isnull=False
                    ).exists()
                    
                    if has_ldap_dn:
                        # Обновляем is_active в LDAP (установит userAccountControl=514)
                        svc.update_user(emp, changes={"is_active": False})
                        
                        # Удаляем из всех отделов (это переместит в OU=Dismissed)
                        active_departments = EmployeeDepartment.objects.filter(
                            employee=emp
                        ).select_related('department')
                        
                        departments_processed = False
                        for emp_dept in active_departments:
                            try:
                                svc.remove_member(emp_dept.department, emp)
                                departments_processed = True
                            except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                                # Логируем ошибку удаления из конкретного отдела
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(
                                    f"Failed to remove dismissed employee from department: "
                                    f"employee_id={emp.id}, department_id={emp_dept.department.id}, error={e}"
                                )
                        
                        # Если сотрудник не состоял ни в одном отделе, вручную перемещаем в OU=Dismissed
                        if not departments_processed:
                            try:
                                from employees.ldap.utils.ldap_utils import get_base_dn_for_employee
                                from employees.ldap.infrastructure.connections import _ldap
                                from employees.ldap.repositories.ldap_repository import ensure_container_exists
                                from employees.models import LdapSyncState
                                
                                sync_state = LdapSyncState.objects.filter(
                                    model='employee',
                                    object_pk=str(emp.pk)
                                ).first()
                                
                                if sync_state and sync_state.ldap_dn:
                                    target_base = get_base_dn_for_employee(emp)
                                    current_dn = sync_state.ldap_dn
                                    
                                    # Проверяем, нужно ли перемещение
                                    if not current_dn.lower().endswith(target_base.lower()):
                                        with _ldap() as conn:
                                            ensure_container_exists(conn, target_base)
                                            new_dn = svc._user_service._move_user_to_base(
                                                conn, current_dn, target_base
                                            )
                                            sync_state.touch(ldap_dn=new_dn, sync_dir="ldap")
                                            import logging
                                            logger = logging.getLogger(__name__)
                                            logger.info(
                                                f"Dismissed employee without department moved to OU=Dismissed: "
                                                f"employee_id={emp.id}, new_dn={new_dn}"
                                            )
                            except Exception as e:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(
                                    f"Failed to move dismissed employee without department to OU=Dismissed: "
                                    f"employee_id={emp.id}, error={e}"
                                )
                except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                    # Логируем ошибку, но не прерываем процесс
                    # БД-изменения уже применены
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(
                        f"Failed to disable user in LDAP during dismissal: "
                        f"employee_id={emp.id}, error={e}"
                    )
        else:
            # любое иное событие делает сотрудника активным в БД
            was_inactive = not emp.is_active
            if was_inactive:
                emp.is_active = True
                emp.save(update_fields=["is_active"])
            
            # синхронизируем с LDAP (enable учётной записи)
            if ldap_enabled:
                try:
                    svc = DirectoryService()
                    from employees.models import LdapSyncState
                    has_ldap_dn = LdapSyncState.objects.filter(
                        model='employee',
                        object_pk=str(emp.pk),
                        ldap_dn__isnull=False
                    ).exists()
                    
                    if has_ldap_dn:
                        # Обновляем is_active в LDAP (установит userAccountControl=512)
                        svc.update_user(emp, changes={"is_active": True})
                        
                        # Если сотрудник был неактивен (восстановление из увольнения),
                        # перемещаем из OU=Dismissed в OU=Users
                        if was_inactive:
                            try:
                                sync_state = LdapSyncState.objects.get(
                                    model='employee',
                                    object_pk=str(emp.pk)
                                )
                                current_dn = sync_state.ldap_dn
                                dismissed_base = getattr(settings, "LDAP_DISMISSED_BASE", "")
                                
                                # Проверяем, находится ли сотрудник в OU=Dismissed
                                if dismissed_base and current_dn.lower().endswith(dismissed_base.lower()):
                                    users_base = getattr(settings, "LDAP_USERS_BASE", None) or getattr(
                                        settings, "LDAP_USER_BASE", None
                                    )
                                    if users_base:
                                        from employees.ldap.infrastructure.connections import _ldap
                                        from employees.ldap.repositories.ldap_repository import ensure_container_exists
                                        
                                        with _ldap() as conn:
                                            ensure_container_exists(conn, users_base)
                                            # Используем внутренний метод UserService
                                            new_dn = svc._user_service._move_user_to_base(
                                                conn, current_dn, users_base
                                            )
                                            sync_state.touch(ldap_dn=new_dn, sync_dir="ldap")
                                            import logging
                                            logger = logging.getLogger(__name__)
                                            logger.info(
                                                f"Restored employee moved from Dismissed to Users: "
                                                f"employee_id={emp.id}, new_dn={new_dn}"
                                            )
                            except Exception as e:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(
                                    f"Failed to move restored employee from Dismissed to Users: "
                                    f"employee_id={emp.id}, error={e}"
                                )
                except (DirectoryLdapError, DirectoryDbError, DirectoryServiceError) as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(
                        f"Failed to enable user in LDAP during action: "
                        f"employee_id={emp.id}, action={action_obj.action}, error={e}"
                    )

    @transaction.atomic
    def perform_create(self, serializer):
        obj = serializer.save()
        self._apply_effects(obj)

    @transaction.atomic
    def perform_update(self, serializer):
        obj = serializer.save()
        self._apply_effects(obj)


class GroupViewSet(viewsets.ModelViewSet):
    """CRUD и LDAP-операции с группами.

    Базовые маршруты:
        GET    /api/v1/groups/               — список (аутентифицированные)
        GET    /api/v1/groups/{id}/          — детальная
        POST   /api/v1/groups/               — создать (staff/superuser или с model perms)
        PATCH  /api/v1/groups/{id}/          — частичное обновление
        DELETE /api/v1/groups/{id}/          — удалить

    Экшены:
        GET    /api/v1/groups/{id}/permissions
        POST   /api/v1/groups/{id}/set-permissions
        POST   /api/v1/groups/{id}/add-permissions
        POST   /api/v1/groups/{id}/remove-permissions

        POST   /api/v1/groups/{id}/rename
        POST   /api/v1/groups/{id}/set-description
        GET    /api/v1/groups/{id}/members
        POST   /api/v1/groups/{id}/add-members
        POST   /api/v1/groups/{id}/remove-members
        POST   /api/v1/groups/{id}/replace-members
    """

    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    pagination_class = None

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name", "id"]
    ordering = ["name"]

    permission_classes = [IsAuthenticated, AdminOrActionOrModelPerms]
    required_perms_by_action = {
        "set_permissions": "employees.assign_group_permissions",
        "add_permissions": "employees.assign_group_permissions",
        "remove_permissions": "employees.assign_group_permissions",
        "rename": "employees.assign_group_permissions",
        "set_description": "employees.assign_group_permissions",
        "add_members": "employees.assign_group_permissions",
        "remove_members": "employees.assign_group_permissions",
        "replace_members": "employees.assign_group_permissions",
    }

    # ---------- queryset ----------

    def get_queryset(self):
        qs = super().get_queryset()

        member_raw = self.request.query_params.get(
            "member"
        ) or self.request.query_params.get("member_id")
        if member_raw is None:
            return qs

        try:
            member_id = int(str(member_raw).strip())
        except (TypeError, ValueError):
            return qs.none()  # некорректный id → пусто

        # У Group нет явного поля на пользователя; используется обратная связь ManyToMany от пользователя:
        # group.user_set (manager). Для ORM-фильтров используется related_query_name — в django.contrib.auth он "user".
        # На всякий случай поддержим оба варианта: user__ и user_set__ (если кастомная модель могла переопределить имя).
        try:
            return qs.filter(
                Q(user__id=member_id) | Q(user_set__id=member_id)
            ).distinct()
        except FieldError:
            # если вдруг один из путей не существует — пробуем максимально совместимый
            try:
                return qs.filter(user__id=member_id).distinct()
            except FieldError:
                try:
                    return qs.filter(user_set__id=member_id).distinct()
                except FieldError:
                    return qs.none()

    # ---------- helpers ----------

    def _validate_permissions_payload(
        self, request
    ) -> tuple[Optional[List[Permission]], Optional[Response]]:
        """Валидирует payload с ID permissions.

        Args:
            request (Request): DRF Request с полем 'permissions' (list[int]).

        Returns:
            tuple[Optional[List[Permission]], Optional[Response]]: (список Permission или None, Response с 400 или None)

        Raises:
            TypeError: Если элементы 'permissions' не приводимы к int.
        """
        ids = request.data.get("permissions")
        if not isinstance(ids, list):
            return None, Response(
                {"detail": "Поле 'permissions' должно быть списком id"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            pid_list = [int(x) for x in ids]
        except (TypeError, ValueError):
            return None, Response(
                {"detail": "Список 'permissions' должен содержать целые числа"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        qs = Permission.objects.filter(id__in=pid_list)
        if qs.count() != len(set(pid_list)):
            return None, Response(
                {"detail": "Некоторые permissions не найдены"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return list(qs), None

    def _resolve_group_dn(self, grp: Group) -> Optional[str]:
        """Определяет DN группы по её CN.

        Сначала пробует DirectoryService.group_find_dn, затем (fallback) find_group_dn.

        Args:
            grp (Group): Экземпляр Group.

        Returns:
            Optional[str]: DN группы или None, если не найдена или LDAP отключен.
        """
        if not _is_ldap_enabled():
            return None
        
        base = getattr(settings, "LDAP_GROUPS_BASE", "") or None
        svc = DirectoryService()
        try:
            return svc.group_find_dn(grp.name, bases=[base] if base else None)
        except AttributeError:
            # совместимость со старым API
            return svc.find_group_dn(grp.name, bases=[base] if base else None)

    def _members_payload_to_dns(self, payload: dict[str, Any]) -> list[str]:
        """Извлекает список DN участников из payload (member_dns|member_ids).

        Args:
            payload (dict[str, Any]): Dict с ключами 'member_dns' (list[str]) и/или 'member_ids' (list[int]).

        Returns:
            list[str]: Уникальные DN (пустой список если LDAP отключен).

        Raises:
            ValueError: Если не удалось получить ни одного DN (только в LDAP режиме).
        """
        dns: list[str] = []
        raw_dns = payload.get("member_dns") or []
        if isinstance(raw_dns, list):
            dns.extend([d.strip() for d in raw_dns if isinstance(d, str) and d.strip()])

        ids = payload.get("member_ids") or []
        if isinstance(ids, list) and ids:
            if _is_ldap_enabled():
                svc = DirectoryService()
                dns.extend(svc.employee_ids_to_dns([i for i in ids if isinstance(i, int)]))

        uniq, seen = [], set()
        for d in dns:
            if d and d not in seen:
                uniq.append(d)
                seen.add(d)
        
        if not _is_ldap_enabled():
            return []  # В non-LDAP режиме возвращаем пустой список
            
        if not uniq:
            raise ValueError("Не переданы корректные member_dns или member_ids")
        return uniq

    def _dns_to_users(self, dns):
        """Маппит DN участников на локальных пользователей через LdapSyncState.

        Args:
            dns (list[str]): Список DN.

        Returns:
            list[Employee]: Пользователи, найденные по LdapSyncState(model='employee', ldap_dn IN dns).
        """
        if not dns:
            return []

        sub = LdapSyncState.objects.filter(
            model="employee", object_pk=OuterRef("pk")
        ).values_list("ldap_dn", flat=True)[:1]

        return list(
            Employee.objects.annotate(ldap_dn=Subquery(sub))
            .filter(ldap_dn__in=dns)
            .only("id")
        )

    # ---------- override CRUD: LDAP → DB ----------

    def create(self, request, *args, **kwargs) -> Response:
        """Создаёт LDAP-группу, затем запись Group в БД и (опционально) назначает permissions.

        Body (частично):
            name (str): CN группы (обязателен)
            ldap_parent_dn (str): контейнер, по умолчанию settings.LDAP_GROUPS_BASE
            ldap_description (str|None)
            ldap_scope (str): 'global'|... (по умолчанию 'global')
            ldap_security (bool): безопасная группа (по умолчанию True)
            permissions (list[int]): список id Django permissions (необязателен)

        Returns:
            Response: 201 с сериализованной группой или ошибка 4xx/5xx.

        Raises:
            ValidationError: При ошибке сериализатора.
            Exception: Пробрасывается как 502/500 в зависимости от источника ошибки.
        """
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        name: str = ser.validated_data["name"]

        parent_dn = request.data.get("ldap_parent_dn") or getattr(
            settings, "LDAP_GROUPS_BASE", None
        )
        description = request.data.get("ldap_description")
        scope = request.data.get("ldap_scope", "global")
        security_enabled = bool(request.data.get("ldap_security", True))

        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # LDAP mode: create group in LDAP first
            svc = DirectoryService()
            try:
                svc.group_create(
                    cn=name,
                    parent_dn=parent_dn,
                    description=description,
                    scope=scope,
                    security_enabled=security_enabled,
                )
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

        # Create DB record
        try:
            grp = Group.objects.create(name=name)
            perms = ser.validated_data.get("permissions")
            if perms:
                grp.permissions.set(perms)
        except Exception as e:
            if ldap_enabled:
                # откат LDAP-создания
                try:
                    svc = DirectoryService()
                    dn = None
                    try:
                        dn = svc.group_find_dn(
                            name, bases=[parent_dn] if parent_dn else None
                        )
                    except AttributeError:
                        dn = svc.find_group_dn(
                            name, bases=[parent_dn] if parent_dn else None
                        )
                    if dn:
                        svc.group_delete(dn)
                except Exception:
                    pass
            return Response(
                {"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        out = self.get_serializer(grp)
        return Response(
            out.data,
            status=status.HTTP_201_CREATED,
            headers=self.get_success_headers(out.data),
        )

    def partial_update(self, request, *args, **kwargs) -> Response:
        """Частичное обновление: сначала LDAP (rename/description), затем БД.

        Body:
            name (str): новое имя (опционально)
            ldap_description (str|None): описание (опционально, "__NO_CHANGE__" — без изменений)
        """
        grp = self.get_object()
        new_name = request.data.get("name")
        new_desc = request.data.get("ldap_description", "__NO_CHANGE__")
        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # LDAP mode: sync name/description with LDAP
            dn = self._resolve_group_dn(grp)
            if not dn:
                return Response(
                    {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
                )

            svc = DirectoryService()
            try:
                if new_name and new_name != grp.name:
                    dn = svc.group_rename(dn, new_name)
                if new_desc != "__NO_CHANGE__":
                    svc.group_set_description(dn, (new_desc or None))
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

        # Update DB
        if new_name and new_name != grp.name:
            grp.name = new_name
            grp.save(update_fields=["name"])

        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs) -> Response:
        """Удаляет LDAP-группу (если найдена), затем запись Group в БД.

        Query params:
            force_db (bool): При ошибке удаления в LDAP всё равно удалить запись БД.
        """
        grp = self.get_object()
        force_db = str(request.query_params.get("force_db", "")).lower() in {
            "1",
            "true",
            "yes",
        }
        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            dn = self._resolve_group_dn(grp)
            if dn:
                try:
                    DirectoryService().group_delete(dn)
                except Exception as e:
                    if not force_db:
                        return Response(
                            {"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY
                        )
                    # мягко логируем и продолжаем удаление из БД
        
        return super().destroy(request, *args, **kwargs)

    def list(self, request, *args, **kwargs) -> Response:
        """Возвращает список групп. Перед выдачей — мягкий LDAP-синк каталога."""
        if _is_ldap_enabled():
            try:
                DirectoryService().sync_groups_catalog(throttle_seconds=60)
            except Exception:
                pass
        return super().list(request, *args, **kwargs)

    # ---------- Actions: Django permissions ----------

    @action(detail=True, methods=["get"])
    def permissions(self, request, pk=None) -> Response:
        """Возвращает permissions, привязанные к группе (через M2M).

        Returns:
            Response: {"count": int, "results": [{id, codename, name, app, model}, ...]}
        """
        grp = self.get_object()
        perms = grp.permissions.select_related("content_type").distinct()
        data = [
            {
                "id": p.id,
                "codename": f"{p.content_type.app_label}.{p.codename}",
                "name": p.name,
                "app": p.content_type.app_label,
                "model": p.content_type.model,
            }
            for p in perms
        ]
        return Response(
            {"count": len(data), "results": data}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="set-permissions")
    def set_permissions(self, request, pk=None) -> Response:
        """Полностью заменяет набор permissions у группы.

        Body:
            permissions (list[int]): список id разрешений.
        """
        grp = self.get_object()
        qs, error = self._validate_permissions_payload(request)
        if error:
            return error
        grp.permissions.set(qs)
        return Response(
            {
                "ok": True,
                "permission_ids": list(grp.permissions.values_list("id", flat=True)),
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="add-permissions")
    def add_permissions(self, request, pk=None) -> Response:
        """Добавляет permissions к группе (без удаления существующих).

        Body:
            permissions (list[int]): список id разрешений.
        """
        grp = self.get_object()
        qs, error = self._validate_permissions_payload(request)
        if error:
            return error
        grp.permissions.add(*qs)
        return Response(
            {
                "ok": True,
                "permission_ids": list(grp.permissions.values_list("id", flat=True)),
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="remove-permissions")
    def remove_permissions(self, request, pk=None) -> Response:
        """Удаляет указанные permissions у группы.

        Body:
            permissions (list[int]): список id разрешений.
        """
        grp = self.get_object()
        qs, error = self._validate_permissions_payload(request)
        if error:
            return error
        grp.permissions.remove(*qs)
        return Response(
            {
                "ok": True,
                "permission_ids": list(grp.permissions.values_list("id", flat=True)),
            },
            status=status.HTTP_200_OK,
        )

    # ---------- Actions: LDAP ----------

    @action(detail=True, methods=["post"])
    def rename(self, request, pk=None) -> Response:
        """Переименовывает LDAP-группу и синхронизирует имя в БД.

        Body:
            new_name (str): Новое имя (CN).

        Returns:
            Response: {"ok": True, "name": str}
        """
        grp = self.get_object()
        new_name = (request.data.get("new_name") or "").strip()
        if not new_name:
            return Response(
                {"detail": "new_name обязателен"}, status=status.HTTP_400_BAD_REQUEST
            )

        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            dn = self._resolve_group_dn(grp)
            if not dn:
                return Response(
                    {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
                )

            svc = DirectoryService()
            try:
                svc.group_rename(dn, new_name)
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)
        
        # Update DB
        grp.name = new_name
        grp.save(update_fields=["name"])
        return Response({"ok": True, "name": grp.name}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="set-description")
    def set_description(self, request, pk=None) -> Response:
        """Устанавливает описание LDAP-группы.

        Body:
            description (str|None): Описание группы.

        Returns:
            Response: {"ok": True}
        """
        grp = self.get_object()
        
        if not _is_ldap_enabled():
            # Non-LDAP mode: description stored only in LDAP, so just return OK
            return Response({"ok": True}, status=status.HTTP_200_OK)
        
        dn = self._resolve_group_dn(grp)
        if not dn:
            return Response(
                {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
            )

        svc = DirectoryService()
        try:
            svc.group_set_description(dn, request.data.get("description"))
            return Response({"ok": True}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

    @action(detail=True, methods=["get"])
    def members(self, request, pk=None) -> Response:
        """Возвращает состав LDAP-группы (DN) и краткую информацию о сопоставленных сотрудниках.

        Returns:
            Response: {"dns": list[str], "employees": list[dict]}
        """
        grp = self.get_object()
        
        if not _is_ldap_enabled():
            # Non-LDAP mode: return DB members only
            users = grp.user_set.all()
            employees = [
                {
                    "id": u.id,
                    "email": u.email,
                    "first_name": u.first_name,
                    "last_name": u.last_name,
                }
                for u in users
            ]
            return Response({"dns": [], "employees": employees}, status=status.HTTP_200_OK)
        
        dn = self._resolve_group_dn(grp)
        if not dn:
            return Response(
                {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
            )

        svc = DirectoryService()
        try:
            dns = svc.group_list_members(dn)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

        employees = svc.employees_brief_by_dns(dns)
        return Response({"dns": dns, "employees": employees}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="add-members")
    def add_members(self, request, pk=None) -> Response:
        """Добавляет участников в LDAP-группу и связывает найденных пользователей в БД.

        Body:
            member_dns (list[str])|None: Прямые DN участников.
            member_ids (list[int])|None: ID сотрудников, которые будут преобразованы в DN.

        Returns:
            Response: {"ok": True, "ldap_added": int, "db_added": int, "ok_dns": list[str], "ok_user_ids": list[int]}

        Raises:
            400: Некорректный payload.
            404: Группа не найдена в LDAP.
            502: Ошибка LDAP.
            500: Ошибка БД (с попыткой компенсировать изменения в LDAP).
        """
        grp = self.get_object()
        ldap_enabled = _is_ldap_enabled()
        
        if ldap_enabled:
            # LDAP mode: sync with LDAP first
            dn = self._resolve_group_dn(grp)
            if not dn:
                return Response(
                    {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
                )
            try:
                member_dns = self._members_payload_to_dns(request.data)
            except ValueError as e:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            svc = DirectoryService()
            try:
                svc.group_add_members(dn, member_dns)
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

            users = self._dns_to_users(member_dns)
            ok_user_ids = [u.id for u in users]
            try:
                with transaction.atomic():
                    grp.user_set.add(*users)
            except Exception as e:
                # компенсируем LDAP
                try:
                    svc.group_remove_members(dn, member_dns)
                except Exception:
                    pass
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            return Response(
                {
                    "ok": True,
                    "ldap_added": len(member_dns),
                    "db_added": len(users),
                    "ok_dns": member_dns,
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )
        else:
            # Non-LDAP mode: add members directly to DB
            member_ids = request.data.get("member_ids") or []
            if not isinstance(member_ids, list):
                return Response(
                    {"detail": "member_ids must be a list"}, status=status.HTTP_400_BAD_REQUEST
                )
            
            users = Employee.objects.filter(id__in=member_ids)
            try:
                with transaction.atomic():
                    grp.user_set.add(*users)
            except Exception as e:
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            
            ok_user_ids = [u.id for u in users]
            return Response(
                {
                    "ok": True,
                    "ldap_added": 0,
                    "db_added": len(users),
                    "ok_dns": [],
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )

    @action(detail=True, methods=["post"], url_path="remove-members")
    def remove_members(self, request, pk=None) -> Response:
        """Удаляет участников из LDAP-группы и разрывает связи в БД.

        Body:
            member_dns (list[str])|None
            member_ids (list[int])|None

        Returns:
            Response: {"ok": True, "ldap_removed": int, "db_removed": int, "ok_dns": list[str], "ok_user_ids": list[int]}

        Errors:
            400/404/502/500 — см. add_members.
        """
        grp = self.get_object()
        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # LDAP mode: sync with LDAP first
            dn = self._resolve_group_dn(grp)
            if not dn:
                return Response(
                    {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
                )
            try:
                member_dns = self._members_payload_to_dns(request.data)
            except ValueError as e:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            svc = DirectoryService()
            try:
                svc.group_remove_members(dn, member_dns)
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

            users = self._dns_to_users(member_dns)
            ok_user_ids = [u.id for u in users]
            try:
                with transaction.atomic():
                    grp.user_set.remove(*users)
            except Exception as e:
                # компенсируем LDAP
                try:
                    svc.group_add_members(dn, member_dns)
                except Exception:
                    pass
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            return Response(
                {
                    "ok": True,
                    "ldap_removed": len(member_dns),
                    "db_removed": len(users),
                    "ok_dns": member_dns,
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )
        else:
            # Non-LDAP mode: remove members directly from DB
            member_ids = request.data.get("member_ids") or []
            if not isinstance(member_ids, list):
                return Response(
                    {"detail": "member_ids must be a list"}, status=status.HTTP_400_BAD_REQUEST
                )
            
            users = Employee.objects.filter(id__in=member_ids)
            try:
                with transaction.atomic():
                    grp.user_set.remove(*users)
            except Exception as e:
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            
            ok_user_ids = [u.id for u in users]
            return Response(
                {
                    "ok": True,
                    "ldap_removed": 0,
                    "db_removed": len(users),
                    "ok_dns": [],
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )

    @action(detail=True, methods=["post"], url_path="replace-members")
    def replace_members(self, request, pk=None) -> Response:
        """Полностью заменяет состав LDAP-группы и синхронизирует M2M в БД.

        Body:
            member_dns (list[str])|None
            member_ids (list[int])|None

        Returns:
            Response: {"ok": True, "ldap_total": int, "db_total": int, "ok_dns": list[str], "ok_user_ids": list[int]}

        Notes:
            При ошибке БД выполняется попытка отката LDAP к исходному составу.
        """
        grp = self.get_object()
        ldap_enabled = _is_ldap_enabled()

        if ldap_enabled:
            # LDAP mode: sync with LDAP first
            dn = self._resolve_group_dn(grp)
            if not dn:
                return Response(
                    {"detail": "Группа не найдена в LDAP"}, status=status.HTTP_404_NOT_FOUND
                )
            try:
                desired_dns = self._members_payload_to_dns(request.data)
            except ValueError as e:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            svc = DirectoryService()
            # сохраняем исходный состав для возможного отката
            try:
                prev_dns = svc.group_list_members(dn)
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

            try:
                svc.group_replace_members(dn, desired_dns)
            except Exception as e:
                return Response({"detail": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

            users = self._dns_to_users(desired_dns)
            ok_user_ids = [u.id for u in users]
            try:
                with transaction.atomic():
                    grp.user_set.set(users)
            except Exception as e:
                # откат LDAP
                try:
                    svc.group_replace_members(dn, prev_dns)
                except Exception:
                    pass
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            return Response(
                {
                    "ok": True,
                    "ldap_total": len(desired_dns),
                    "db_total": len(users),
                    "ok_dns": desired_dns,
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )
        else:
            # Non-LDAP mode: replace members directly in DB
            member_ids = request.data.get("member_ids") or []
            if not isinstance(member_ids, list):
                return Response(
                    {"detail": "member_ids must be a list"}, status=status.HTTP_400_BAD_REQUEST
                )
            
            users = Employee.objects.filter(id__in=member_ids)
            try:
                with transaction.atomic():
                    grp.user_set.set(users)
            except Exception as e:
                return Response(
                    {"detail": f"DB error: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            
            ok_user_ids = [u.id for u in users]
            return Response(
                {
                    "ok": True,
                    "ldap_total": 0,
                    "db_total": len(users),
                    "ok_dns": [],
                    "ok_user_ids": ok_user_ids,
                },
                status=status.HTTP_200_OK,
            )

        return Response(
            {
                "ok": True,
                "ldap_total": len(desired_dns),
                "db_total": len(users),
                "ok_dns": desired_dns,
                "ok_user_ids": ok_user_ids,
            },
            status=status.HTTP_200_OK,
        )
